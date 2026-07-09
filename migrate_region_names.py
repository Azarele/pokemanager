"""
migrate_region_names.py — one-time script to append (Korean) / (Japanese)
suffixes to Card_Name for existing KR/JP inventory items.

Safe to run multiple times — strips any existing suffix before reapplying the
correct one, so double-suffixes and wrong-suffix cases are fixed too.
Run with bot stopped: python3 migrate_region_names.py
"""

import sys
from pathlib import Path

import openpyxl

INVENTORY_FILE = Path(__file__).parent / "inventory.xlsx"

# Column indices (0-based) matching excel_db.COLUMNS order
COL_ITEM_ID   = 0
COL_CARD_NAME = 1
COL_REGION    = 4
COL_STATUS    = 10


def _fix_name(card_name: str, region: str) -> tuple[str, bool]:
    """Return (corrected_name, was_changed)."""
    name = card_name or ""
    # Strip any existing region suffix
    for suffix in [" (Korean)", " (Japanese)"]:
        if name.endswith(suffix):
            name = name[: -len(suffix)]

    if region == "KR":
        correct = f"{name} (Korean)"
    elif region == "JP":
        correct = f"{name} (Japanese)"
    else:
        correct = name

    return correct, correct != (card_name or "")


def main() -> None:
    if not INVENTORY_FILE.exists():
        print(f"ERROR: {INVENTORY_FILE} not found.")
        sys.exit(1)

    wb = openpyxl.load_workbook(INVENTORY_FILE)
    ws = wb.active

    # Resolve column positions from header row (handles schema migrations)
    headers = {cell.value: cell.column - 1 for cell in ws[1]}
    col_item_id   = headers.get("Item_ID",   COL_ITEM_ID)
    col_card_name = headers.get("Card_Name", COL_CARD_NAME)
    col_region    = headers.get("Region",    COL_REGION)
    col_status    = headers.get("Status",    COL_STATUS)

    renamed = []
    skipped = []
    unchanged = []

    data_rows = list(ws.iter_rows(min_row=2))
    print(f"Scanning inventory ({len(data_rows)} rows)...\n")

    for row in data_rows:
        card_name = row[col_card_name].value or ""
        region    = row[col_region].value or ""
        item_id   = row[col_item_id].value

        if region not in ("KR", "JP"):
            unchanged.append(item_id)
            continue

        new_name, changed = _fix_name(card_name, region)

        if changed:
            note = ""
            for sfx in [" (Korean)", " (Japanese)"]:
                if card_name.endswith(sfx) and sfx != f" ({'Korean' if region == 'KR' else 'Japanese'})":
                    note = "  [wrong suffix replaced]"
                    break
            print(f"  Item {item_id}: {card_name!r}")
            print(f"           → {new_name!r}{note}")
            renamed.append((row[col_card_name], new_name))
        else:
            print(f"  Item {item_id}: already correct — skipped")
            skipped.append(item_id)

    print()
    print(
        f"Total: {len(renamed)} renamed, {len(skipped)} skipped (already correct), "
        f"{len(unchanged)} unchanged (English/no region)"
    )

    if not renamed:
        print("Nothing to change.")
        return

    print()
    answer = input("Proceed with these changes? [y/N]: ").strip().lower()
    if answer != "y":
        print("Aborted — no changes saved.")
        return

    for cell, new_name in renamed:
        cell.value = new_name

    wb.save(INVENTORY_FILE)
    print(f"Saved {INVENTORY_FILE}. Done.")


if __name__ == "__main__":
    main()
