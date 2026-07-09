"""
One-off backfill: finds Price_History rows present in inventory.xlsx but
missing from Supabase (gap left by an earlier partial run of
migrate_price_history_watchlist.py) and inserts only those rows.

Dedup key: (item_id, recorded_at string, live_price_gbp) — the original
migration writes the Excel timestamp string through verbatim with no
timezone conversion, so historical rows match digit-for-digit against
Excel's Timestamp column.

Run once: python3 backfill_price_history_gap.py
"""
import openpyxl
from web.database import get_db

USER_ID = "fa6d2b09-e53a-4245-bec7-6995e5be8307"


def fetch_all_supabase_rows(db):
    rows = []
    page = 1000
    offset = 0
    while True:
        resp = db.table("price_history").select("item_id, recorded_at, live_price_gbp") \
            .eq("user_id", USER_ID).range(offset, offset + page - 1).execute()
        batch = resp.data
        rows.extend(batch)
        if len(batch) < page:
            break
        offset += page
    return rows


def main():
    wb = openpyxl.load_workbook("inventory.xlsx", data_only=True)
    ws = wb["Price_History"]
    headers = [c.value for c in ws[1]]
    idx_item_id = headers.index("Item_ID")
    idx_price = headers.index("Live_Price_GBP")
    idx_timestamp = headers.index("Timestamp")

    excel_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item_id = row[idx_item_id]
        price = row[idx_price]
        timestamp = row[idx_timestamp]
        if item_id is None or price is None or not timestamp:
            continue
        excel_rows.append((int(item_id), str(timestamp), float(price)))

    db = get_db()
    supabase_rows = fetch_all_supabase_rows(db)
    print(f"Excel rows: {len(excel_rows)}, Supabase rows: {len(supabase_rows)}")

    existing = set()
    for r in supabase_rows:
        ts = r["recorded_at"]
        # Normalise "2026-07-01 21:05:05+00" -> "2026-07-01T21:05:05" for comparison
        ts_key = ts.replace(" ", "T").split("+")[0].split(".")[0]
        existing.add((r["item_id"], ts_key, round(float(r["live_price_gbp"]), 2)))

    missing = [
        {"user_id": USER_ID, "item_id": item_id, "live_price_gbp": price, "recorded_at": ts}
        for item_id, ts, price in excel_rows
        if (item_id, ts, round(price, 2)) not in existing
    ]

    print(f"Missing rows to insert: {len(missing)}")
    for i in range(0, len(missing), 100):
        batch = missing[i:i + 100]
        db.table("price_history").insert(batch).execute()
        print(f"Inserted {i + len(batch)}/{len(missing)}")

    print("Done.")


if __name__ == "__main__":
    main()
