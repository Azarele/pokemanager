"""
One-off migration: copies Price_History and Watchlist sheets from
inventory.xlsx into the Supabase price_history / watchlist tables.

Run once: python3 migrate_price_history_watchlist.py
"""
import openpyxl
from web.database import get_db

USER_ID = "fa6d2b09-e53a-4245-bec7-6995e5be8307"


def migrate_price_history(wb):
    if "Price_History" not in wb.sheetnames:
        print("No Price_History sheet found - skipping")
        return

    ws = wb["Price_History"]
    headers = [c.value for c in ws[1]]
    print(f"Price_History headers: {headers}")

    idx_item_id = headers.index("Item_ID")
    idx_price = headers.index("Live_Price_GBP")
    idx_timestamp = headers.index("Timestamp")

    db = get_db()
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item_id = row[idx_item_id]
        price = row[idx_price]
        timestamp = row[idx_timestamp]

        if item_id is None or price is None:
            continue

        rows.append({
            "user_id": USER_ID,
            "item_id": int(item_id),
            "live_price_gbp": float(price),
            "recorded_at": str(timestamp) if timestamp else None,
        })

    total = len(rows)
    for i in range(0, total, 100):
        batch = rows[i:i + 100]
        db.table("price_history").insert(batch).execute()
        print(f"Inserted {i + len(batch)}/{total} price history rows")

    print(f"Done: {total} price history rows migrated")


def migrate_watchlist(wb):
    if "Watchlist" not in wb.sheetnames:
        print("No Watchlist sheet found - skipping")
        return

    ws = wb["Watchlist"]
    headers = [c.value for c in ws[1]]
    print(f"Watchlist headers: {headers}")

    idx_card_name = headers.index("Card_Name")
    idx_pc_url = headers.index("PC_URL")
    idx_target = headers.index("Target_Price_GBP")
    idx_current = headers.index("Current_Price_GBP")
    idx_added = headers.index("Added_Date") if "Added_Date" in headers else -1
    idx_alert = headers.index("Alert_Sent") if "Alert_Sent" in headers else -1

    db = get_db()
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[idx_card_name]:
            continue
        entry = {
            "user_id": USER_ID,
            "card_name": str(row[idx_card_name] or ""),
            "pc_url": str(row[idx_pc_url] or ""),
            "target_price_gbp": float(row[idx_target] or 0),
            "current_price_gbp": float(row[idx_current] or 0),
        }
        if idx_added >= 0 and row[idx_added]:
            entry["added_date"] = str(row[idx_added])[:10]
        if idx_alert >= 0 and row[idx_alert] is not None:
            entry["alert_sent"] = bool(row[idx_alert])
        rows.append(entry)

    if rows:
        db.table("watchlist").insert(rows).execute()
        print(f"Migrated {len(rows)} watchlist entries")
    else:
        print("No watchlist entries found")


def main():
    wb = openpyxl.load_workbook("inventory.xlsx", data_only=True)
    print(f"Sheets found: {wb.sheetnames}")
    migrate_price_history(wb)
    migrate_watchlist(wb)


if __name__ == "__main__":
    main()
