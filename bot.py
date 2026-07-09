"""
Discord trading-inventory bot — Slash Commands.

Commands
--------
/add  <pc_url> <condition> <purchase_price>
      Scrape PriceCharting for the card name and live price, save to inventory.

/sell <item_ids> <sell_prices>
      Mark one or more items sold and record profit. Accepts comma-separated IDs and prices.

/list <item_id> <image1> [image2-5] [price_override]
      Cross-list on eBay UK and Vinted UK.

/stock
      Show all cards currently in inventory.

/summary
      Aggregate profit/loss across all sold items.
"""

import asyncio
import calendar
import json
import random
import shutil
import traceback
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Optional

import discord
from discord import app_commands
from discord.ext import commands, tasks

import ai_helper
import audit
import backups
import config
import excel_db
import lister_ebay_api as lister_ebay
import lister_vinted
import scraper

try:
    from web import bot_sync
    _SYNC_ENABLED = True
except Exception as _sync_exc:
    _SYNC_ENABLED = False
    print(f"[bot] Supabase sync disabled: {_sync_exc}")

# ---------------------------------------------------------------------------
# Bot setup
# ---------------------------------------------------------------------------

intents = discord.Intents.default()
intents.message_content = True

bot = commands.Bot(command_prefix=config.COMMAND_PREFIX, intents=intents)

STATUS_INVENTORY = "Inventory"

# ---------------------------------------------------------------------------
# eBay sales deduplication — persists across restarts via a flat text file
# ---------------------------------------------------------------------------

_PROCESSED_ORDERS_FILE = Path(config.AUDIT_LOG_DIR) / "processed_orders.txt"
_processed_order_ids: set[str] = set()

# ---------------------------------------------------------------------------
# Bot health stats — updated by background loops and scraper callbacks
# ---------------------------------------------------------------------------

_bot_start_time: datetime                    = datetime.now(timezone.utc)
_last_price_update: Optional[datetime]       = None
_last_ebay_sales_check: Optional[datetime]   = None
_price_update_stats: dict                    = {"checked": 0, "updated": 0, "pc_429s": 0, "pc_success": 0}
_watchdog_alerts_sent: set[str]              = set()


def _fmt_delta(td: timedelta) -> str:
    """Format a timedelta as '3d 14h 22m' (always includes minutes)."""
    total = int(abs(td.total_seconds()))
    d = total // 86400
    h = (total % 86400) // 3600
    m = (total % 3600) // 60
    parts = []
    if d:
        parts.append(f"{d}d")
    if h:
        parts.append(f"{h}h")
    parts.append(f"{m}m")
    return " ".join(parts)


def _load_processed_order_ids() -> None:
    if _PROCESSED_ORDERS_FILE.exists():
        for line in _PROCESSED_ORDERS_FILE.read_text().splitlines():
            oid = line.strip()
            if oid:
                _processed_order_ids.add(oid)
    print(f"[ebay_sales] Loaded {len(_processed_order_ids)} previously processed order ID(s).")


def _persist_processed_order_id(order_id: str) -> None:
    try:
        _PROCESSED_ORDERS_FILE.parent.mkdir(exist_ok=True)
        with _PROCESSED_ORDERS_FILE.open("a") as f:
            f.write(order_id + "\n")
    except Exception as exc:
        print(f"[ebay_sales] Warning: could not persist order ID {order_id}: {exc}")


async def _maybe_create_monthly_sheet() -> None:
    """Create the previous month's summary sheet if it doesn't already exist."""
    today = date.today()
    if today.month == 1:
        prev_year, prev_month = today.year - 1, 12
    else:
        prev_year, prev_month = today.year, today.month - 1
    created = await excel_db.ensure_monthly_sheet_async(prev_year, prev_month)
    if created:
        print(f"[bot] Created monthly summary sheet: {prev_year}-{prev_month:02d}")

# Condition choices presented in the /add dropdown
_REGION_CHOICES = [
    app_commands.Choice(name="JP — Japanese print", value="JP"),
    app_commands.Choice(name="KR — Korean print (70% of JP price)", value="KR"),
]

# Note: PSA 1-5 removed from dropdown but still valid values in existing inventory.
# BGS 8.5 and CGC 8.5 omitted to stay within Discord's 25-choice limit; still
# supported in scraper/eBay mappings for manually-entered values.
_CONDITION_CHOICES = [
    # Ungraded
    app_commands.Choice(name="Near mint or better", value="Near mint or better"),
    app_commands.Choice(name="Lightly played",       value="Lightly played"),
    app_commands.Choice(name="Moderately played",    value="Moderately played"),
    app_commands.Choice(name="Heavily played",       value="Heavily played"),
    # PSA (Professional Sports Authenticator)
    app_commands.Choice(name="PSA 10",  value="PSA 10"),
    app_commands.Choice(name="PSA 9",   value="PSA 9"),
    app_commands.Choice(name="PSA 8",   value="PSA 8"),
    app_commands.Choice(name="PSA 7",   value="PSA 7"),
    app_commands.Choice(name="PSA 6",   value="PSA 6"),
    # BGS (Beckett Grading Services)
    app_commands.Choice(name="BGS 10 Black Label", value="BGS 10"),
    app_commands.Choice(name="BGS 9.5",            value="BGS 9.5"),
    app_commands.Choice(name="BGS 9",              value="BGS 9"),
    app_commands.Choice(name="BGS 8",              value="BGS 8"),
    # CGC (Certified Guaranty Company)
    app_commands.Choice(name="CGC 10",             value="CGC 10"),
    app_commands.Choice(name="CGC 9.5",            value="CGC 9.5"),
    app_commands.Choice(name="CGC 9",              value="CGC 9"),
    app_commands.Choice(name="CGC 8",              value="CGC 8"),
    # SGC (Sportscard Guaranty Corporation)
    app_commands.Choice(name="SGC 10",             value="SGC 10"),
    app_commands.Choice(name="SGC 9.5",            value="SGC 9.5"),
    app_commands.Choice(name="SGC 9",              value="SGC 9"),
    app_commands.Choice(name="SGC 8",              value="SGC 8"),
    # ACE Grading (UK-based)
    app_commands.Choice(name="ACE 10",             value="ACE 10"),
    app_commands.Choice(name="ACE 9.5",            value="ACE 9.5"),
    # GetGraded (UK-based) — ACE 9 and ACE 8 dropped to stay within 25-choice limit
    app_commands.Choice(name="GetGraded 10",       value="GetGraded 10"),
    app_commands.Choice(name="GetGraded 9.5",      value="GetGraded 9.5"),
]


# ---------------------------------------------------------------------------
# Lifecycle
# ---------------------------------------------------------------------------

@bot.event
async def on_ready():
    print(f"[bot] Logged in as {bot.user} (id={bot.user.id})")
    await excel_db.init_excel_async()
    print("[bot] Inventory file ready.")
    await _maybe_create_monthly_sheet()
    try:
        if config.TEST_GUILD_ID:
            test_guild = discord.Object(id=config.TEST_GUILD_ID)
            # Push the full command set to the test guild (instant propagation).
            bot.tree.copy_global_to(guild=test_guild)
            synced = await bot.tree.sync(guild=test_guild)
            print(f"[bot] Synced {len(synced)} slash command(s) to guild {config.TEST_GUILD_ID}.")
            # Wipe any leftover global registrations from previous runs so commands
            # don't appear twice in the picker.  clear_commands() empties the
            # in-memory global list, so we snapshot first and restore afterward —
            # this only pushes an empty payload to Discord's global endpoint;
            # the local tree remains intact for any future code that inspects it.
            global_cmds = bot.tree.get_commands(guild=None)
            bot.tree.clear_commands(guild=None)
            await bot.tree.sync(guild=None)
            for cmd in global_cmds:
                bot.tree.add_command(cmd)
            print("[bot] Cleared global commands (per-guild sync active; duplicates removed).")
        else:
            # Global sync: works on any server but takes up to 1 hour to propagate.
            synced = await bot.tree.sync()
            print(f"[bot] Synced {len(synced)} slash command(s) globally.")
    except Exception as exc:
        print(f"[bot] Tree sync failed: {exc}")
    _load_processed_order_ids()
    if not Path(config.VINTED_STATE_PATH).exists():
        print("[bot] Vinted session not found — /listvinted will fail until you run generate_cookies.py vinted")
    else:
        print("[bot] Vinted session found — /listvinted available")
    def _on_pc_429_stats():
        _price_update_stats["pc_429s"] += 1

    def _on_pc_success_stats():
        _price_update_stats["pc_success"] += 1

    scraper.register_callbacks(on_429=_on_pc_429_stats, on_success=_on_pc_success_stats)
    await _check_excel_integrity()

    if not price_update_loop.is_running():
        price_update_loop.start()
        print(f"[bot] Price update loop started (every {config.UPDATE_INTERVAL_HOURS}h).")
    if not check_ebay_sales_loop.is_running():
        check_ebay_sales_loop.start()
        print("[bot] eBay sales check loop started (every 30 min).")
    if not watchdog_loop.is_running():
        watchdog_loop.start()
        print("[bot] Watchdog loop started (every 15 min).")


@bot.tree.error
async def on_app_command_error(
    interaction: discord.Interaction,
    error: app_commands.AppCommandError,
) -> None:
    msg = f"Command error: {error}"
    try:
        if interaction.response.is_done():
            await interaction.followup.send(msg, ephemeral=True)
        else:
            await interaction.response.send_message(msg, ephemeral=True)
    except discord.errors.HTTPException:
        print(f"[bot] Could not send error to expired interaction: {msg}")


# ---------------------------------------------------------------------------
# Background price updater — private helpers
# ---------------------------------------------------------------------------

_CIRCUIT_BREAKER_THRESHOLD = 5


class NoPriceAvailable(ValueError):
    """Raised when a card's PriceCharting page exists but no price is listed.

    This is not a system failure — it is a normal outcome for cards with sparse
    data (e.g. uncommon cards, pre-release pages). The circuit breaker must NOT
    count these; only genuine scraper/network errors should trip it.
    """


async def _scrape_fresh_price(
    pc_url: str,
    card_name: str,
    condition: str = "ungraded",
    region: str = "",
) -> float:
    """Fetch the latest PriceCharting live price (GBP).

    Raises NoPriceAvailable when the page exists but has no price listed.
    Other scraper/network exceptions propagate unchanged and count toward
    the circuit breaker.
    """
    _, live_price_gbp = await scraper.scrape_card(pc_url, condition, region)
    if live_price_gbp is None:
        raise NoPriceAvailable(f"No price listed on PriceCharting for '{card_name}'")
    return live_price_gbp


def _check_price_shift(item: dict, new_live: float, big_shifts: list[dict]) -> None:
    old_live = item.get("live_price")
    if not old_live:
        return
    shift = abs(new_live - old_live) / old_live
    if shift > 0.10:
        big_shifts.append({
            "name":  item["card_name"],
            "old":   old_live,
            "new":   new_live,
            "shift": shift,
        })


def _build_update_embed(
    checked: int,
    updated: int,
    skipped: int,
    big_shifts: list[dict],
    aborted: bool,
    abort_reason: str,
    title_prefix: str = "",
    summary: Optional[dict] = None,
    invalid_url_count: int = 0,
    underwater_items: Optional[list[dict]] = None,
    ebay_sync_stats: Optional[dict] = None,
    pending_verification: int = 0,
    health_issues: Optional[dict] = None,
) -> discord.Embed:
    if aborted:
        embed = discord.Embed(
            title="Price Update Aborted",
            description=(
                f"Circuit breaker triggered after **{_CIRCUIT_BREAKER_THRESHOLD}** "
                f"consecutive failures.\n**Reason:** {abort_reason[:300]}\n\n"
                "Paused until the next scheduled cycle."
            ),
            colour=discord.Colour.red(),
        )
        embed.add_field(name="Items Checked Before Abort", value=str(checked), inline=True)
        return embed

    label = title_prefix or f"{config.UPDATE_INTERVAL_HOURS:.0f}-Hour"
    embed = discord.Embed(
        title=f"{label} Price Update Complete",
        colour=discord.Colour.green(),
    )
    embed.add_field(name="Items Checked",      value=str(checked),        inline=True)
    embed.add_field(name="Prices Refreshed",   value=str(updated),        inline=True)
    embed.add_field(name="Skipped (no data)",  value=str(skipped),        inline=True)
    embed.add_field(name=">10% Price Shifts",  value=str(len(big_shifts)), inline=True)

    if big_shifts:
        lines = [
            f"• **{s['name'][:35]}**: £{s['old']:.2f} → £{s['new']:.2f} "
            f"({s['shift'] * 100:.0f}%)"
            for s in big_shifts[:10]
        ]
        embed.add_field(name="Notable Shifts", value="\n".join(lines), inline=False)

    if summary:
        totals_text = (
            f"Cost in stock:          £{summary.get('total_cost_in_stock', 0):.2f}\n"
            f"Potential value:         £{summary.get('total_potential_in_stock', 0):.2f}\n"
            f"Potential profit:        £{summary.get('total_potential_profit_in_stock', 0):.2f}\n"
            f"Sold revenue (lifetime): £{summary.get('total_sold_revenue', 0):.2f}"
        )
        embed.add_field(name="Inventory Totals", value=totals_text, inline=False)

    if invalid_url_count > 0:
        embed.add_field(
            name="⚠️ Invalid URLs",
            value=(
                f"{invalid_url_count} item(s) have a URL that doesn't start with "
                f"`https://www.pricecharting.com/` and will never update. "
                f"Run `/stock` to find them and `/remove` to delete them."
            ),
            inline=False,
        )

    if underwater_items:
        lines = []
        for u in underwater_items[:10]:
            diff = u["purchase"] - u["live"]
            lines.append(
                f"  #{u['item_id']} {u['name'][:30]} — "
                f"bought £{u['purchase']:.2f}, now £{u['live']:.2f} (-£{diff:.2f})"
            )
        embed.add_field(
            name=f"⚠️ Underwater cards ({len(underwater_items)})",
            value="\n".join(lines),
            inline=False,
        )

    if not config.AUTO_SYNC_EBAY_PRICES:
        embed.add_field(
            name="📦 eBay price sync",
            value="Disabled (`AUTO_SYNC_EBAY_PRICES=false`)",
            inline=True,
        )
    elif ebay_sync_stats and (ebay_sync_stats.get("updated", 0) + ebay_sync_stats.get("not_managed", 0)) > 0:
        embed.add_field(
            name="📦 eBay prices auto-synced",
            value=f"{ebay_sync_stats['updated']} updated, {ebay_sync_stats['not_managed']} not API-managed",
            inline=True,
        )

    if pending_verification > 0:
        embed.add_field(
            name="🔍 Pending verification",
            value=(
                f"{pending_verification} item(s) with price changes\n"
                "Run `/verify mode:all` to confirm, or `/verify mode:single item_id:X` to review individually"
            ),
            inline=False,
        )

    if health_issues:
        total_issues = (len(health_issues.get("zero_price", []))
                        + len(health_issues.get("no_pc_url", []))
                        + len(health_issues.get("stale_price", [])))
        if total_issues > 0:
            lines = []
            for item in health_issues["zero_price"][:5]:
                lines.append(f"💰 #{item['item_id']} {item['card_name'][:30]} — £0.00 market price")
            for item in health_issues["no_pc_url"][:5]:
                lines.append(f"🔗 #{item['item_id']} {item['card_name'][:30]} — missing PC URL")
            for item in health_issues["stale_price"][:5]:
                lines.append(f"⏰ #{item['item_id']} {item['card_name'][:30]} — price not updated in 14+ days")
            overflow = total_issues - len(lines)
            embed.add_field(
                name=f"🔍 Data Quality ({total_issues} issues)",
                value="\n".join(lines) + (f"\n…and {overflow} more" if overflow > 0 else ""),
                inline=False,
            )

    return embed


# ---------------------------------------------------------------------------
# Price data quality check
# ---------------------------------------------------------------------------

async def _check_price_health() -> dict:
    """Identify inventory items with missing, zero, or stale price data."""
    all_items = await excel_db.get_stock_async(status_filter="Inventory")

    zero_price:  list[dict] = []
    no_pc_url:   list[dict] = []
    stale_price: list[dict] = []

    now = datetime.now(timezone.utc)

    for item in all_items:
        live   = float(item.get("live_price") or 0)
        pc_url = item.get("pc_url", "")

        if not pc_url:
            no_pc_url.append(item)
            continue

        if live <= 0:
            zero_price.append(item)
            continue

        # Stale: no price update recorded in the last 14 days
        try:
            history = await excel_db.get_item_price_history_async(item["item_id"], limit=1)
            if history:
                last_ts = datetime.fromisoformat(str(history[-1]["timestamp"]))
                if last_ts.tzinfo is None:
                    last_ts = last_ts.replace(tzinfo=timezone.utc)
                if (now - last_ts).days > 14:
                    stale_price.append(item)
        except Exception:
            pass

    return {
        "zero_price":  zero_price,
        "no_pc_url":   no_pc_url,
        "stale_price": stale_price,
    }


# ---------------------------------------------------------------------------
# Shared price-refresh core (used by both the loop and /update)
# ---------------------------------------------------------------------------

async def _run_price_refresh(
    sleep_range: tuple[float, float] = (120.0, 420.0),
    progress_callback=None,
) -> tuple:
    """
    Refresh Live_Price for every Inventory item.

    Parameters
    ----------
    sleep_range       : (min_secs, max_secs) random sleep between requests.
                        Background loop uses (120, 420); /update uses (3, 8).
    progress_callback : Optional async callable(checked, updated) invoked every
                        5 successful items so callers can edit a progress message.

    Returns
    -------
    (checked, updated, skipped, big_shifts, aborted, abort_reason,
     invalid_url_count, underwater_items, ebay_sync_stats)
    """
    items = await excel_db.get_stock_async(STATUS_INVENTORY)
    if not items:
        return 0, 0, 0, [], False, "", 0, [], {"updated": 0, "not_managed": 0}

    _PC_BASE = "https://www.pricecharting.com/"
    invalid_url_count = sum(
        1 for it in items
        if not (it.get("pc_url") or "").startswith(_PC_BASE)
    )

    checked                = 0
    updated                = 0
    skipped                = 0
    big_shifts: list[dict] = []
    underwater_items: list[dict] = []
    ebay_updated           = 0
    ebay_not_managed       = 0
    consecutive_failures   = 0
    aborted                = False
    abort_reason           = ""

    for item in items:
        pc_url = item.get("pc_url") or ""
        if not pc_url:
            continue

        item_id   = item["item_id"]
        card_name = item["card_name"]
        condition = item.get("condition") or "ungraded"
        region    = item.get("region") or ""

        if scraper.get_pc_429_streak() >= scraper._PC_429_ABORT_THRESHOLD:
            aborted      = True
            abort_reason = f"Aborted after {scraper.get_pc_429_streak()} consecutive 429s from PriceCharting"
            print(f"[price_update] {abort_reason}")
            break

        try:
            new_live = await _scrape_fresh_price(pc_url, card_name, condition, region)
            consecutive_failures = 0
        except NoPriceAvailable as exc:
            print(f"[price_update] No price for item {item_id} ({card_name}) — skipping")
            skipped += 1
            await asyncio.sleep(random.uniform(*sleep_range))
            continue
        except Exception as exc:
            consecutive_failures += 1
            print(f"[price_update] Error for item {item_id} ({card_name}): {exc}")
            if consecutive_failures >= _CIRCUIT_BREAKER_THRESHOLD:
                aborted      = True
                abort_reason = str(exc)
                break
            delay = min(random.uniform(*sleep_range) + scraper._pc_base_delay, 600.0)
            await asyncio.sleep(delay)
            continue

        _check_price_shift(item, new_live, big_shifts)
        await excel_db.update_live_price_async(
            item_id, new_live,
            card_name=card_name,
            region=region,
            old_price=item.get("live_price"),
        )
        checked += 1
        updated += 1

        # Underwater detection
        purchase_price = item.get("purchase_price") or 0.0
        if purchase_price > 0 and new_live < purchase_price:
            underwater_items.append({
                "item_id":  item_id,
                "name":     card_name,
                "purchase": purchase_price,
                "live":     new_live,
            })

        # Quick sell price — eBay competitor lookup (non-critical, never blocks)
        quick_price = None
        try:
            comp = await lister_ebay.get_competitor_price(card_name, condition)
            if comp:
                purchase_price_val = float(item.get("purchase_price") or 0)
                min_profitable = max(round(purchase_price_val * 1.10, 2), config.EBAY_MIN_PRICE_GBP)
                quick_price = max(comp["quick_sell"], min_profitable)
                await excel_db.update_quick_price_async(item_id, quick_price)
                print(
                    f"[price_loop] #{item_id} quick_price £{quick_price:.2f}"
                    f" — {comp['strategy_note']}"
                )
            await asyncio.sleep(0.2)  # rate-limit eBay Browse API calls
        except Exception:
            pass

        if _SYNC_ENABLED:
            try:
                potential_profit = round(new_live - (item.get("purchase_price") or 0), 2)
                bot_sync.sync_price_update(item_id, new_live, quick_price, potential_profit)
            except Exception as exc:
                print(f"[bot_sync] price sync failed for item {item_id}: {exc}")

        # eBay repricing — push Quick_Price (or 95% live) after every refresh
        ebay_id = str(item.get("ebay_listing_id", "") or "").strip()
        if item.get("ebay_listed") == "Yes" and ebay_id.isdigit():
            if config.AUTO_SYNC_EBAY_PRICES:
                purchase_price_val = float(item.get("purchase_price") or 0)
                min_profitable     = max(round(purchase_price_val * 1.10, 2), config.EBAY_MIN_PRICE_GBP)
                stored_quick       = float(item.get("quick_price") or 0)
                if stored_quick >= min_profitable:
                    new_ebay_price = stored_quick
                else:
                    new_ebay_price = max(round(new_live * 0.95, 2), min_profitable)
                try:
                    sync_ok = await lister_ebay.revise_listing_price(
                        ebay_id, new_ebay_price, sku=f"pokemaz-{item_id}"
                    )
                    if sync_ok:
                        ebay_updated += 1
                        print(f"[price_sync] Updated eBay listing {ebay_id}: £{new_ebay_price:.2f}")
                    else:
                        ebay_not_managed += 1
                except Exception as exc:
                    ebay_not_managed += 1
                    print(f"[price_sync] Error for item {item_id}: {exc}")
                await asyncio.sleep(0.3)

        if progress_callback is not None and checked % 5 == 0:
            try:
                await progress_callback(checked, updated)
            except Exception:
                pass

        await asyncio.sleep(random.uniform(*sleep_range))

    return (checked, updated, skipped, big_shifts, aborted, abort_reason,
            invalid_url_count, underwater_items, {"updated": ebay_updated, "not_managed": ebay_not_managed})


# ---------------------------------------------------------------------------
# Excel integrity check
# ---------------------------------------------------------------------------

async def _check_excel_integrity() -> bool:
    """Verify inventory.xlsx is readable and has expected structure."""
    try:
        import openpyxl as _openpyxl
        wb = _openpyxl.load_workbook(config.EXCEL_FILE, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        required = ["Item_ID", "Card_Name", "Status", "Purchase_Price"]
        missing = [h for h in required if h not in headers]
        wb.close()
        if missing:
            print(f"[bot] ⚠️ Excel missing columns: {missing}")
            return False
        row_count = ws.max_row - 1
        print(f"[bot] ✅ Excel integrity OK — {row_count} rows, {len(headers)} columns")
        return True
    except Exception as e:
        print(f"[bot] 🔴 Excel integrity FAILED: {e}")
        backup_dir = Path(config.BACKUP_DIR)
        if backup_dir.exists():
            backups_list = sorted(backup_dir.glob("inventory-*.xlsx"),
                                  key=lambda p: p.stat().st_mtime, reverse=True)
            if backups_list:
                print(f"[bot] Attempting restore from {backups_list[0].name}...")
                import shutil as _shutil
                _shutil.copy2(str(backups_list[0]), config.EXCEL_FILE)
                print(f"[bot] ✅ Restored from backup")
                return True
        return False


# ---------------------------------------------------------------------------
# Watchdog loop — health monitoring
# ---------------------------------------------------------------------------

@tasks.loop(minutes=15)
async def watchdog_loop():
    """Monitor bot health and alert on failures."""
    await bot.wait_until_ready()

    issues = []
    now = datetime.now(timezone.utc)

    if _last_price_update:
        hours_since = (now - _last_price_update).total_seconds() / 3600
        if hours_since > 14:
            issues.append(f"⏰ Price update overdue — last ran {hours_since:.1f}h ago (expected every 12h)")

    if _last_ebay_sales_check:
        mins_since = (now - _last_ebay_sales_check).total_seconds() / 60
        if mins_since > 45:
            issues.append(f"📦 eBay sales check overdue — last ran {mins_since:.0f}m ago (expected every 30m)")

    total_pc = _price_update_stats["pc_success"] + _price_update_stats["pc_429s"]
    if total_pc > 10:
        error_rate = _price_update_stats["pc_429s"] / total_pc
        if error_rate > 0.5:
            issues.append(f"🚫 PriceCharting rate-limited — {error_rate:.0%} of requests failing")

    try:
        import openpyxl as _openpyxl2
        _wb = _openpyxl2.load_workbook(config.EXCEL_FILE, data_only=True)
        _wb.close()
    except Exception as e:
        issues.append(f"🔴 Excel file corrupted or locked: {e}")

    if not issues:
        _watchdog_alerts_sent.clear()
        return

    new_issues = [i for i in issues if i not in _watchdog_alerts_sent]
    if not new_issues:
        return

    channel = bot.get_channel(config.PRICE_UPDATE_CHANNEL_ID)
    if channel and new_issues:
        owner_id = getattr(config, "OWNER_USER_ID", 0)
        owner_mention = f"<@{owner_id}>" if owner_id else "@here"
        embed = discord.Embed(
            title="⚠️ Pokemaz Health Alert",
            description="\n".join(new_issues),
            colour=discord.Colour.red(),
        )
        embed.set_footer(text="This alert won't repeat until the issue resolves")
        await channel.send(content=owner_mention, embed=embed)
        _watchdog_alerts_sent.update(new_issues)


@watchdog_loop.error
async def _watchdog_loop_error(error: Exception) -> None:
    print(f"[watchdog] Unhandled error: {error}")


# ---------------------------------------------------------------------------
# Background price updater — task
# ---------------------------------------------------------------------------

@tasks.loop(hours=config.UPDATE_INTERVAL_HOURS)
async def price_update_loop() -> None:
    """
    Silently refresh Live_Price for every Inventory item once per cycle.
    Uses long random jitter (2–7 min) between requests to avoid bot detection.
    """
    global _last_price_update
    _price_update_stats["pc_429s"]   = 0
    _price_update_stats["pc_success"] = 0

    await backups.snapshot_inventory("daily-snapshot")
    await _maybe_create_monthly_sheet()

    channel = (
        bot.get_channel(config.PRICE_UPDATE_CHANNEL_ID)
        if config.PRICE_UPDATE_CHANNEL_ID
        else None
    )

    (checked, updated, skipped, big_shifts, aborted, abort_reason,
     invalid_url_count, underwater_items, ebay_sync_stats) = (
        await _run_price_refresh(sleep_range=(120.0, 420.0))
    )

    _price_update_stats["checked"] = checked
    _price_update_stats["updated"] = updated
    _last_price_update = datetime.now(timezone.utc)

    try:
        await _check_watchlist()
    except Exception as exc:
        print(f"[watchlist] Error during watchlist check: {exc}")

    if checked == 0 and skipped == 0 and not aborted:
        return  # empty inventory or nothing scraped — skip report

    if channel is None:
        print(
            f"[price_update] Cycle done — checked={checked}, updated={updated}, "
            f"skipped={skipped}, aborted={aborted}, invalid_urls={invalid_url_count}"
        )
        return

    try:
        loop_summary = await excel_db.get_summary_async()
    except Exception:
        loop_summary = None

    try:
        loop_unverified = await excel_db.get_unverified_price_changes_async()
        loop_pending_verification = len(loop_unverified)
    except Exception:
        loop_pending_verification = 0

    try:
        loop_health = await _check_price_health()
    except Exception as exc:
        print(f"[price_update] Health check error: {exc}")
        loop_health = None

    await channel.send(
        embed=_build_update_embed(
            checked, updated, skipped, big_shifts, aborted, abort_reason,
            summary=loop_summary,
            invalid_url_count=invalid_url_count,
            underwater_items=underwater_items,
            ebay_sync_stats=ebay_sync_stats,
            pending_verification=loop_pending_verification,
            health_issues=loop_health,
        )
    )


@price_update_loop.error
async def _price_update_loop_error(error: Exception) -> None:
    print(f"[price_update] Unhandled task error: {error}")
    traceback.print_exception(type(error), error, error.__traceback__)


# ---------------------------------------------------------------------------
# Background eBay sales detector
# ---------------------------------------------------------------------------

async def _get_recent_orders(access_token: str) -> list[dict]:
    """Fetch orders created in the last 30 days via the eBay Sell Fulfillment API."""
    import aiohttp
    since = (datetime.now(timezone.utc) - timedelta(days=30)).strftime("%Y-%m-%dT%H:%M:%SZ")
    async with aiohttp.ClientSession() as session:
        async with session.get(
            "https://api.ebay.com/sell/fulfillment/v1/order",
            headers={
                "Authorization": f"Bearer {access_token}",
                "X-EBAY-C-MARKETPLACE-ID": "EBAY_GB",
            },
            params={
                "filter": f"creationdate:[{since}]",
                "limit": "50",
            },
            timeout=aiohttp.ClientTimeout(total=15),
        ) as resp:
            if resp.status != 200:
                print(f"[ebay_sales] Orders API HTTP {resp.status}: {await resp.text()}")
                return []
            data = await resp.json()
            return data.get("orders", [])


async def _check_cancelled_orders(orders: list[dict]) -> list[dict]:
    """
    Detect fully-cancelled eBay orders and revert any of their items that were
    already marked Sold back to Inventory. A single listing can map to several
    inventory rows (a multi-card set sold under one listing), so every matching
    Sold row for a cancelled listing is reverted.
    """
    channel = (
        bot.get_channel(config.PRICE_UPDATE_CHANNEL_ID)
        if config.PRICE_UPDATE_CHANNEL_ID
        else None
    )
    reverted: list[dict] = []
    all_items = await excel_db.get_stock_async(status_filter=None)

    for order in orders:
        cancel_status = order.get("cancelStatus", {}) or {}
        if cancel_status.get("cancelState") != "CANCEL_COMPLETE":
            continue

        order_id = order.get("orderId", "")
        order_reverted: list[dict] = []

        for line_item in order.get("lineItems", []):
            legacy_item_id = str(line_item.get("legacyItemId", "") or "").strip()
            if not legacy_item_id:
                continue

            matching_items = [
                i for i in all_items
                if str(i.get("ebay_listing_id", "") or "").strip() == legacy_item_id
                and str(i.get("status", "")).strip() == "Sold"
            ]
            if not matching_items:
                continue

            for match_item in matching_items:
                inv_item_id = int(float(str(match_item["item_id"])))
                try:
                    await excel_db.unsell_item_async(inv_item_id)
                except ValueError as exc:
                    print(f"[ebay_sales] Could not revert item {inv_item_id}: {exc}")
                    continue

                print(f"[ebay_sales] Order {order_id} cancelled — reverted #{inv_item_id} "
                      f"{match_item['card_name']} back to Inventory")

                if _SYNC_ENABLED:
                    try:
                        bot_sync.sync_unsell_item(inv_item_id)
                    except Exception as exc:
                        print(f"[bot_sync] unsell sync failed for item {inv_item_id}: {exc}")

                entry = {"item_id": inv_item_id, "card_name": match_item["card_name"], "order_id": order_id}
                order_reverted.append(entry)
                reverted.append(entry)

        if order_reverted and channel is not None:
            names = ", ".join(f"#{r['item_id']} {r['card_name'][:30]}" for r in order_reverted)
            embed = discord.Embed(
                title="Order Cancelled — Reverted to Inventory",
                colour=discord.Colour.orange(),
            )
            embed.add_field(name="Item(s)", value=names, inline=False)
            embed.add_field(name="Order",   value=order_id, inline=True)
            try:
                await channel.send(embed=embed)
            except Exception as exc:
                print(f"[ebay_sales] Could not send cancellation notification: {exc}")

    return reverted


async def _sync_relisted_items() -> None:
    """
    Verify active eBay listings for Inventory items are still live. A listing
    can end without our knowledge if eBay cancels/expires it, so a stale
    ebay_listing_id would otherwise block re-listing that item. Clear it once
    the listing is confirmed inactive.
    """
    all_items = await excel_db.get_stock_async(status_filter="Inventory")
    listed_items = [
        item for item in all_items
        if str(item.get("ebay_listing_id", "") or "").strip().isdigit()
    ]
    if not listed_items:
        return

    for item in listed_items:
        listing_id = str(item["ebay_listing_id"]).strip()
        try:
            is_active = await lister_ebay.check_listing_active(listing_id)
        except Exception as exc:
            print(f"[ebay_sync] Could not check listing {listing_id}: {exc}")
            continue

        if is_active:
            continue

        print(f"[ebay_sync] Listing {listing_id} for #{item['item_id']} {item['card_name']} "
              f"is no longer active — clearing.")
        try:
            await excel_db.mark_ebay_delisted_async(item["item_id"])
        except Exception as exc:
            print(f"[ebay_sync] Could not clear listing for item {item['item_id']}: {exc}")
            continue

        if _SYNC_ENABLED:
            try:
                bot_sync.sync_ebay_delisted(item["item_id"])
            except Exception as exc:
                print(f"[bot_sync] ebay delisted sync failed for item {item['item_id']}: {exc}")


async def _process_ebay_sales() -> None:
    """
    1. Get all active eBay listing IDs from inventory
    2. Fetch recent orders from the eBay Sell Fulfillment API
    3. Match orders to inventory items by SKU (pokemaz-{item_id})
    4. Mark matched items sold at net proceeds and post a Discord notification

    NOTE: The Sell Fulfillment API requires sell.fulfillment.readonly scope.
    If orders return 403, re-run generate_ebay_token.py to refresh the token scope.
    The refresh token in .env must have been generated AFTER the scope was added.
    """
    # Step 1 — filter inventory to items with an active eBay listing
    all_items = await excel_db.get_stock_async(status_filter=None)
    inventory_items = [
        item for item in all_items
        if str(item.get("ebay_listing_id", "") or "").strip().isdigit()
        and str(item.get("status", "")).strip() == "Inventory"
    ]
    print(f"[ebay_sales] Found {len(inventory_items)} active eBay-listed item(s) in inventory")
    # Key by eBay listing ID string — matches legacyItemId returned by the Orders API
    listed = {}
    for item in inventory_items:
        ebay_id = str(item.get("ebay_listing_id", "") or "").strip()
        if ebay_id and ebay_id.isdigit():
            listed[ebay_id] = item
    if not listed:
        print("[ebay_sales] No eBay-listed items in inventory — skipping.")
        return

    # Step 2 — fetch recent orders
    access_token = await lister_ebay._get_access_token()
    orders = await _get_recent_orders(access_token)
    print(f"[ebay_sales] Fetched {len(orders)} order(s) from eBay (last 30 days)")
    if not orders:
        print("[ebay_sales] No recent orders found.")
        return

    # Step 2b — revert any orders that were sold then cancelled, before matching new sales
    await _check_cancelled_orders(orders)

    channel = (
        bot.get_channel(config.PRICE_UPDATE_CHANNEL_ID)
        if config.PRICE_UPDATE_CHANNEL_ID
        else None
    )

    matched = 0
    skipped_already_sold = 0

    # Step 3 — match orders to inventory items by legacyItemId (= eBay listing ID)
    for order in orders:
        order_id = order.get("orderId", "")
        if not order_id or order_id in _processed_order_ids:
            continue

        if (order.get("cancelStatus", {}) or {}).get("cancelState") == "CANCEL_COMPLETE":
            continue

        line_items = order.get("lineItems", [])
        num_line_items = len(line_items)

        # Order-level totals cover ALL line items combined — for a bundle order
        # (multiple distinct listings in one order) these must be prorated per
        # line item rather than credited in full to every matched listing.
        order_pricing = order.get("pricingSummary", {})
        order_gross_sale = float(order_pricing.get("total", {}).get("value", 0) or 0)
        order_delivery_raw = order_pricing.get("deliveryCost", {}).get("value", "0")
        order_delivery_cost = float(order_delivery_raw) if order_delivery_raw else 0.0

        order_marketplace_fee = order.get("totalMarketplaceFee", {})
        if order_marketplace_fee and order_marketplace_fee.get("value"):
            order_ebay_fees = float(order_marketplace_fee["value"])
            fee_source = "actual"
        else:
            order_ebay_fees = round(order_gross_sale * config.EBAY_FEE_RATE, 2)
            fee_source = "estimated"

        line_item_totals = [float(li.get("total", {}).get("value", 0) or 0) for li in line_items]
        sum_line_item_totals = sum(line_item_totals) or order_gross_sale

        for line_item, line_item_total in zip(line_items, line_item_totals):
            legacy_item_id = str(line_item.get("legacyItemId", "") or "").strip()
            if not legacy_item_id:
                continue

            print(f"[ebay_sales/debug] Order listing ID: {legacy_item_id!r} — {line_item.get('title', '')[:40]}")
            matching_items = [
                i for i in inventory_items
                if str(i.get("ebay_listing_id", "") or "").strip() == legacy_item_id
            ]
            print(f"[ebay_sales/debug] Match found: {len(matching_items)} item(s)")

            if not matching_items:
                continue

            # Step 4 — calculate this line item's share of the order-level totals
            share = (line_item_total / sum_line_item_totals) if (num_line_items > 1 and sum_line_item_totals) else 1.0
            gross_sale = round(order_gross_sale * share, 2) if num_line_items > 1 else order_gross_sale
            if not gross_sale:
                print(f"[ebay_sales] Order {order_id}: zero gross sale for listing {legacy_item_id} — skipping.")
                continue

            ebay_fees = round(order_ebay_fees * share, 2) if num_line_items > 1 else order_ebay_fees
            print(f"[ebay_sales] Fees ({fee_source}): £{ebay_fees:.2f}")

            # Delivery cost passes through eBay to cover the postage label — not profit
            delivery_cost = round(order_delivery_cost * share, 2) if num_line_items > 1 else order_delivery_cost
            item_sale_value = gross_sale - delivery_cost if delivery_cost > 0 else gross_sale
            net_proceeds = round(item_sale_value - ebay_fees, 2)

            print(
                f"[ebay_sales] Order {order_id}: gross=£{gross_sale:.2f}, delivery=£{delivery_cost:.2f}, "
                f"fees=£{ebay_fees:.2f} ({fee_source}), net=£{net_proceeds:.2f}, "
                f"items={len(matching_items)}"
            )

            # Mark order as processed before the DB write to avoid double-processing
            _processed_order_ids.add(order_id)
            _persist_processed_order_id(order_id)

            # Record sale for each matched item
            sold_results: list[tuple] = []
            for match_item in matching_items:
                inv_item_id = int(float(str(match_item["item_id"])))
                item_net = float(match_item.get("sell_price") or 0)
                if item_net <= 0:
                    item_net = round(net_proceeds / len(matching_items), 2)
                print(f"[ebay_sales] Bundle item sold — #{inv_item_id} ({match_item['card_name'][:40]}): net=£{item_net:.2f}")
                try:
                    result = await excel_db.sell_item_async(inv_item_id, item_net)
                    sold_results.append((match_item, inv_item_id, item_net, result))
                    matched += 1
                    if _SYNC_ENABLED:
                        try:
                            bot_sync.sync_ebay_sold(inv_item_id, item_net, result["profit"])
                        except Exception as exc:
                            print(f"[bot_sync] ebay sold sync failed for item {inv_item_id}: {exc}")
                except ValueError as exc:
                    print(f"[ebay_sales] sell_item_async skipped item {inv_item_id}: {exc}")
                    skipped_already_sold += 1
                except Exception as exc:
                    print(f"[ebay_sales] sell_item_async error for item {inv_item_id}: {exc}")

            if not sold_results:
                continue

            for match_item, inv_item_id, item_net, result in sold_results:
                audit.log_mutation("ebay_auto_sell", inv_item_id, "auto_sold_via_ebay", {
                    "gross_sale_gbp":    gross_sale,
                    "delivery_cost_gbp": delivery_cost,
                    "item_sale_gbp":     item_sale_value,
                    "ebay_fees_gbp":     ebay_fees,
                    "fee_source":        fee_source,
                    "net_proceeds_gbp":  net_proceeds,
                    "item_net_gbp":      item_net,
                    "bundle_size":       len(matching_items),
                    "order_id":          order_id,
                    "buyer":             order.get("buyer", {}).get("username", "unknown"),
                })

            # Step 5 — post Discord notification
            if channel is None:
                continue

            is_bundle = len(matching_items) > 1
            fee_pct = round(ebay_fees / item_sale_value * 100, 2) if item_sale_value else config.EBAY_FEE_RATE * 100
            fee_label = f"£{ebay_fees:.2f} ({fee_pct:.2f}%, {fee_source})"

            if is_bundle:
                bundle_names = ", ".join(i["card_name"][:20] for i in matching_items)
                total_purchase = sum(float(r.get("purchase_price", 0)) for _, _, _, r in sold_results)
                total_profit   = sum(float(r.get("profit", 0)) for _, _, _, r in sold_results)
                profit_sign = "+" if total_profit >= 0 else ""
                embed = discord.Embed(
                    title="eBay Bundle Sale Detected — Auto-recorded",
                    colour=discord.Colour.green(),
                )
                embed.add_field(name=f"Bundle ({len(matching_items)} items)", value=bundle_names,             inline=False)
                embed.add_field(name="Buyer paid",    value=f"£{gross_sale:.2f}",                             inline=True)
                if delivery_cost > 0:
                    embed.add_field(name="Postage",   value=f"£{delivery_cost:.2f}",                          inline=True)
                embed.add_field(name="Item value",    value=f"£{item_sale_value:.2f}",                        inline=True)
                embed.add_field(name="eBay fees",     value=fee_label,                                        inline=True)
                embed.add_field(name="Net proceeds",  value=f"£{net_proceeds:.2f}",                           inline=True)
                embed.add_field(name="Purchase cost", value=f"£{total_purchase:.2f}",                         inline=True)
                embed.add_field(name="Profit",        value=f"**{profit_sign}£{total_profit:.2f}**",          inline=True)
            else:
                match_item, inv_item_id, item_net, result = sold_results[0]
                purchase_price = result.get("purchase_price", 0.0)
                profit = result.get("profit", 0.0)
                profit_sign = "+" if profit >= 0 else ""
                embed = discord.Embed(
                    title="eBay Sale Detected — Auto-recorded",
                    colour=discord.Colour.green(),
                )
                embed.add_field(name="Card",          value=match_item["card_name"],                          inline=False)
                embed.add_field(name="Item ID",       value=str(inv_item_id),                                 inline=True)
                embed.add_field(name="Buyer paid",    value=f"£{gross_sale:.2f}",                             inline=True)
                if delivery_cost > 0:
                    embed.add_field(name="Postage",   value=f"£{delivery_cost:.2f}",                          inline=True)
                embed.add_field(name="Item value",    value=f"£{item_sale_value:.2f}",                        inline=True)
                embed.add_field(name="eBay fees",     value=fee_label,                                        inline=True)
                embed.add_field(name="Net proceeds",  value=f"£{net_proceeds:.2f}",                           inline=True)
                embed.add_field(name="Purchase cost", value=f"£{purchase_price:.2f}",                         inline=True)
                embed.add_field(name="Profit",        value=f"**{profit_sign}£{profit:.2f}**",                inline=True)

            embed.set_footer(text=f"Order ID: {order_id}")
            try:
                await channel.send(embed=embed)
            except Exception as exc:
                print(f"[ebay_sales] Could not send sale notification: {exc}")

    print(f"[ebay_sales] Matched {matched} order(s) to inventory items, {skipped_already_sold} already sold")
    if matched == 0:
        print("[ebay_sales] No new sales to record")


@tasks.loop(minutes=30)
async def check_ebay_sales_loop() -> None:
    """Poll eBay API for completed sales and auto-record them."""
    global _last_ebay_sales_check
    await bot.wait_until_ready()
    print("[ebay_sales] Checking for completed eBay sales...")
    try:
        await _process_ebay_sales()
    except Exception as exc:
        print(f"[ebay_sales] Error: {exc}")
    try:
        await _sync_relisted_items()
    except Exception as exc:
        print(f"[ebay_sync] Error: {exc}")
    _last_ebay_sales_check = datetime.now(timezone.utc)


@check_ebay_sales_loop.error
async def _check_ebay_sales_loop_error(error: Exception) -> None:
    print(f"[ebay_sales] Unhandled task error: {error}")
    traceback.print_exception(type(error), error, error.__traceback__)


# ---------------------------------------------------------------------------
# /add
# ---------------------------------------------------------------------------

@bot.tree.command(name="add", description="Add a Pokémon card to inventory by scraping PriceCharting.")
@app_commands.describe(
    pc_url="Full PriceCharting URL for the card",
    condition="Card condition — select from the list",
    purchase_price="Price you paid in GBP (e.g. 50.00)",
    region="Region (JP/KR) — leave blank for standard English cards",
    quantity="Number of copies (default 1). Use for packs or multiples.",
)
@app_commands.choices(condition=_CONDITION_CHOICES, region=_REGION_CHOICES)
async def slash_add(
    interaction: discord.Interaction,
    pc_url: str,
    condition: app_commands.Choice[str],
    purchase_price: float,
    region: Optional[app_commands.Choice[str]] = None,
    quantity: int = 1,
) -> None:
    if quantity < 1 or quantity > 50:
        await interaction.response.send_message(
            "❌ `quantity` must be between 1 and 50.", ephemeral=True
        )
        return

    _PC_PREFIX = "https://www.pricecharting.com/game/"
    if not pc_url.lower().startswith(_PC_PREFIX) or len(pc_url) <= len(_PC_PREFIX):
        await interaction.response.send_message(
            "❌ That doesn't look like a PriceCharting product URL. Expected something like:\n"
            "`https://www.pricecharting.com/game/<set-slug>/<card-slug>`",
            ephemeral=True,
        )
        return

    await interaction.response.defer()

    region_str = region.value if region else ""
    # Scrape once regardless of quantity
    card_name, live_price_gbp = await scraper.scrape_card(pc_url, condition.value, region_str)

    for _sfx in [" (Korean)", " (Japanese)"]:
        if card_name.endswith(_sfx):
            card_name = card_name[: -len(_sfx)]
    if region_str == "KR":
        card_name = f"{card_name} (Korean)"
    elif region_str == "JP":
        card_name = f"{card_name} (Japanese)"

    if quantity == 1:
        try:
            item_id = await excel_db.add_item_async(
                card_name, pc_url, condition.value, region_str, purchase_price, live_price_gbp,
            )
        except Exception as exc:
            await interaction.followup.send(f"Database error: {exc}")
            return

        if _SYNC_ENABLED:
            try:
                bot_sync.sync_add_item(await excel_db.get_item_async(item_id))
            except Exception as exc:
                print(f"[bot_sync] add sync failed for item {item_id}: {exc}")

        live_display = f"£{live_price_gbp:.2f}" if live_price_gbp is not None else "N/A *(not listed)*"
        margin       = round((live_price_gbp or purchase_price) - purchase_price, 2)
        sign         = "+" if margin >= 0 else ""

        embed = discord.Embed(title="Added to Inventory", colour=discord.Colour.green())
        embed.add_field(name="Card",           value=card_name,                  inline=False)
        embed.add_field(name="ID",             value=f"`{item_id}`",             inline=True)
        embed.add_field(name="Condition",      value=condition.value,            inline=True)
        embed.add_field(name="Region",         value=region_str or "—",          inline=True)
        embed.add_field(name="Purchase Price", value=f"£{purchase_price:.2f}",   inline=True)
        embed.add_field(name="Live Price",     value=live_display,               inline=True)
        embed.add_field(name="Est. Margin",    value=f"**{sign}£{margin:.2f}**", inline=True)
        embed.set_footer(text=f"Run /list with item_id={item_id} and attach photos to cross-list.")
        await interaction.followup.send(embed=embed)
        return

    # quantity > 1 — create separate rows, each with Quantity=1
    item_ids: list[int] = []
    errors: list[str] = []
    for _ in range(quantity):
        try:
            iid = await excel_db.add_item_async(
                card_name, pc_url, condition.value, region_str, purchase_price, live_price_gbp,
            )
            item_ids.append(iid)
            if _SYNC_ENABLED:
                try:
                    bot_sync.sync_add_item(await excel_db.get_item_async(iid))
                except Exception as exc:
                    print(f"[bot_sync] add sync failed for item {iid}: {exc}")
        except Exception as exc:
            errors.append(str(exc))
            break  # stop on first DB error

    ids_str  = ", ".join(str(i) for i in item_ids)
    live_display = f"£{live_price_gbp:.2f}" if live_price_gbp is not None else "N/A *(not listed)*"
    margin   = round((live_price_gbp or purchase_price) - purchase_price, 2)
    sign     = "+" if margin >= 0 else ""

    embed = discord.Embed(
        title="Added to Inventory",
        description=f"Added **{len(item_ids)} × {card_name}** to inventory (IDs: {ids_str})",
        colour=discord.Colour.green(),
    )
    embed.add_field(name="Condition",      value=condition.value,            inline=True)
    embed.add_field(name="Region",         value=region_str or "—",          inline=True)
    embed.add_field(name="Purchase Price", value=f"£{purchase_price:.2f} each", inline=True)
    embed.add_field(name="Live Price",     value=live_display,               inline=True)
    embed.add_field(name="Est. Margin",    value=f"**{sign}£{margin:.2f}**", inline=True)
    if errors:
        embed.add_field(name="⚠️ Errors", value="\n".join(errors), inline=False)
    embed.set_footer(text=f"Use /list item_id:<ID> quantity:{len(item_ids)} to list all at once.")
    await interaction.followup.send(embed=embed)


# ---------------------------------------------------------------------------
# /sell
# ---------------------------------------------------------------------------

@bot.tree.command(name="sell", description="Mark one or more items sold. Use commas for multiple: item_ids:45,46 sell_prices:10,12")
@app_commands.describe(
    item_ids="Item ID(s), comma-separated for multiple (e.g. 45 or 45,46,47)",
    sell_prices="Sell price(s) in GBP, comma-separated, matching the order of item_ids",
)
async def slash_sell(
    interaction: discord.Interaction,
    item_ids: str,
    sell_prices: str,
) -> None:
    # Parse item IDs
    id_tokens = [t.strip() for t in item_ids.split(",")]
    parsed_ids: list[int] = []
    for tok in id_tokens:
        if not tok.lstrip("-").isdigit() or int(tok) <= 0:
            await interaction.response.send_message(
                f"`{tok}` is not a valid item ID — item IDs must be positive integers.",
                ephemeral=True,
            )
            return
        parsed_ids.append(int(tok))

    # Parse sell prices
    price_tokens = [t.strip() for t in sell_prices.split(",")]
    parsed_prices: list[float] = []
    for tok in price_tokens:
        try:
            parsed_prices.append(float(tok))
        except ValueError:
            await interaction.response.send_message(
                f"`{tok}` is not a valid price.", ephemeral=True
            )
            return

    # Expand single price to all items
    if len(parsed_prices) == 1 and len(parsed_ids) > 1:
        parsed_prices = parsed_prices * len(parsed_ids)

    if len(parsed_ids) != len(parsed_prices):
        await interaction.response.send_message(
            f"Number of item IDs ({len(parsed_ids)}) doesn't match number of prices ({len(parsed_prices)}). "
            "Provide one price per item, or a single price to apply to all.",
            ephemeral=True,
        )
        return

    if len(parsed_ids) > 3:
        await interaction.response.defer()
    else:
        await interaction.response.defer()

    successes: list[dict] = []
    failures: list[tuple[int, str]] = []
    auto_delisted: set[int] = set()

    for iid, price in zip(parsed_ids, parsed_prices):
        # Auto-delist eBay listing if one is recorded
        try:
            item_data = await excel_db.get_item_async(iid)
            if item_data.get("ebay_listing_id"):
                auto_delisted.add(iid)
                try:
                    await lister_ebay.end_ebay_listing(item_data["ebay_listing_id"])
                    print(f"[sell] Auto-delisted eBay listing {item_data['ebay_listing_id']} for item {iid}")
                except Exception as exc:
                    print(f"[sell] Warning: could not auto-delist {item_data['ebay_listing_id']}: {exc}")
        except Exception as exc:
            print(f"[sell] Warning: could not fetch item {iid} for eBay check: {exc}")

        try:
            result = await excel_db.sell_item_async(iid, price)
            successes.append({"id": iid, "price": price, **result})
            if _SYNC_ENABLED:
                try:
                    bot_sync.sync_sell_item(iid, price, result["profit"])
                except Exception as exc:
                    print(f"[bot_sync] sell sync failed for item {iid}: {exc}")
        except ValueError as exc:
            failures.append((iid, str(exc)))
        except Exception as exc:
            failures.append((iid, f"Database error: {exc}"))

    if len(parsed_ids) == 1 and successes:
        r = successes[0]
        profit = r["profit"]
        colour = discord.Colour.green() if profit >= 0 else discord.Colour.red()
        sign = "+" if profit >= 0 else ""
        embed = discord.Embed(title="Item Sold", colour=colour)
        embed.add_field(name="Card",           value=r["card_name"],               inline=False)
        embed.add_field(name="ID",             value=f"`{r['id']}`",               inline=True)
        embed.add_field(name="Purchase Price", value=f"£{r['purchase_price']:.2f}", inline=True)
        embed.add_field(name="Sell Price",     value=f"£{r['price']:.2f}",         inline=True)
        embed.add_field(name="Profit",         value=f"**{sign}£{profit:.2f}**",   inline=True)
        if r["id"] in auto_delisted:
            embed.add_field(name="eBay", value="Listing ended automatically", inline=True)
        await interaction.followup.send(embed=embed)
        return

    total_profit = sum(r["profit"] for r in successes)
    overall_colour = discord.Colour.green() if total_profit >= 0 else discord.Colour.red()
    embed = discord.Embed(title=f"Bulk Sell — {len(successes)} sold", colour=overall_colour)

    for r in successes:
        profit = r["profit"]
        sign = "+" if profit >= 0 else ""
        ebay_note = " | eBay ended" if r["id"] in auto_delisted else ""
        embed.add_field(
            name=f"ID {r['id']} — {r['card_name']}",
            value=f"Sold £{r['price']:.2f} | Profit **{sign}£{profit:.2f}**{ebay_note}",
            inline=False,
        )

    if successes:
        total_sign = "+" if total_profit >= 0 else ""
        embed.add_field(name="Total Profit", value=f"**{total_sign}£{total_profit:.2f}**", inline=False)

    if failures:
        fail_lines = "\n".join(f"• Item {iid}: {msg}" for iid, msg in failures)
        embed.add_field(name="Failures", value=fail_lines, inline=False)

    await interaction.followup.send(embed=embed)


# ---------------------------------------------------------------------------
# /remove
# ---------------------------------------------------------------------------

@bot.tree.command(name="remove", description="Remove an item from inventory (use for accidental additions).")
@app_commands.describe(
    item_id="Item ID shown in /stock",
    confirm="Set to True to actually delete (defaults to False — first call shows a preview)",
)
async def slash_remove(
    interaction: discord.Interaction,
    item_id: int,
    confirm: bool = False,
) -> None:
    await interaction.response.defer(ephemeral=True)

    try:
        item = await excel_db.get_item_async(item_id)
    except ValueError:
        await interaction.followup.send(f"Item ID {item_id} not found.")
        return

    if item["status"] == "Sold":
        await interaction.followup.send(
            f"Item {item_id} is marked Sold. Sold items can't be removed via this command — "
            "edit the Excel file directly if absolutely necessary. "
            "This protects your sales history from accidental deletion."
        )
        return

    if not confirm:
        live_display = f"£{item['live_price']:.2f}" if item["live_price"] is not None else "N/A"
        embed = discord.Embed(
            title=f"Preview — Remove Item #{item_id}",
            colour=discord.Colour.orange(),
        )
        embed.add_field(name="Card",           value=item["card_name"],                          inline=False)
        embed.add_field(name="Condition",      value=item["condition"] or "—",                   inline=True)
        embed.add_field(name="Region",         value=item["region"] or "—",                      inline=True)
        embed.add_field(name="Purchase Price", value=f"£{item['purchase_price']:.2f}",           inline=True)
        embed.add_field(name="Live Price",     value=live_display,                               inline=True)
        embed.add_field(name="Status",         value=item["status"],                             inline=True)
        embed.add_field(
            name="⚠️ This will permanently delete the row.",
            value=f"Run `/remove item_id:{item_id} confirm:True` to confirm.",
            inline=False,
        )
        await interaction.followup.send(embed=embed)
        return

    try:
        removed = await excel_db.remove_item_async(item_id)
    except ValueError as exc:
        await interaction.followup.send(f"Error: {exc}")
        return

    if _SYNC_ENABLED:
        try:
            bot_sync.sync_remove_item(item_id)
        except Exception as exc:
            print(f"[bot_sync] remove sync failed for item {item_id}: {exc}")

    live_display = f"£{removed['live_price']:.2f}" if removed["live_price"] is not None else "N/A"
    embed = discord.Embed(
        title=f"Removed Item #{item_id}",
        colour=discord.Colour.red(),
    )
    embed.add_field(name="Card",           value=removed["card_name"],               inline=False)
    embed.add_field(name="Condition",      value=removed["condition"] or "—",        inline=True)
    embed.add_field(name="Region",         value=removed["region"] or "—",           inline=True)
    embed.add_field(name="Purchase Price", value=f"£{removed['purchase_price']:.2f}", inline=True)
    embed.add_field(name="Live Price",     value=live_display,                       inline=True)
    embed.set_footer(text="Row permanently deleted. ID will not be reused.")
    await interaction.followup.send(embed=embed)


# ---------------------------------------------------------------------------
# /edit
# ---------------------------------------------------------------------------

@bot.tree.command(name="edit", description="Edit a field on an inventory item.")
@app_commands.describe(
    item_id="Item ID to edit",
    field="Field to change",
    value="New value (use decimal point for prices, e.g. 5.50)",
)
@app_commands.choices(field=[
    app_commands.Choice(name="Card Name",      value="card_name"),
    app_commands.Choice(name="Purchase Price", value="purchase_price"),
    app_commands.Choice(name="PC URL",         value="pc_url"),
    app_commands.Choice(name="Condition",      value="condition"),
    app_commands.Choice(name="Region",         value="region"),
    app_commands.Choice(name="Sell Price",     value="sell_price"),
    app_commands.Choice(name="eBay Listed",    value="ebay_listed"),
])
async def slash_edit(
    interaction: discord.Interaction,
    item_id: int,
    field: app_commands.Choice[str],
    value: str,
) -> None:
    await interaction.response.defer(ephemeral=True)
    try:
        old_val, new_val = await excel_db.edit_item_async(item_id, field.value, value)
    except ValueError as exc:
        await interaction.followup.send(f"❌ {exc}")
        return
    except Exception as exc:
        await interaction.followup.send(f"Database error: {exc}")
        return

    embed = discord.Embed(
        title=f"✏️ Item {item_id} updated",
        colour=discord.Colour.blurple(),
    )
    embed.add_field(name="Field",     value=field.name,                                   inline=True)
    embed.add_field(name="Old value", value=str(old_val) if old_val is not None else "—", inline=True)
    embed.add_field(name="New value", value=str(new_val),                                 inline=True)
    await interaction.followup.send(embed=embed)


# ---------------------------------------------------------------------------
# /list — private helpers
# ---------------------------------------------------------------------------

_E = {
    "wait": "⏳",
    "ok":   "✅",
    "fail": "❌",
    "img":  "🖼️",
    "tag":  "🏷️",
}


async def _download_attachments(
    attachments: list[discord.Attachment],
    dest_dir: Path,
) -> list[Path]:
    dest_dir.mkdir(parents=True, exist_ok=True)

    async def _save_one(att: discord.Attachment, idx: int) -> Path:
        ext  = Path(att.filename).suffix or ".jpg"
        path = dest_dir / f"{idx:02d}{ext}"
        await att.save(path)
        return path

    return list(await asyncio.gather(*[_save_one(a, i) for i, a in enumerate(attachments)]))


def _listing_embed(
    item:        dict,
    price:       float,
    image_count: int,
    ebay_status: str,
    ai_title:    str = "",
    colour:      discord.Colour = discord.Colour.blurple(),
) -> discord.Embed:
    embed = discord.Embed(
        title=f"{_E['tag']} Listing Item #{item['item_id']}",
        colour=colour,
    )
    embed.add_field(name="Card",    value=item["card_name"], inline=False)
    if ai_title:
        embed.add_field(name="AI Title", value=f"`{ai_title}`", inline=False)
    embed.add_field(name="Price",   value=f"£{price:.2f}",   inline=True)
    embed.add_field(name="Images",  value=f"{_E['img']} {image_count}", inline=True)
    embed.add_field(name="eBay UK", value=ebay_status,       inline=False)
    return embed


def _resolve_listing_price(
    item: dict, price_override: float, use_market_price: bool = False
) -> Optional[float]:
    if price_override > 0:
        return price_override
    if not use_market_price and item.get("quick_price") is not None:
        return max(float(item["quick_price"]), config.EBAY_MIN_PRICE_GBP)
    if item.get("live_price") is not None:
        return max(round(item["live_price"] * 1.15, 2), config.EBAY_MIN_PRICE_GBP)
    if item.get("purchase_price") is not None:
        return item["purchase_price"]
    return None


def _price_source(item: dict, price_override: float, use_market_price: bool = False) -> str:
    if price_override > 0:
        return "Manual override"
    if not use_market_price and item.get("quick_price") is not None:
        return "Quick sell (eBay -7%)"
    if item.get("live_price") is not None:
        return "Market +15% (PriceCharting)"
    return "Purchase price (fallback)"


async def _generate_ai_content(
    item: dict,
    status_msg,          # discord.Message or discord.WebhookMessage
    listing_price: float,
    n_images: int,
) -> tuple[str, str]:
    try:
        ai_item_name = item["card_name"]
        if item.get("region") == "KR":
            ai_item_name = ai_item_name.replace(" (Korean)", "") + " - Korean Print"
        elif item.get("region") == "JP":
            ai_item_name = ai_item_name.replace(" (Japanese)", "") + " - Japanese"
        ai_content = await ai_helper.generate_listing_content(
            item_name=ai_item_name,
            condition=item.get("condition") or "ungraded",
            uk_avg_price_gbp=item.get("live_price"),
        )
        ai_title       = ai_content["title"]
        ai_description = ai_content["description"]
        print(f"[list] AI title: {ai_title!r}")
    except Exception as exc:
        print(f"[list] AI helper failed (falling back to card name): {exc}")
        ai_title       = item["card_name"]
        ai_description = ""

    await status_msg.edit(embed=_listing_embed(
        item, listing_price, n_images,
        ebay_status=f"{_E['wait']} Queued",
        ai_title=ai_title,
    ))
    return ai_title, ai_description


def _resolve_listing_result(result, platform: str):
    if isinstance(result, Exception):
        return lister_ebay.ListingResult(platform=platform, success=False, error=str(result))
    return result


def _format_status_line(r) -> str:
    if r.success:
        url_part = f"\n{r.listing_url}" if r.listing_url else " *(URL not captured)*"
        return f"{_E['ok']} Listed{url_part}"
    return f"{_E['fail']} Failed — {r.error}"


# ---------------------------------------------------------------------------
# /list
# ---------------------------------------------------------------------------

async def _list_single_item(
    item: dict,
    attachments: list[discord.Attachment],
    image_paths: list[Path],
    ai_title: str,
    ai_description: str,
    listing_price: float,
    price_override: float,
) -> lister_ebay.ListingResult:
    """List one item on eBay and update the DB record. Returns the ListingResult."""
    item_id = item["item_id"]
    try:
        ebay_result = await lister_ebay.list_item_on_ebay(
            item_name=ai_title,
            price_gbp=listing_price,
            image_paths=image_paths,
            condition=item.get("condition") or "Used",
            description=ai_description,
            region=item.get("region") or "",
            card_name=item.get("card_name") or "",
            pc_url=item.get("pc_url") or "",
            item_id=item_id,
        )
    except Exception as exc:
        return lister_ebay.ListingResult(platform="eBay", success=False, error=str(exc))

    ebay_r = _resolve_listing_result(ebay_result, "eBay")
    if ebay_r.success and ebay_r.listing_url:
        listing_id = ebay_r.listing_url.split("/itm/")[-1]
        if listing_id:
            try:
                await excel_db.mark_ebay_listed_async(item_id, listing_id)
                audit.log_mutation("list", item_id, "listed on eBay", {
                    "ebay_listing_id": listing_id,
                    "price":           listing_price,
                })
                if _SYNC_ENABLED:
                    try:
                        bot_sync.sync_ebay_listed(item_id, listing_id, listing_price)
                    except Exception as exc:
                        print(f"[bot_sync] ebay listed sync failed for item {item_id}: {exc}")
            except Exception as exc:
                print(f"[list] Failed to record eBay listing for item {item_id}: {exc}")
    return ebay_r


async def _slash_list_bulk(
    interaction: discord.Interaction,
    item_ids: list[int],
    raw_attachments: list,
    price_override: float,
    use_market_price: bool = False,
) -> None:
    """Handle bulk /list for comma-separated IDs — sequential, same images for all."""
    attachments = [a for a in raw_attachments if a is not None]
    n_images    = len(attachments)

    # Validate all items exist and are in Inventory
    items: list[dict] = []
    for iid in item_ids:
        try:
            it = await excel_db.get_item_async(iid)
        except ValueError as exc:
            await interaction.followup.send(f"❌ Item {iid} not found: {exc}")
            return
        if it.get("status") != STATUS_INVENTORY:
            await interaction.followup.send(f"❌ Item {iid} ({it['card_name']}) is not in Inventory.")
            return
        items.append(it)

    # Download images once — shared for all listings
    temp_dir = Path(config.TEMP_IMAGES_DIR) / f"bulk_{'_'.join(str(i) for i in item_ids[:3])}"
    results: list[tuple[dict, lister_ebay.ListingResult]] = []

    status_msg = await interaction.followup.send(
        f"🔄 Bulk listing {len(items)} item(s)… (0/{len(items)})"
    )

    try:
        image_paths = await _download_attachments(attachments, temp_dir)

        for idx, item in enumerate(items, 1):
            try:
                await status_msg.edit(
                    content=f"🔄 Listing {idx}/{len(items)}: {item['card_name'][:40]}…"
                )
            except Exception:
                pass

            listing_price = _resolve_listing_price(item, price_override, use_market_price)
            if listing_price is None:
                results.append((item, lister_ebay.ListingResult(
                    platform="eBay", success=False,
                    error="No price recorded — use price_override"
                )))
                continue

            try:
                ai_content   = await ai_helper.generate_listing_content(
                    item_name  = item["card_name"],
                    condition  = item.get("condition") or "ungraded",
                    uk_avg_price_gbp = item.get("live_price"),
                )
                ai_title       = ai_content["title"]
                ai_description = ai_content["description"]
            except Exception:
                ai_title       = item["card_name"]
                ai_description = ""

            ebay_r = await _list_single_item(
                item, attachments, image_paths,
                ai_title, ai_description, listing_price, price_override,
            )
            results.append((item, ebay_r))

        successes = [(it, r) for it, r in results if r.success]
        failures  = [(it, r) for it, r in results if not r.success]
        total_value = sum(
            _resolve_listing_price(it, price_override, use_market_price) or 0.0
            for it, r in successes
        )

        colour = discord.Colour.green() if successes else discord.Colour.red()
        embed  = discord.Embed(
            title=f"✅ Bulk List — {len(successes)}/{len(items)} listed",
            description=f"Total value: £{total_value:.2f}",
            colour=colour,
        )
        for it, r in successes:
            embed.add_field(
                name=f"#{it['item_id']} {it['card_name'][:30]}",
                value=r.listing_url or "*(URL not captured)*",
                inline=False,
            )
        if failures:
            fail_lines = "\n".join(f"• #{it['item_id']} {it['card_name'][:25]}: {r.error}" for it, r in failures)
            embed.add_field(name="Failures", value=fail_lines, inline=False)

        await status_msg.edit(content=None, embed=embed)

    finally:
        if temp_dir.exists():
            await asyncio.to_thread(shutil.rmtree, temp_dir, ignore_errors=True)


@bot.tree.command(name="list", description="Cross-list a card on eBay UK and Vinted UK.")
@app_commands.describe(
    item_id="Item ID shown in /stock — use commas for multiple (e.g. 5,6,7)",
    image1="First listing photo (required)",
    image2="Second photo (optional)",
    image3="Third photo (optional)",
    image4="Fourth photo (optional)",
    image5="Fifth photo (optional)",
    price_override="Listing price in GBP — leave blank to use quick-sell price",
    quantity="Number of identical listings to create (default 1).",
    use_market_price="Use PriceCharting market price (+15%) instead of quick-sell price",
)
async def slash_list(
    interaction: discord.Interaction,
    item_id: str,
    image1: discord.Attachment,
    image2: Optional[discord.Attachment] = None,
    image3: Optional[discord.Attachment] = None,
    image4: Optional[discord.Attachment] = None,
    image5: Optional[discord.Attachment] = None,
    price_override: float = 0.0,
    quantity: int = 1,
    use_market_price: bool = False,
) -> None:
    # Parse item_id — may be a single int or comma-separated list
    id_tokens = [t.strip() for t in str(item_id).split(",")]
    is_multi  = len(id_tokens) > 1

    if is_multi:
        if len(id_tokens) > 10:
            await interaction.response.send_message(
                "❌ Maximum 10 items per bulk call.", ephemeral=True
            )
            return
        parsed_ids: list[int] = []
        for tok in id_tokens:
            if not tok.isdigit() or int(tok) <= 0:
                await interaction.response.send_message(
                    f"❌ `{tok}` is not a valid item ID.", ephemeral=True
                )
                return
            parsed_ids.append(int(tok))
        if quantity > 1:
            await interaction.response.send_message(
                "❌ Cannot combine comma-separated IDs with `quantity`. "
                "Use commas for different items, or a single ID with `quantity` for identical copies.",
                ephemeral=True,
            )
            return
        await interaction.response.defer()
        await _slash_list_bulk(
            interaction, parsed_ids, [image1, image2, image3, image4, image5],
            price_override, use_market_price,
        )
        return

    # Single-item mode — parse the single token as int
    if not id_tokens[0].isdigit() or int(id_tokens[0]) <= 0:
        await interaction.response.send_message("❌ Invalid item ID.", ephemeral=True)
        return
    item_id_int = int(id_tokens[0])

    if quantity < 1 or quantity > 20:
        await interaction.response.send_message(
            "❌ `quantity` must be between 1 and 20.", ephemeral=True
        )
        return

    await interaction.response.defer()

    attachments = [a for a in [image1, image2, image3, image4, image5] if a is not None]

    try:
        anchor_item = await excel_db.get_item_async(item_id_int)
    except ValueError as exc:
        await interaction.followup.send(f"Item not found: {exc}")
        return

    # For quantity > 1, find consecutive matching items (same name + condition, ID >= item_id_int)
    if quantity == 1:
        items_to_list = [anchor_item]
    else:
        all_stock = await excel_db.get_stock_async("Inventory")
        matching = sorted(
            [
                i for i in all_stock
                if i["card_name"] == anchor_item["card_name"]
                and i["condition"] == anchor_item["condition"]
                and i["item_id"] >= item_id_int
            ],
            key=lambda i: i["item_id"],
        )
        items_to_list = matching[:quantity]
        found = len(items_to_list)
        if found == 0:
            await interaction.followup.send(
                f"No matching Inventory items found for item #{item_id_int}."
            )
            return

    listing_price = _resolve_listing_price(anchor_item, price_override, use_market_price)
    if listing_price is None:
        await interaction.followup.send(
            f"Item #{item_id_int} has no price recorded. "
            f"Specify one with the `price_override` parameter."
        )
        return

    n_images   = len(attachments)
    status_msg = await interaction.followup.send(embed=_listing_embed(
        anchor_item, listing_price, n_images,
        ebay_status=f"{_E['wait']} Queued",
        ai_title="🤖 Generating…",
    ))

    ai_title, ai_description = await _generate_ai_content(
        anchor_item, status_msg, listing_price, n_images
    )

    # Fetch competitor prices (best-effort — shown in final embed as info)
    competitor_data: dict | None = None
    try:
        competitor_data = await lister_ebay.get_competitor_price(
            anchor_item["card_name"], anchor_item.get("condition") or ""
        )
        if competitor_data:
            print(
                f"[list] Competitors: {competitor_data['count']} active, "
                f"lowest £{competitor_data['lowest']:.2f}, quick_sell £{competitor_data['quick_sell']:.2f}"
                f" — {competitor_data['strategy_note']}"
            )
    except Exception:
        pass

    temp_dir = Path(config.TEMP_IMAGES_DIR) / f"item_{item_id_int}"

    try:
        image_paths = await _download_attachments(attachments, temp_dir)
        print(f"[list] Downloaded {len(image_paths)} images to {temp_dir}")

        if quantity == 1:
            await status_msg.edit(embed=_listing_embed(
                anchor_item, listing_price, n_images,
                ebay_status=f"{_E['wait']} Listing…",
                ai_title=ai_title,
            ))
            ebay_r = await _list_single_item(
                anchor_item, attachments, image_paths,
                ai_title, ai_description, listing_price, price_override,
            )
            colour = discord.Colour.green() if ebay_r.success else discord.Colour.red()
            final_embed = _listing_embed(
                anchor_item, listing_price, n_images,
                ebay_status=_format_status_line(ebay_r),
                ai_title=ai_title,
                colour=colour,
            )
            if ebay_r.success and ebay_r.listing_url:
                listing_id = ebay_r.listing_url.split("/itm/")[-1]
                if listing_id:
                    final_embed.add_field(name="eBay Listed",  value="Yes",                                               inline=True)
                    final_embed.add_field(name="Price source", value=_price_source(anchor_item, price_override, use_market_price), inline=True)
                    final_embed.add_field(name="eBay URL",     value=f"https://www.ebay.co.uk/itm/{listing_id}",          inline=False)

            # Pricing breakdown
            pc_live = anchor_item.get("live_price")
            stored_quick = anchor_item.get("quick_price")
            pricing_lines = []
            if pc_live:
                pricing_lines.append(f"PriceCharting market:  £{pc_live:.2f}")
            if competitor_data:
                qs = competitor_data["quick_sell"]
                pricing_lines.append(
                    f"Quick sell: £{qs:.2f}  "
                    f"({competitor_data['strategy_note']})"
                )
            elif stored_quick:
                pricing_lines.append(f"Quick sell (stored):   £{stored_quick:.2f}")
            if pricing_lines:
                using = "quick sell" if (not use_market_price and (competitor_data or stored_quick)) else "market +15%"
                pricing_lines.append(f"Your price:            £{listing_price:.2f}  ← using {using}")
                final_embed.add_field(name="💰 Pricing", value="\n".join(pricing_lines), inline=False)

            await status_msg.edit(embed=final_embed)
        else:
            # Multi-quantity path
            found     = len(items_to_list)
            results: list[tuple[dict, lister_ebay.ListingResult]] = []
            await status_msg.edit(embed=_listing_embed(
                anchor_item, listing_price, n_images,
                ebay_status=f"{_E['wait']} Listing 0/{found}…",
                ai_title=ai_title,
            ))
            for idx, item in enumerate(items_to_list, 1):
                ebay_r = await _list_single_item(
                    item, attachments, image_paths,
                    ai_title, ai_description, listing_price, price_override,
                )
                results.append((item, ebay_r))
                try:
                    await status_msg.edit(embed=_listing_embed(
                        anchor_item, listing_price, n_images,
                        ebay_status=f"{_E['wait']} Listed {idx}/{found}…",
                        ai_title=ai_title,
                    ))
                except Exception:
                    pass

            successes = [(it, r) for it, r in results if r.success]
            failures  = [(it, r) for it, r in results if not r.success]
            short_msg = quantity != found

            colour = discord.Colour.green() if successes else discord.Colour.red()
            embed  = discord.Embed(
                title=f"Bulk List — {len(successes)}/{found} listed",
                colour=colour,
            )
            if short_msg:
                embed.description = (
                    f"Only {found} of {quantity} matching items found — listed {len(successes)}."
                )

            urls = [r.listing_url for _, r in successes if r.listing_url]
            if len(urls) <= 5:
                for (it, r) in successes:
                    url_val = r.listing_url or "*(URL not captured)*"
                    embed.add_field(
                        name=f"ID {it['item_id']}",
                        value=url_val,
                        inline=False,
                    )
            elif urls:
                embed.add_field(
                    name=f"Created {len(successes)} listings",
                    value=f"First: {urls[0]}\nLast: {urls[-1]}",
                    inline=False,
                )

            if failures:
                fail_lines = "\n".join(
                    f"• ID {it['item_id']}: {r.error}" for it, r in failures
                )
                embed.add_field(name="Failures", value=fail_lines, inline=False)

            await status_msg.edit(embed=embed)

    finally:
        if temp_dir.exists():
            await asyncio.to_thread(shutil.rmtree, temp_dir, ignore_errors=True)
            print(f"[list] Cleaned up {temp_dir}")



# ---------------------------------------------------------------------------
# /listbundle
# ---------------------------------------------------------------------------

@bot.tree.command(name="listbundle", description="List multiple cards as a single eBay bundle listing.")
@app_commands.describe(
    item_ids="Comma-separated item IDs (e.g. 328,329,330)",
    image1="Primary photo (required)",
    image2="Second photo (optional)",
    image3="Third photo (optional)",
    image4="Fourth photo (optional)",
    image5="Fifth photo (optional)",
    price_override="Override the combined price (optional)",
    discount_pct="Bundle discount % off combined market price (default 10)",
)
async def slash_listbundle(
    interaction: discord.Interaction,
    item_ids: str,
    image1: discord.Attachment,
    image2: Optional[discord.Attachment] = None,
    image3: Optional[discord.Attachment] = None,
    image4: Optional[discord.Attachment] = None,
    image5: Optional[discord.Attachment] = None,
    price_override: float = 0.0,
    discount_pct: float = 10.0,
) -> None:
    await interaction.response.defer()
    print(f"[listbundle] price_override={price_override!r} discount_pct={discount_pct!r}")
    price_override = float(price_override) if price_override else 0.0

    # Parse item IDs
    try:
        ids = [int(x.strip()) for x in item_ids.split(",") if x.strip()]
    except ValueError:
        await interaction.followup.send("❌ Invalid item IDs — use comma-separated numbers e.g. `328,329,330`")
        return

    if len(ids) < 2:
        await interaction.followup.send("❌ Bundle needs at least 2 items.")
        return
    if len(ids) > 10:
        await interaction.followup.send("❌ Maximum 10 items per bundle.")
        return

    # Fetch all items and validate
    items: list[dict] = []
    for iid in ids:
        try:
            item = await excel_db.get_item_async(iid)
        except ValueError:
            await interaction.followup.send(f"❌ Item {iid} not found.")
            return
        if item.get("status") == "Sold":
            await interaction.followup.send(f"❌ Item {iid} ({item['card_name']}) is already sold.")
            return
        if item.get("ebay_listed") == "Yes":
            await interaction.followup.send(
                f"❌ Item {iid} ({item['card_name']}) is already marked as listed on eBay.\n"
                f"If you deleted the listing, reset it with: `/edit item_id:{iid} field:ebay_listed value:No`"
            )
            return
        items.append(item)

    # Calculate prices
    combined_market = sum(float(item.get("live_price") or 0) for item in items)
    combined_cost   = sum(float(item.get("purchase_price") or 0) for item in items)

    if price_override > 0:
        bundle_price = round(price_override, 2)
        print(f"[listbundle] Using price override: £{bundle_price:.2f}")
    else:
        discount       = discount_pct / 100
        bundle_price   = round(combined_market * (1 - discount), 2)
        min_profitable = max(round(combined_cost * 1.10, 2), config.EBAY_MIN_PRICE_GBP)
        bundle_price   = max(bundle_price, min_profitable)
        print(f"[listbundle] Using calculated price: £{bundle_price:.2f} ({discount_pct}% discount)")

    bundle_profit = round(bundle_price - combined_cost, 2)
    roi_pct       = round((bundle_profit / combined_cost * 100) if combined_cost else 0, 1)

    # Generate AI title and description
    bundle_name = " + ".join(item["card_name"] for item in items)
    try:
        content     = await ai_helper.generate_listing_content(
            item_name        = bundle_name,
            condition        = "Near mint or better",
            uk_avg_price_gbp = combined_market,
        )
        title       = content["title"]
        description = content["description"]
    except Exception as e:
        print(f"[listbundle] AI generation failed: {e}")
        title       = f"{bundle_name[:75]} Bundle"[:79]
        description = (
            f"Bundle of {len(items)} Pokémon TCG cards in excellent condition. "
            "Dispatched in protective sleeves and toploaders via tracked postage."
        )

    # Download images
    attachments = [a for a in [image1, image2, image3, image4, image5] if a is not None]
    temp_dir    = Path(config.TEMP_IMAGES_DIR) / f"bundle_{'_'.join(str(i) for i in ids)}"

    try:
        image_paths = await _download_attachments(attachments, temp_dir)

        # List on eBay — use first item's metadata for condition/category aspects
        sku    = f"pokemaz-bundle-{'-'.join(str(i) for i in ids)}"
        first  = items[0]
        result = await lister_ebay.list_item_on_ebay(
            item_name   = title,
            price_gbp   = bundle_price,
            image_paths = image_paths,
            condition   = first.get("condition") or "Near mint or better",
            description = description,
            region      = first.get("region") or "",
            card_name   = first.get("card_name") or "",
            pc_url      = first.get("pc_url") or "",
            sku         = sku,
        )

        if not result.success:
            await interaction.followup.send(f"❌ eBay listing failed: {result.error}")
            return

        listing_url = result.listing_url
        listing_id  = listing_url.split("/itm/")[-1].rstrip("/").split("?")[0]

        # Update each item: mark as listed with proportional sell price
        has_zero_price = any(float(item.get("live_price") or 0) <= 0 for item in items)
        for item in items:
            item_market = float(item.get("live_price") or 0)
            if has_zero_price or combined_market <= 0:
                item_share = round(bundle_price / len(items), 2)
            else:
                item_share = round(bundle_price * item_market / combined_market, 2)
            try:
                await excel_db.update_item_listing_async(item["item_id"], listing_id, item_share)
            except Exception as e:
                print(f"[listbundle] Failed to update item {item['item_id']}: {e}")

        # Build result embed
        embed = discord.Embed(
            title="📦 Bundle Listed on eBay",
            description=f"[{title}]({listing_url})",
            colour=discord.Colour.green(),
        )

        item_lines = []
        for item in items:
            item_market = float(item.get("live_price") or 0)
            if has_zero_price or combined_market <= 0:
                item_share = round(bundle_price / len(items), 2)
            else:
                item_share = round(bundle_price * item_market / combined_market, 2)
            item_lines.append(
                f"• #{item['item_id']} {item['card_name'][:30]}"
                f"  market £{item_market:.2f} → share £{item_share:.2f}"
            )

        embed.add_field(name=f"Items ({len(items)})",  value="\n".join(item_lines),                         inline=False)
        embed.add_field(name="Combined market",         value=f"£{combined_market:.2f}",                    inline=True)
        price_label = f"£{bundle_price:.2f}  (override)" if price_override > 0 else f"£{bundle_price:.2f}  (-{discount_pct:.0f}%)"
        embed.add_field(name="Bundle price",            value=price_label, inline=True)
        embed.add_field(name="Combined cost",           value=f"£{combined_cost:.2f}",                      inline=True)
        embed.add_field(name="Bundle profit",           value=f"**+£{bundle_profit:.2f}**  (ROI: {roi_pct}%)", inline=True)

        audit.log_mutation("listbundle", None, "bundle_listed", {
            "item_ids":      ids,
            "listing_id":    listing_id,
            "bundle_price":  bundle_price,
            "combined_cost": combined_cost,
            "profit":        bundle_profit,
            "sku":           sku,
        })

        await interaction.followup.send(embed=embed)

    finally:
        if temp_dir.exists():
            await asyncio.to_thread(shutil.rmtree, temp_dir, ignore_errors=True)


# ---------------------------------------------------------------------------
# /delist
# ---------------------------------------------------------------------------

@bot.tree.command(name="delist", description="End one or more eBay listings. Use commas: item_ids:45,46,47")
@app_commands.describe(item_ids="Item ID(s), comma-separated for multiple")
async def slash_delist(interaction: discord.Interaction, item_ids: str) -> None:
    # Parse item IDs
    id_tokens = [t.strip() for t in item_ids.split(",")]
    parsed_ids: list[int] = []
    for tok in id_tokens:
        if not tok.lstrip("-").isdigit() or int(tok) <= 0:
            await interaction.response.send_message(
                f"`{tok}` is not a valid item ID — item IDs must be positive integers.",
                ephemeral=True,
            )
            return
        parsed_ids.append(int(tok))

    await interaction.response.defer()

    ended: list[dict] = []
    skipped: list[tuple[int, str]] = []
    failed: list[tuple[int, str]] = []

    for iid in parsed_ids:
        try:
            item = await excel_db.get_item_async(iid)
        except ValueError as exc:
            failed.append((iid, f"Item not found: {exc}"))
            continue

        listing_id = item.get("ebay_listing_id") or ""
        if item.get("ebay_listed") != "Yes" or not listing_id:
            skipped.append((iid, item.get("card_name", "Unknown")))
            continue

        try:
            await lister_ebay.end_ebay_listing(listing_id)
        except Exception as exc:
            failed.append((iid, f"Failed to end listing `{listing_id}`: {exc}"))
            continue

        try:
            await excel_db.mark_ebay_delisted_async(iid)
            audit.log_mutation("delist", iid, "delisted from eBay", {
                "ebay_listing_id": listing_id,
            })
            if _SYNC_ENABLED:
                try:
                    bot_sync.sync_ebay_delisted(iid)
                except Exception as exc:
                    print(f"[bot_sync] ebay delisted sync failed for item {iid}: {exc}")
        except Exception as exc:
            print(f"[delist] Failed to update inventory after delisting item {iid}: {exc}")

        ended.append({"id": iid, "card_name": item["card_name"], "listing_id": listing_id})

    if len(parsed_ids) == 1 and ended:
        r = ended[0]
        embed = discord.Embed(
            title=f"eBay Listing Ended — Item #{r['id']}",
            colour=discord.Colour.orange(),
        )
        embed.add_field(name="Card",               value=r["card_name"],                                        inline=False)
        embed.add_field(name="eBay listing ended",  value=f"https://www.ebay.co.uk/itm/{r['listing_id']}",      inline=False)
        embed.add_field(name="Status",              value="Item returned to inventory (no longer listed)",        inline=False)
        await interaction.followup.send(embed=embed)
        return

    embed = discord.Embed(
        title=f"Bulk Delist — {len(ended)} ended",
        colour=discord.Colour.orange(),
    )

    if ended:
        ended_lines = "\n".join(
            f"• ID {r['id']} — {r['card_name']} ([listing](https://www.ebay.co.uk/itm/{r['listing_id']}))"
            for r in ended
        )
        embed.add_field(name="Listings Ended", value=ended_lines, inline=False)

    if skipped:
        skip_lines = "\n".join(f"• ID {iid} — {name}: not listed on eBay" for iid, name in skipped)
        embed.add_field(name="Skipped (not listed)", value=skip_lines, inline=False)

    if failed:
        fail_lines = "\n".join(f"• Item {iid}: {msg}" for iid, msg in failed)
        embed.add_field(name="Failures", value=fail_lines, inline=False)

    await interaction.followup.send(embed=embed)


# ---------------------------------------------------------------------------
# /cancelled
# ---------------------------------------------------------------------------

@bot.tree.command(name="cancelled", description="Check for cancelled eBay orders and revert inventory.")
async def slash_cancelled(interaction: discord.Interaction) -> None:
    await interaction.response.defer()

    try:
        access_token = await lister_ebay._get_access_token()
        orders = await _get_recent_orders(access_token)
        cancelled_orders = [
            o for o in orders
            if (o.get("cancelStatus", {}) or {}).get("cancelState") == "CANCEL_COMPLETE"
        ]

        if not cancelled_orders:
            await interaction.followup.send("No cancelled orders found in the last 30 days.")
            return

        embed = discord.Embed(
            title=f"{len(cancelled_orders)} Cancelled Order(s) Found",
            colour=discord.Colour.orange(),
        )
        for order in cancelled_orders[:5]:
            order_id = order.get("orderId", "")
            titles = [li.get("title", "Unknown") for li in order.get("lineItems", [])]
            embed.add_field(name=f"Order {order_id}", value="\n".join(titles[:3]) or "—", inline=False)
        await interaction.followup.send(embed=embed)

        reverted = await _check_cancelled_orders(cancelled_orders)
        if reverted:
            names = ", ".join(f"#{r['item_id']} {r['card_name']}" for r in reverted)
            await interaction.followup.send(f"Reverted to Inventory: {names}")
    except Exception as exc:
        await interaction.followup.send(f"Error checking cancelled orders: {exc}")


# ---------------------------------------------------------------------------
# /update
# ---------------------------------------------------------------------------

@bot.tree.command(name="update", description="Manually refresh live prices for all inventory items.")
async def slash_update(interaction: discord.Interaction) -> None:
    await interaction.response.defer()

    channel_id = interaction.channel_id
    interaction_expired = False

    items = await excel_db.get_stock_async(STATUS_INVENTORY)
    if not items:
        await interaction.followup.send("Nothing to update — inventory is empty.")
        return

    await backups.snapshot_inventory("pre-manual-update")

    status_msg = await interaction.followup.send(
        f"🔄 Refreshing prices for {len(items)} item(s)…"
    )

    async def _progress(checked: int, updated: int) -> None:
        nonlocal interaction_expired
        if interaction_expired:
            return
        try:
            await status_msg.edit(
                content=f"🔄 Refreshing prices… ({checked}/{len(items)} done, {updated} updated)"
            )
        except discord.errors.HTTPException as exc:
            if exc.code == 50027:
                interaction_expired = True
                print("[bot] Interaction token expired — progress silenced, work continues")
        except Exception:
            pass

    (checked, updated, skipped, big_shifts, aborted, abort_reason,
     invalid_url_count, underwater_items, ebay_sync_stats) = (
        await _run_price_refresh(sleep_range=(3.0, 8.0), progress_callback=_progress)
    )

    try:
        update_summary = await excel_db.get_summary_async()
    except Exception:
        update_summary = None

    try:
        update_unverified = await excel_db.get_unverified_price_changes_async()
        update_pending_verification = len(update_unverified)
    except Exception:
        update_pending_verification = 0

    embed = _build_update_embed(
        checked, updated, skipped, big_shifts, aborted, abort_reason,
        title_prefix="Manual",
        summary=update_summary,
        invalid_url_count=invalid_url_count,
        underwater_items=underwater_items,
        ebay_sync_stats=ebay_sync_stats,
        pending_verification=update_pending_verification,
    )

    if interaction_expired:
        channel = bot.get_channel(channel_id)
        if channel:
            await channel.send(embed=embed)
        else:
            print("[bot] Interaction expired and channel not found — update result lost")
    else:
        await status_msg.edit(content=None, embed=embed)


# ---------------------------------------------------------------------------
# /updatelistings — helpers
# ---------------------------------------------------------------------------

async def _calculate_new_listing_price(
    item: dict,
    strategy: str,
    custom_price: float = 0.0,
) -> tuple[float, str, bool]:
    """
    Calculate the new eBay listing price for an item.
    Returns (new_price, reasoning, was_floored).
    was_floored is True when the profit floor raised the price above the raw calculation.
    """
    purchase_price = float(item.get("purchase_price") or 0)
    live_price     = float(item.get("live_price") or 0)
    stored_quick   = float(item.get("quick_price") or 0)
    min_profitable = max(round(purchase_price * 1.10, 2), config.EBAY_MIN_PRICE_GBP)

    if strategy == "custom":
        if custom_price <= 0:
            raise ValueError("custom_price must be > 0 when using custom strategy")
        raw = round(custom_price, 2)
        if raw < min_profitable:
            return min_profitable, f"custom £{raw:.2f} raised to floor £{min_profitable:.2f}", True
        return raw, f"custom price £{raw:.2f}", False

    elif strategy == "quicksell":
        # 1. Use stored quick_price if it covers the floor
        if stored_quick >= min_profitable:
            return stored_quick, f"quick sell £{stored_quick:.2f} (cached)", False

        # 2. Fetch fresh competitor data
        try:
            comp = await lister_ebay.get_competitor_price(
                item["card_name"], item.get("condition") or ""
            )
            if comp:
                raw = comp["quick_sell"]
                note = comp.get("strategy_note", "competitor-aware")
                if raw >= min_profitable:
                    return raw, f"£{raw:.2f} ({note})", False
                return min_profitable, (
                    f"floor £{min_profitable:.2f} (quick sell £{raw:.2f} below cost — {note})"
                ), True
        except Exception:
            pass

        # 3. Fallback: 95% of live price
        if live_price > 0:
            raw = round(live_price * 0.95, 2)
            floored = raw < min_profitable
            final = max(raw, min_profitable)
            return final, f"fallback £{final:.2f} (95% of market, no competitor data)", floored

        return min_profitable, f"floor £{min_profitable:.2f} (no price data)", True

    else:  # market
        if live_price <= 0:
            return min_profitable, f"floor £{min_profitable:.2f} (no live price)", True
        raw = round(live_price * 1.15, 2)
        floored = raw < min_profitable
        final = max(raw, min_profitable)
        return final, f"market +15% = £{final:.2f}", floored


def _build_updatelistings_embed(
    mode_name: str,
    strategy: str,
    updated: list,
    not_api_managed: list,
    failed: list,
    floored_count: int,
    dry_run: bool,
) -> discord.Embed:
    strategy_labels = {
        "quicksell": "quick sell (beat competitors)",
        "market":    "market +15%",
        "custom":    "custom price",
    }
    strategy_label = strategy_labels.get(strategy, strategy)

    if dry_run:
        title  = f"🔍 Dry Run — Price Update Preview ({mode_name})"
        colour = discord.Colour.blue()
    elif failed and not updated:
        title  = f"❌ eBay Price Update Failed ({mode_name})"
        colour = discord.Colour.red()
    elif failed:
        title  = f"⚠️ eBay Listings Updated with Errors ({mode_name})"
        colour = discord.Colour.orange()
    else:
        title  = f"✅ eBay Listings Updated ({mode_name})"
        colour = discord.Colour.green()

    embed = discord.Embed(title=title, colour=colour)
    embed.add_field(name="Strategy",        value=strategy_label,          inline=True)
    embed.add_field(name="Updated",         value=str(len(updated)),       inline=True)
    embed.add_field(name="Not API-managed", value=str(len(not_api_managed)), inline=True)
    if failed:
        embed.add_field(name="Failed", value=str(len(failed)), inline=True)

    if updated:
        lines = []
        for item_id, card_name, new_price, reasoning, _ in updated[:10]:
            lines.append(f"• #{item_id} {card_name[:26]:<26} → £{new_price:.2f}  {reasoning[:35]}")
        if len(updated) > 10:
            lines.append(f"…and {len(updated) - 10} more")
        field_label = "Would update" if dry_run else "Price changes"
        embed.add_field(name=f"{field_label} ({len(updated)})", value="\n".join(lines), inline=False)

    if not_api_managed:
        nam_lines = [f"  #{iid} {name[:40]}" for iid, name in not_api_managed[:5]]
        if len(not_api_managed) > 5:
            nam_lines.append(f"  …and {len(not_api_managed) - 5} more")
        embed.add_field(
            name=f"Not API-managed ({len(not_api_managed)}) — revise manually on eBay",
            value="\n".join(nam_lines),
            inline=False,
        )

    if failed:
        fail_lines = [f"  #{iid} {name[:28]}: {err[:60]}" for iid, name, err in failed[:5]]
        if len(failed) > 5:
            fail_lines.append(f"  …and {len(failed) - 5} more")
        embed.add_field(name=f"Failed ({len(failed)})", value="\n".join(fail_lines), inline=False)

    if floored_count > 0:
        embed.add_field(
            name="💡 Profit floor applied",
            value=f"{floored_count} listing(s) raised to min 10% above cost",
            inline=False,
        )

    if dry_run:
        embed.set_footer(text="Dry run — no changes made. Remove dry_run:True to apply.")

    return embed


# ---------------------------------------------------------------------------
# /updatelistings
# ---------------------------------------------------------------------------

@bot.tree.command(
    name="updatelistings",
    description="Update eBay listing prices to quick-sell or market price.",
)
@app_commands.describe(
    mode="all = all active listings, specific = named items, underwater = items below cost",
    price_strategy="quicksell = beat competitors, market = PC +15%, custom = fixed price",
    item_ids="Comma-separated item IDs (required for mode:specific)",
    custom_price="Fixed price in GBP (required for price_strategy:custom)",
    dry_run="Preview changes without touching eBay",
)
@app_commands.choices(mode=[
    app_commands.Choice(name="All listings",    value="all"),
    app_commands.Choice(name="Specific items",  value="specific"),
    app_commands.Choice(name="Underwater only", value="underwater"),
])
@app_commands.choices(price_strategy=[
    app_commands.Choice(name="Quick sell (beat competitors)", value="quicksell"),
    app_commands.Choice(name="Market price (+15%)",           value="market"),
    app_commands.Choice(name="Custom price",                  value="custom"),
])
async def slash_updatelistings(
    interaction: discord.Interaction,
    mode: app_commands.Choice[str],
    price_strategy: app_commands.Choice[str],
    item_ids: str = "",
    custom_price: float = 0.0,
    dry_run: bool = False,
) -> None:
    await interaction.response.defer()

    strategy = price_strategy.value

    # Validate custom strategy early
    if strategy == "custom" and custom_price <= 0:
        await interaction.followup.send(
            "❌ `custom_price` must be > 0 when using the custom strategy.", ephemeral=True
        )
        return

    # ── Step 1: build item list ───────────────────────────────────────────────
    all_inventory = await excel_db.get_stock_async("Inventory")
    active_listed = [
        item for item in all_inventory
        if item.get("ebay_listed") == "Yes"
        and excel_db.is_valid_ebay_listing_id(str(item.get("ebay_listing_id") or ""))
    ]

    mode_value = mode.value

    if mode_value == "all":
        items_to_process = active_listed

    elif mode_value == "underwater":
        items_to_process = [
            item for item in active_listed
            if (item.get("live_price") is not None and item.get("purchase_price") is not None
                and float(item["live_price"]) < float(item["purchase_price"]))
        ]

    else:  # specific
        if not item_ids.strip():
            await interaction.followup.send(
                "❌ Provide `item_ids` when using mode:specific.", ephemeral=True
            )
            return
        parsed_ids: list[int] = []
        for tok in item_ids.split(","):
            tok = tok.strip()
            if not tok.isdigit() or int(tok) <= 0:
                await interaction.followup.send(
                    f"❌ `{tok}` is not a valid item ID.", ephemeral=True
                )
                return
            parsed_ids.append(int(tok))

        listed_by_id = {item["item_id"]: item for item in active_listed}
        items_to_process = []
        for iid in parsed_ids:
            if iid not in listed_by_id:
                await interaction.followup.send(
                    f"❌ Item #{iid} not found or has no active eBay listing.", ephemeral=True
                )
                return
            items_to_process.append(listed_by_id[iid])

    if not items_to_process:
        await interaction.followup.send(
            f"No active eBay listings found for mode **{mode.name}**."
        )
        return

    # ── Step 2: status message ────────────────────────────────────────────────
    dry_prefix = "🔍 DRY RUN — " if dry_run else ""
    status_msg = await interaction.followup.send(
        f"{dry_prefix}Calculating prices for {len(items_to_process)} listing(s)…"
    )

    # ── Step 3: process each item ─────────────────────────────────────────────
    updated:         list[tuple] = []  # (item_id, card_name, new_price, reasoning, was_floored)
    not_api_managed: list[tuple] = []  # (item_id, card_name)
    failed:          list[tuple] = []  # (item_id, card_name, error)

    for idx, item in enumerate(items_to_process):
        item_id    = item["item_id"]
        card_name  = item["card_name"]
        listing_id = str(item["ebay_listing_id"]).strip()

        # Calculate new price
        try:
            new_price, reasoning, was_floored = await _calculate_new_listing_price(
                item, strategy, custom_price
            )
        except Exception as exc:
            failed.append((item_id, card_name, str(exc)))
            await asyncio.sleep(0.5)
            continue

        # Dry run: record what would happen and move on (no API calls)
        if dry_run:
            updated.append((item_id, card_name, new_price, reasoning, was_floored))
            await asyncio.sleep(0.5)
            continue

        # Attempt the price update — Inventory API via SKU/listing_id lookup,
        # falling back to the Trading API for manually-created listings
        success = await lister_ebay.revise_listing_price(
            listing_id, new_price, sku=f"pokemaz-{item_id}"
        )
        if success:
            updated.append((item_id, card_name, new_price, reasoning, was_floored))
            try:
                await excel_db.update_sell_price_async(item_id, new_price)
            except Exception:
                pass
            audit.log_mutation("update_listing_price", item_id, "ebay_price_updated", {
                "new_price":  new_price,
                "strategy":   strategy,
                "reasoning":  reasoning,
                "listing_id": listing_id,
            })
            print(
                f"[updatelistings] #{item_id} {card_name[:30]}: "
                f"→ £{new_price:.2f}  ({reasoning})"
            )
        else:
            not_api_managed.append((item_id, card_name))

        await asyncio.sleep(0.5)

        # Progress updates every 5 items
        if (idx + 1) % 5 == 0:
            done = len(updated) + len(not_api_managed) + len(failed)
            try:
                await status_msg.edit(
                    content=(
                        f"{dry_prefix}Processing {done}/{len(items_to_process)}… "
                        f"({len(updated)} updated, {len(not_api_managed)} not API-managed)"
                    )
                )
            except Exception:
                pass

    # ── Step 4: build result embed ────────────────────────────────────────────
    floored_count = sum(1 for *_, was_floored in updated if was_floored)
    result_embed  = _build_updatelistings_embed(
        mode.name, strategy, updated, not_api_managed, failed, floored_count, dry_run
    )

    # Post to price update channel for visibility; ack in interaction
    post_channel = (
        bot.get_channel(config.PRICE_UPDATE_CHANNEL_ID)
        if config.PRICE_UPDATE_CHANNEL_ID else None
    )
    if post_channel and post_channel.id != interaction.channel_id and not dry_run:
        try:
            await post_channel.send(embed=result_embed)
        except Exception:
            pass
        ack = discord.Embed(
            description=(
                f"✅ Done — {len(updated)} updated, "
                f"{len(not_api_managed)} not API-managed, "
                f"{len(failed)} failed. Full results posted to <#{post_channel.id}>."
            ),
            colour=discord.Colour.green() if not failed else discord.Colour.orange(),
        )
        await status_msg.edit(content=None, embed=ack)
    else:
        await status_msg.edit(content=None, embed=result_embed)


# ---------------------------------------------------------------------------
# /verify
# ---------------------------------------------------------------------------

@bot.tree.command(name="verify", description="Review and confirm price changes from the last update.")
@app_commands.describe(
    mode="all = confirm all changes, single = verify one item, underwater = show underwater only",
    item_id="Item ID to verify individually (required for mode:single)",
)
@app_commands.choices(mode=[
    app_commands.Choice(name="Verify all",            value="all"),
    app_commands.Choice(name="Single item",           value="single"),
    app_commands.Choice(name="Verify underwater only", value="underwater"),
])
async def slash_verify(
    interaction: discord.Interaction,
    mode: app_commands.Choice[str],
    item_id: int = 0,
) -> None:
    await interaction.response.defer()

    if mode.value == "all":
        unverified = await excel_db.get_unverified_price_changes_async()
        if not unverified:
            await interaction.followup.send("✅ No pending verifications.")
            return
        ids = [it["item_id"] for it in unverified]
        await excel_db.set_price_verified_bulk_async(ids)
        await interaction.followup.send(f"✅ Verified {len(ids)} price change(s).")

    elif mode.value == "underwater":
        unverified = await excel_db.get_unverified_price_changes_async()
        underwater = [
            it for it in unverified
            if it.get("live_price") is not None
            and it.get("purchase_price") is not None
            and it["live_price"] < it["purchase_price"]
        ]
        if not underwater:
            await interaction.followup.send("✅ No unverified underwater items.")
            return
        ids = [it["item_id"] for it in underwater]
        lines = []
        for it in underwater[:20]:
            diff = it["purchase_price"] - it["live_price"]
            lines.append(
                f"• #{it['item_id']} {it['card_name'][:35]} — "
                f"bought £{it['purchase_price']:.2f}, now £{it['live_price']:.2f} (-£{diff:.2f})"
            )
        await excel_db.set_price_verified_bulk_async(ids)
        embed = discord.Embed(
            title=f"Verified {len(ids)} underwater item(s)",
            description="\n".join(lines),
            colour=discord.Colour.orange(),
        )
        await interaction.followup.send(embed=embed)

    elif mode.value == "single":
        if item_id <= 0:
            await interaction.followup.send("❌ Provide `item_id` when using `mode:single`.")
            return
        try:
            item = await excel_db.get_item_async(item_id)
        except ValueError:
            await interaction.followup.send(f"❌ Item {item_id} not found.")
            return

        history = await excel_db.get_item_price_history_async(item_id, limit=2)
        prev_price = history[-2]["live_price_gbp"] if len(history) >= 2 else None

        already_verified = item.get("price_verified", True)
        status_str = "✅ Already verified" if already_verified else "⚠️ Unverified"
        embed = discord.Embed(
            title=f"Item #{item_id} — {item['card_name']}",
            colour=discord.Colour.blurple(),
        )
        embed.add_field(name="Verification", value=status_str, inline=True)
        live_str = f"£{item['live_price']:.2f}" if item.get("live_price") is not None else "N/A"
        embed.add_field(name="Current Price", value=live_str, inline=True)
        if prev_price is not None:
            embed.add_field(name="Previous Price", value=f"£{prev_price:.2f}", inline=True)
        embed.add_field(name="Purchase Price", value=f"£{item['purchase_price']:.2f}", inline=True)

        if not already_verified:
            await excel_db.set_price_verified_bulk_async([item_id])
            embed.set_footer(text="Marked as verified.")
        else:
            embed.set_footer(text="No action needed — already verified.")

        await interaction.followup.send(embed=embed)


# ---------------------------------------------------------------------------
# /stock
# ---------------------------------------------------------------------------

@bot.tree.command(name="stock", description="Show all cards currently in inventory.")
async def slash_stock(interaction: discord.Interaction) -> None:
    await interaction.response.defer()

    try:
        items = await excel_db.get_stock_async(STATUS_INVENTORY)
    except Exception as exc:
        await interaction.followup.send(f"Error reading inventory: {exc}")
        return

    if not items:
        await interaction.followup.send("No items currently in inventory.")
        return

    header = f"{'ID':>4}  {'Condition':<14}  {'Buy':>7}  {'Market':>7}  {'Quick':>8}  Card"
    sep    = "-" * 80

    _effort_hours = config.EFFORT_MINUTES_PER_CARD / 60

    def _row(item: dict) -> str:
        live           = f"£{item['live_price']:.2f}" if item["live_price"] else "   N/A"
        listed_tag     = " 🏪" if item.get("ebay_listed") == "Yes" else ""
        unverified_tag = " ⚠️" if not item.get("price_verified", True) else ""
        quick = item.get("quick_price")
        if quick is None:
            quick_str = "      —"
        elif item["live_price"] and quick < item["live_price"]:
            quick_str = f"£{quick:.2f} ↓"
        else:
            quick_str = f"£{quick:.2f} "
        est_price      = float(quick or item.get("live_price") or 0)
        est_profit     = est_price - float(item.get("purchase_price") or 0)
        low_efficiency = _effort_hours > 0 and (est_profit / _effort_hours) < config.HOURLY_RATE_GBP
        efficiency_tag = " ⚡" if low_efficiency else ""
        return (
            f"{item['item_id']:>4}  {(item['condition'] or ''):.<14}  "
            f"£{item['purchase_price']:>6.2f}  {live:>7}  {quick_str:>8}  "
            f"{item['card_name']}{listed_tag}{unverified_tag}{efficiency_tag}"
        )

    try:
        s = await excel_db.get_summary_async()
    except Exception:
        s = {}

    def _totals_footer(s: dict) -> str:
        div = "─" * 41
        return (
            f"{div}\n"
            f"Total cost (in stock):         £{s.get('total_cost_in_stock', 0):.2f}\n"
            f"Total potential value:          £{s.get('total_potential_in_stock', 0):.2f}\n"
            f"Total potential profit:         £{s.get('total_potential_profit_in_stock', 0):.2f}\n"
            f"Total sold revenue (lifetime): £{s.get('total_sold_revenue', 0):.2f}"
        )

    footer = _totals_footer(s)

    lines = ["**Current Inventory**", "```", header, sep]
    lines.extend(_row(i) for i in items)
    lines.extend(["", footer, "```"])
    message = "\n".join(lines)

    if len(message) <= 2000:
        await interaction.followup.send(message)
        return

    # Chunk for large inventories
    chunk = ["**Current Inventory**", "```", header, sep]
    for item in items:
        line = _row(item)
        if sum(len(ln) + 1 for ln in chunk) + len(line) > 1900:
            chunk.append("```")
            await interaction.followup.send("\n".join(chunk))
            chunk = ["```", line]
        else:
            chunk.append(line)
    chunk.extend(["", footer, "```"])
    await interaction.followup.send("\n".join(chunk))


# ---------------------------------------------------------------------------
# /search
# ---------------------------------------------------------------------------

@bot.tree.command(name="search", description="Search inventory by name, condition, or region.")
@app_commands.describe(query="Search term — card name, set, condition, or region (e.g. 'charizard', 'PSA 10', 'KR')")
async def slash_search(interaction: discord.Interaction, query: str) -> None:
    await interaction.response.defer()

    all_items = await excel_db.get_stock_async(status_filter=None)
    q = query.lower()
    matches = [
        item for item in all_items
        if q in (item.get("card_name") or "").lower()
        or q in (item.get("condition") or "").lower()
        or q in (item.get("region") or "").lower()
    ]

    if not matches:
        await interaction.followup.send(f"No items found matching '{query}'.")
        return

    display   = matches[:20]
    truncated = len(matches) > 20

    _STATUS_ICON = {"Inventory": "🟢", "Sold": "✅"}

    def _row(item: dict) -> str:
        icon  = _STATUS_ICON.get(item.get("status", ""), "❓")
        live  = item.get("live_price")
        quick = item.get("quick_price")
        sell  = item.get("sell_price")
        buy   = item.get("purchase_price") or 0

        live_str  = f"£{live:.2f}" if live else "  N/A "
        quick_str = f"£{quick:.2f}" if quick else "  —   "

        underwater = (
            item.get("status") == "Inventory"
            and live is not None and buy > 0 and live < buy
        )
        uw_tag = "  ⚠️ UNDERWATER" if underwater else ""
        sold_str = f"  (sold £{sell:.2f})" if sell else ""

        return (
            f"#{item['item_id']:<4} {icon}  {item.get('card_name', '')[:35]:<35}  "
            f"Market: {live_str}  Quick: {quick_str}  Bought: £{buy:.2f}"
            f"{sold_str}{uw_tag}"
        )

    header = f"#ID   St  {'Card':<35}  Market    Quick     Bought"
    sep    = "-" * 82
    lines  = [f"**Search: '{query}'**", "```", header, sep]
    lines.extend(_row(i) for i in display)
    if truncated:
        lines.append(f"\n(Showing 20 of {len(matches)} — refine your search)")
    lines.append("```")

    message = "\n".join(lines)
    if len(message) <= 2000:
        await interaction.followup.send(message)
        return

    # Chunk for large results
    chunk = [f"**Search: '{query}'**", "```", header, sep]
    for item in display:
        line = _row(item)
        if sum(len(ln) + 1 for ln in chunk) + len(line) > 1900:
            chunk.append("```")
            await interaction.followup.send("\n".join(chunk))
            chunk = ["```", line]
        else:
            chunk.append(line)
    if truncated:
        chunk.append(f"\n(Showing 20 of {len(matches)} — refine your search)")
    chunk.append("```")
    await interaction.followup.send("\n".join(chunk))


# ---------------------------------------------------------------------------
# /price
# ---------------------------------------------------------------------------

@bot.tree.command(name="price", description="Check market and quick-sell price for an item.")
@app_commands.describe(item_id="Item ID to check")
async def slash_price(interaction: discord.Interaction, item_id: int) -> None:
    await interaction.response.defer()

    try:
        item = await excel_db.get_item_async(item_id)
    except ValueError:
        await interaction.followup.send(f"❌ Item #{item_id} not found.", ephemeral=True)
        return

    card_name      = item["card_name"]
    condition      = item.get("condition") or "ungraded"
    purchase_price = float(item.get("purchase_price") or 0)
    pc_url         = item.get("pc_url") or ""

    # Fetch live PriceCharting price
    pc_price: float | None = None
    try:
        pc_price = await _scrape_fresh_price(pc_url, card_name, condition, item.get("region") or "")
    except Exception:
        pc_price = item.get("live_price")  # fall back to stored price

    # Fetch eBay competitor data
    comp: dict | None = None
    try:
        comp = await lister_ebay.get_competitor_price(card_name, condition)
    except Exception:
        pass

    embed = discord.Embed(
        title=f"💰 Price Check — {card_name[:50]} (#{item_id})",
        colour=discord.Colour.blurple(),
    )

    # PriceCharting section
    if pc_price is not None:
        embed.add_field(name="PriceCharting", value=f"£{pc_price:.2f}  ({condition})", inline=True)
    else:
        embed.add_field(name="PriceCharting", value="N/A (no data)", inline=True)
    embed.add_field(name="Your cost", value=f"£{purchase_price:.2f}", inline=True)
    embed.add_field(name="​", value="​", inline=True)

    # eBay competitors section
    if comp:
        embed.add_field(
            name="eBay competitors",
            value=(
                f"{comp['count']} active listings\n"
                f"Lowest:  £{comp['lowest']:.2f}\n"
                f"Median:  £{comp['median']:.2f}"
            ),
            inline=True,
        )

    # Price recommendations
    market_price = round(pc_price * 1.15, 2) if pc_price else None
    if comp:
        min_profitable = max(round(purchase_price * 1.10, 2), config.EBAY_MIN_PRICE_GBP)
        quick_price = max(comp["quick_sell"], min_profitable)
    else:
        quick_price = None

    price_lines = []
    if quick_price is not None:
        price_lines.append(f"Quick sell:   £{quick_price:.2f}")
    if market_price is not None:
        price_lines.append(f"Market price: £{market_price:.2f}  (15% above PC)")
    if price_lines:
        embed.add_field(name="Recommended prices", value="\n".join(price_lines), inline=False)

    if comp and comp.get("strategy_note"):
        embed.add_field(
            name="Quick sell strategy",
            value=f"_{comp['strategy_note']}_",
            inline=False,
        )

    # ROI section
    roi_lines = []
    if quick_price is not None and purchase_price > 0:
        profit_q = round(quick_price - purchase_price, 2)
        roi_q    = round(profit_q / purchase_price * 100, 0)
        sign_q   = "+" if profit_q >= 0 else ""
        roi_lines.append(f"At quick sell:   {sign_q}£{profit_q:.2f}  (ROI: {roi_q:.0f}%)")
    if market_price is not None and purchase_price > 0:
        profit_m = round(market_price - purchase_price, 2)
        roi_m    = round(profit_m / purchase_price * 100, 0)
        sign_m   = "+" if profit_m >= 0 else ""
        roi_lines.append(f"At market price: {sign_m}£{profit_m:.2f}  (ROI: {roi_m:.0f}%)")
    if roi_lines:
        embed.add_field(name="Potential profit", value="\n".join(roi_lines), inline=False)

    if pc_url:
        embed.add_field(name="PriceCharting", value=pc_url, inline=False)

    await interaction.followup.send(embed=embed)


# ---------------------------------------------------------------------------
# /summary
# ---------------------------------------------------------------------------

def _calculate_forecast(sold_items: list[dict], inventory_count: int) -> dict:
    """
    Returns forecast based on items sold in the last 90 days.
    sold_items must have keys: date_added, date_sold, profit (str dates YYYY-MM-DD).
    Returns {"insufficient_data": True} if fewer than 3 qualifying items.
    """
    cutoff = datetime.now() - timedelta(days=90)
    recent: list[dict] = []
    for item in sold_items:
        date_added = str(item.get("date_added", "") or "")[:10]
        date_sold  = str(item.get("date_sold", "") or "")[:10]
        if not date_added or not date_sold:
            continue
        sell_price     = float(item.get("sell_price")     or 0)
        purchase_price = float(item.get("purchase_price") or 0)
        if sell_price <= 0 or purchase_price <= 0:
            continue
        try:
            dt_added = datetime.strptime(date_added, "%Y-%m-%d")
            dt_sold  = datetime.strptime(date_sold,  "%Y-%m-%d")
        except ValueError:
            continue
        if dt_sold < cutoff:
            continue
        days   = max((dt_sold - dt_added).days, 1)
        profit = sell_price - purchase_price
        recent.append({"days": days, "profit": profit, "sell_price": sell_price})

    if len(recent) < 3:
        return {"insufficient_data": True, "sample_size": len(recent)}

    avg_days        = sum(i["days"]   for i in recent) / len(recent)
    avg_profit      = sum(i["profit"] for i in recent) / len(recent)
    avg_sell        = sum(i["sell_price"] for i in recent) / len(recent)
    avg_margin_pct  = (avg_profit / avg_sell * 100) if avg_sell else 0
    proj_items      = round(30 / avg_days) if avg_days > 0 else 0
    proj_profit     = round(avg_profit * proj_items, 2)

    return {
        "insufficient_data":    False,
        "avg_days_to_sell":     round(avg_days, 1),
        "avg_profit_per_item":  round(avg_profit, 2),
        "avg_margin_pct":       round(avg_margin_pct, 1),
        "projected_30d_items":  proj_items,
        "projected_30d_profit": proj_profit,
        "sample_size":          len(recent),
    }


def _count_ebay_auto_sales_last_30_days() -> int:
    """Count audit log entries where command == 'ebay_auto_sell' in the last 30 days."""
    cutoff = datetime.now() - timedelta(days=30)
    count = 0
    log_dir = Path(config.AUDIT_LOG_DIR)
    for log_file in sorted(log_dir.glob("audit-*.jsonl")):
        try:
            for line in log_file.read_text(encoding="utf-8").splitlines():
                if not line.strip():
                    continue
                try:
                    entry = json.loads(line)
                except json.JSONDecodeError:
                    continue
                if entry.get("command") != "ebay_auto_sell":
                    continue
                try:
                    ts = datetime.fromisoformat(entry["ts"])
                    if ts >= cutoff:
                        count += 1
                except (KeyError, ValueError):
                    pass
        except Exception:
            pass
    return count


@bot.tree.command(name="summary", description="Show aggregate profit/loss across all sold items.")
async def slash_summary(interaction: discord.Interaction) -> None:
    await interaction.response.defer()

    try:
        s = await excel_db.get_summary_async()
    except Exception as exc:
        await interaction.followup.send(f"Error: {exc}")
        return

    sold_items = await excel_db.get_stock_async(status_filter="Sold")
    forecast   = _calculate_forecast(sold_items, s["in_stock"])

    profit = s["total_profit"]
    colour = discord.Colour.green() if profit >= 0 else discord.Colour.red()

    today      = date.today()
    month_name = today.strftime("%B %Y")

    ebay_auto_count = await asyncio.to_thread(_count_ebay_auto_sales_last_30_days)

    roi_sign = "+" if s["roi_pct"] >= 0 else ""

    embed = discord.Embed(title="Inventory Summary", colour=colour)
    embed.add_field(name="In Inventory",   value=str(s["in_stock"]),            inline=True)
    embed.add_field(name="Sold",           value=str(s["sold"]),                inline=True)
    embed.add_field(name="​",              value="​",                           inline=True)
    embed.add_field(name="Total Invested", value=f"£{s['total_invested']:.2f}", inline=True)
    embed.add_field(name="Total Revenue",  value=f"£{s['total_revenue']:.2f}",  inline=True)
    embed.add_field(
        name="Net Profit",
        value=f"**£{profit:.2f}** ({roi_sign}{s['roi_pct']}% ROI on lifetime spend)",
        inline=True,
    )
    embed.add_field(
        name="Avg Margin",
        value=(
            f"**{s['avg_margin_pct']}%**   "
            f"Avg sell: £{s['avg_sell_price']:.2f}   "
            f"Avg profit/item: £{s['avg_profit']:.2f}"
        ),
        inline=False,
    )
    embed.add_field(name="Cost in Stock",        value=f"£{s['total_cost_in_stock']:.2f}",             inline=True)
    embed.add_field(name="Potential in Stock",   value=f"£{s['total_potential_in_stock']:.2f}",         inline=True)
    embed.add_field(
        name="Potential Profit in Stock",
        value=f"**£{s['total_potential_profit_in_stock']:.2f}** ({s['stock_roi_pct']}% ROI if all sold)",
        inline=True,
    )
    embed.add_field(name="eBay Auto-Sales (last 30 days)", value=str(ebay_auto_count), inline=True)
    # Lifetime section
    lifetime_potential_sign = "+" if s["total_lifetime_potential"] >= 0 else ""
    embed.add_field(
        name="Lifetime",
        value=(
            f"Total spent (ever):    £{s['total_lifetime_cost']:.2f}\n"
            f"Total realised profit: **£{s['total_lifetime_profit']:.2f}**\n"
            f"Best-case total return: **{lifetime_potential_sign}£{s['total_lifetime_potential']:.2f}**"
        ),
        inline=False,
    )
    # Month-to-date section
    mtd_profit_sign = "+" if s["mtd_profit"] >= 0 else ""
    embed.add_field(
        name=f"This Month ({month_name})",
        value=(
            f"Sold: **{s['mtd_sold_count']}**  "
            f"Revenue: £{s['mtd_revenue']:.2f}  "
            f"Profit: **{mtd_profit_sign}£{s['mtd_profit']:.2f}**\n"
            f"Cost of sold: £{s['mtd_cost_of_sold']:.2f}  "
            f"Monthly ROI: **{s['mtd_monthly_roi']}%**\n"
            f"Added: {s['mtd_added_count']}"
        ),
        inline=False,
    )
    # Forecast section
    if forecast.get("insufficient_data"):
        sample = forecast.get("sample_size", 0)
        embed.add_field(
            name="📈 30-Day Forecast",
            value=f"Not enough data yet (need 3+ sales with dates, have {sample})",
            inline=False,
        )
    else:
        sign = "+" if forecast["projected_30d_profit"] >= 0 else ""
        cost_in_stock   = s.get("total_cost_in_stock", 0)
        est_roi_on_stock = (forecast["projected_30d_profit"] / cost_in_stock * 100) if cost_in_stock else 0
        embed.add_field(
            name="📈 30-Day Forecast",
            value=(
                f"Avg days to sell:  **{forecast['avg_days_to_sell']}** days\n"
                f"Avg profit/item:   £{forecast['avg_profit_per_item']:.2f}  "
                f"(avg margin {forecast['avg_margin_pct']}%)\n"
                f"Est. items sold:   ~{forecast['projected_30d_items']}\n"
                f"Est. profit:       **{sign}£{forecast['projected_30d_profit']:.2f}**\n"
                f"Est. ROI on stock: ~{est_roi_on_stock:.1f}%  "
                f"(£{forecast['projected_30d_profit']:.2f} profit on £{cost_in_stock:.0f} in stock)\n"
                f"*(based on last {forecast['sample_size']} sold items)*"
            ),
            inline=False,
        )
    await interaction.followup.send(embed=embed)


# ---------------------------------------------------------------------------
# /bestsellers
# ---------------------------------------------------------------------------

@bot.tree.command(name="bestsellers", description="Show your best performing cards by profit, margin, and volume.")
@app_commands.describe(
    metric="Sort by: profit, margin, or volume",
    period="all = all time, month = current month, 30d = last 30 days",
)
@app_commands.choices(metric=[
    app_commands.Choice(name="By profit",  value="profit"),
    app_commands.Choice(name="By margin",  value="margin"),
    app_commands.Choice(name="By volume",  value="volume"),
])
@app_commands.choices(period=[
    app_commands.Choice(name="All time",     value="all"),
    app_commands.Choice(name="This month",   value="month"),
    app_commands.Choice(name="Last 30 days", value="30d"),
])
async def slash_bestsellers(
    interaction: discord.Interaction,
    metric: app_commands.Choice[str],
    period: Optional[app_commands.Choice[str]] = None,
) -> None:
    await interaction.response.defer()

    period_value = period.value if period else "all"
    today        = date.today()

    all_sold = await excel_db.get_stock_async(status_filter="Sold")

    if period_value == "month":
        prefix   = today.strftime("%Y-%m")
        filtered = [i for i in all_sold if str(i.get("date_sold") or "").startswith(prefix)]
        period_label = today.strftime("%B %Y")
    elif period_value == "30d":
        cutoff_dt = today - timedelta(days=30)
        def _parse_sold(i):
            try:
                return datetime.strptime(str(i.get("date_sold", "") or "")[:10], "%Y-%m-%d").date()
            except ValueError:
                return None
        filtered     = [i for i in all_sold if (_d := _parse_sold(i)) is not None and _d >= cutoff_dt]
        period_label = "Last 30 Days"
    else:
        filtered     = all_sold
        period_label = "All Time"

    sold_items = [
        i for i in filtered
        if float(i.get("sell_price") or 0) > 0
        and float(i.get("purchase_price") or 0) > 0
    ]

    if not sold_items:
        await interaction.followup.send(
            f"No sold items found for period: **{period_label}**.", ephemeral=True
        )
        return

    LIMIT = 5

    def _item_margin(i):
        sell = float(i.get("sell_price") or 0)
        buy  = float(i.get("purchase_price") or 0)
        return ((sell - buy) / sell * 100) if sell else 0

    metric_value = metric.value

    if metric_value == "profit":
        top = sorted(
            sold_items,
            key=lambda i: float(i.get("sell_price", 0)) - float(i.get("purchase_price", 0)),
            reverse=True,
        )[:LIMIT]
        lines = []
        for idx, i in enumerate(top, 1):
            profit = float(i["sell_price"]) - float(i["purchase_price"])
            lines.append(
                f"{idx}. **{i['card_name'][:45]}**\n"
                f"   bought £{i['purchase_price']:.2f} → sold £{i['sell_price']:.2f}  "
                f"profit **+£{profit:.2f}**"
            )
        embed = discord.Embed(
            title=f"🏆 Best Sellers — By Profit ({period_label})",
            description="\n".join(lines),
            colour=discord.Colour.gold(),
        )

    elif metric_value == "margin":
        top = sorted(sold_items, key=_item_margin, reverse=True)[:LIMIT]
        lines = []
        for idx, i in enumerate(top, 1):
            m = _item_margin(i)
            lines.append(
                f"{idx}. **{i['card_name'][:45]}**\n"
                f"   bought £{i['purchase_price']:.2f} → sold £{i['sell_price']:.2f}  "
                f"margin **{m:.1f}%**"
            )
        embed = discord.Embed(
            title=f"📈 Best Sellers — By Margin ({period_label})",
            description="\n".join(lines),
            colour=discord.Colour.green(),
        )

    else:  # volume
        from collections import Counter
        name_counts = Counter(i["card_name"] for i in sold_items)
        top         = name_counts.most_common(LIMIT)
        lines = [
            f"{idx}. **{name[:45]}** — {count} sale{'s' if count > 1 else ''}"
            for idx, (name, count) in enumerate(top, 1)
        ]
        embed = discord.Embed(
            title=f"🔄 Most Sold Cards ({period_label})",
            description="\n".join(lines),
            colour=discord.Colour.blurple(),
        )

    embed.set_footer(text=f"{len(sold_items)} qualifying sold item(s) in period")
    await interaction.followup.send(embed=embed)


# ---------------------------------------------------------------------------
# /monthly
# ---------------------------------------------------------------------------

@bot.tree.command(name="monthly", description="Show monthly sales analysis.")
@app_commands.describe(month="Month to analyse in YYYY-MM format (default: current month)")
async def slash_monthly(interaction: discord.Interaction, month: str = "") -> None:
    await interaction.response.defer()

    today = date.today()
    if not month:
        year, mon = today.year, today.month
    else:
        try:
            parts = month.strip().split("-")
            if len(parts) != 2:
                raise ValueError
            year, mon = int(parts[0]), int(parts[1])
            if not (1 <= mon <= 12) or year < 2000:
                raise ValueError
        except ValueError:
            await interaction.followup.send(
                "❌ Invalid month format. Use `YYYY-MM` (e.g. `2026-05`).", ephemeral=True
            )
            return

    try:
        stats = await excel_db.get_monthly_stats_async(year, mon)
    except Exception as exc:
        await interaction.followup.send(f"Error reading stats: {exc}")
        return

    month_label = date(year, mon, 1).strftime("%B %Y")
    profit_sign = "+" if stats["total_profit"] >= 0 else ""
    colour = discord.Colour.green() if stats["total_profit"] >= 0 else discord.Colour.red()

    embed = discord.Embed(title=f"Monthly Analysis — {month_label}", colour=colour)
    embed.add_field(name="Items Sold",    value=str(stats["sold_count"]),                    inline=True)
    embed.add_field(name="Items Added",   value=str(stats["added_count"]),                   inline=True)
    embed.add_field(name="​",             value="​",                                         inline=True)
    embed.add_field(name="Revenue",       value=f"£{stats['total_revenue']:.2f}",            inline=True)
    embed.add_field(name="Cost of Sold",  value=f"£{stats['total_cost']:.2f}",              inline=True)
    embed.add_field(name="Profit",        value=f"**{profit_sign}£{stats['total_profit']:.2f}**", inline=True)

    if stats["sold_count"] > 0:
        embed.add_field(name="Avg Profit/Item", value=f"£{stats['avg_profit']:.2f}", inline=True)

    if stats["best_name"]:
        embed.add_field(
            name="Best Performer",
            value=f"{stats['best_name']} (+£{stats['best_profit']:.2f})",
            inline=False,
        )
    if stats["worst_name"] and stats["worst_name"] != stats["best_name"]:
        worst_sign = "+" if stats["worst_profit"] >= 0 else ""
        embed.add_field(
            name="Worst Performer",
            value=f"{stats['worst_name']} ({worst_sign}£{stats['worst_profit']:.2f})",
            inline=False,
        )

    embed.add_field(
        name="Currently In Stock",
        value=f"{stats['stock_count']} items — cost £{stats['stock_cost']:.2f}",
        inline=False,
    )
    await interaction.followup.send(embed=embed)


# ---------------------------------------------------------------------------
# /velocity
# ---------------------------------------------------------------------------

@bot.tree.command(name="velocity", description="Show how fast different cards sell on average.")
@app_commands.describe(
    group_by="Group results by: set, condition, or price range",
    min_sales="Minimum sales needed to show a group (default 2)",
)
@app_commands.choices(group_by=[
    app_commands.Choice(name="By set",         value="set"),
    app_commands.Choice(name="By condition",   value="condition"),
    app_commands.Choice(name="By price range", value="price"),
])
async def slash_velocity(
    interaction: discord.Interaction,
    group_by: app_commands.Choice[str],
    min_sales: int = 2,
) -> None:
    await interaction.response.defer()

    results = await excel_db.get_velocity_stats_async(group_by.value)
    results = [r for r in results if r["count"] >= min_sales]

    if not results:
        await interaction.followup.send(
            f"Not enough sales data yet — need at least {min_sales} sales per group."
        )
        return

    lines = []
    for r in results[:20]:
        if r["avg_days"] < 7:
            indicator = "🟢"
        elif r["avg_days"] < 21:
            indicator = "🟡"
        else:
            indicator = "🔴"
        lines.append(
            f"{indicator} {r['group'][:22]:<22} {r['avg_days']:>5.1f}d  "
            f"£{r['avg_profit']:>5.2f}  ({r['count']} sales)"
        )

    group_label = {"set": "Set", "condition": "Condition", "price": "Price Range"}.get(group_by.value, group_by.value)
    embed = discord.Embed(
        title=f"⚡ Sales Velocity — By {group_label}",
        colour=discord.Colour.blurple(),
    )
    embed.add_field(
        name=f"{group_label} Performance",
        value="```\n" + "\n".join(lines) + "\n```",
        inline=False,
    )

    fastest   = results[0]
    most_prof = max(results, key=lambda x: x["avg_profit"])
    slowest   = results[-1]
    tips = [f"💡 Fastest: {fastest['group']} ({fastest['avg_days']} days avg)"]
    if most_prof["group"] != fastest["group"]:
        tips.append(f"💡 Most profitable: {most_prof['group']} (£{most_prof['avg_profit']:.2f} avg)")
    if slowest["avg_days"] > 14:
        tips.append(f"💡 Slowest: {slowest['group']} ({slowest['avg_days']} days) — consider pricing lower")
    embed.add_field(name="Insights", value="\n".join(tips), inline=False)

    await interaction.followup.send(embed=embed)


# ---------------------------------------------------------------------------
# /efficiency
# ---------------------------------------------------------------------------

@bot.tree.command(name="efficiency", description="Show profit per hour of effort by card.")
@app_commands.describe(show="all = all items, flagged = below minimum wage, top = top earners")
@app_commands.choices(show=[
    app_commands.Choice(name="All cards",          value="all"),
    app_commands.Choice(name="Below minimum wage", value="flagged"),
    app_commands.Choice(name="Top earners",        value="top"),
])
async def slash_efficiency(
    interaction: discord.Interaction,
    show: app_commands.Choice[str],
) -> None:
    await interaction.response.defer()

    effort_hours = config.EFFORT_MINUTES_PER_CARD / 60
    hourly_rate  = config.HOURLY_RATE_GBP

    sold_items = await excel_db.get_stock_async(status_filter="Sold")

    scored = []
    for item in sold_items:
        sell_price     = float(item.get("sell_price")     or 0)
        purchase_price = float(item.get("purchase_price") or 0)
        if sell_price <= 0:
            continue
        profit    = sell_price - purchase_price
        profit_hr = round(profit / effort_hours, 2) if effort_hours > 0 else 0.0
        scored.append({
            "item_id":   item["item_id"],
            "card_name": item["card_name"],
            "profit":    profit,
            "profit_hr": profit_hr,
            "flagged":   profit_hr < hourly_rate,
        })

    if not scored:
        await interaction.followup.send("No sold items to analyse yet.")
        return

    if show.value == "flagged":
        display = [s for s in scored if s["flagged"]]
        display.sort(key=lambda x: x["profit_hr"])
        section_title = f"Below minimum wage (£{hourly_rate:.0f}/hr)"
    elif show.value == "top":
        display = sorted(scored, key=lambda x: x["profit_hr"], reverse=True)
        section_title = "Top earners"
    else:
        display = sorted(scored, key=lambda x: x["profit_hr"], reverse=True)
        section_title = "All cards"

    MAX_ITEMS   = 30
    CHUNK_SIZE  = 15
    total_count = len(display)
    truncated   = total_count > MAX_ITEMS
    display     = display[:MAX_ITEMS]

    lines = []
    for s in display:
        icon = "🔴" if s["flagged"] else "💚"
        loss = "  LOSS" if s["profit"] < 0 else ""
        lines.append(
            f"{icon} {s['card_name'][:35]:<35}  "
            f"£{s['profit']:>6.2f}  → £{s['profit_hr']:>6.2f}/hr{loss}"
        )

    above = sum(1 for s in scored if not s["flagged"])
    below = len(scored) - above
    avg_hr = round(sum(s["profit_hr"] for s in scored) / len(scored), 2)

    embed = discord.Embed(
        title=f"⚙️ Profit Efficiency ({config.EFFORT_MINUTES_PER_CARD} min/card, £{hourly_rate:.0f}/hr minimum)",
        colour=discord.Colour.orange(),
    )

    if not lines:
        embed.add_field(name=section_title, value="*None*", inline=False)
    else:
        for i in range(0, len(lines), CHUNK_SIZE):
            chunk      = lines[i:i + CHUNK_SIZE]
            field_name = section_title if i == 0 else f"Results {i + 1}–{min(i + CHUNK_SIZE, len(lines))}"
            embed.add_field(
                name=field_name,
                value="```\n" + "\n".join(chunk) + "\n```",
                inline=False,
            )
    if truncated:
        embed.add_field(
            name="Note",
            value=f"Showing {MAX_ITEMS} of {total_count} items.",
            inline=False,
        )
    summary_text = (
        f"Above minimum wage: {above} cards ({round(above / len(scored) * 100)}%)\n"
        f"Below minimum wage: {below} cards ({round(below / len(scored) * 100)}%)\n"
        f"Average profit/hr:  £{avg_hr:.2f}"
    )
    embed.add_field(name="Summary", value=summary_text, inline=False)
    if below > 0:
        embed.add_field(
            name="💡 Tip",
            value="Cards under £3 profit rarely justify the effort — price higher or skip.",
            inline=False,
        )

    await interaction.followup.send(embed=embed)


# ---------------------------------------------------------------------------
# /status
# ---------------------------------------------------------------------------

@bot.tree.command(name="status", description="Show bot health, uptime, and system stats.")
async def slash_status(interaction: discord.Interaction) -> None:
    await interaction.response.defer(ephemeral=True)

    now = datetime.now(timezone.utc)

    uptime_str = _fmt_delta(now - _bot_start_time)

    if _last_price_update:
        last_pu_str = _fmt_delta(now - _last_price_update) + " ago"
        cycle_stats = f"({_price_update_stats['checked']} checked, {_price_update_stats['updated']} updated)"
    else:
        last_pu_str = "never"
        cycle_stats = ""

    next_pu = price_update_loop.next_iteration
    if next_pu:
        delta = next_pu - now
        next_pu_str = ("in " + _fmt_delta(delta)) if delta.total_seconds() > 0 else "imminent"
    else:
        next_pu_str = "not running"

    if _last_ebay_sales_check:
        last_es_str = _fmt_delta(now - _last_ebay_sales_check) + " ago"
    else:
        last_es_str = "never"

    next_es = check_ebay_sales_loop.next_iteration
    if next_es:
        delta_es = next_es - now
        next_es_str = ("in " + _fmt_delta(delta_es)) if delta_es.total_seconds() > 0 else "imminent"
    else:
        next_es_str = "not running"

    ok_count  = _price_update_stats["pc_success"]
    err_count = _price_update_stats["pc_429s"]
    if ok_count + err_count > 0:
        rate = ok_count / (ok_count + err_count) * 100
        pc_health_str = f"{rate:.1f}%  ({ok_count} ok / {err_count} errors)"
    else:
        pc_health_str = "No data yet (first cycle pending)"

    try:
        all_items = await excel_db.get_stock_async(status_filter=None)
        inv_items     = [i for i in all_items if i.get("status") == "Inventory"]
        ebay_count    = sum(1 for i in inv_items if i.get("ebay_listed") == "Yes")
        unverified    = sum(1 for i in inv_items if not i.get("price_verified", True))
        inv_count     = len(inv_items)
    except Exception:
        inv_count = ebay_count = unverified = 0

    embed = discord.Embed(title="🤖 Pokemaz Bot Status", colour=discord.Colour.blurple())

    timing_lines = (
        f"Uptime:                {uptime_str}\n"
        f"Last price update:     {last_pu_str}  {cycle_stats}\n"
        f"Next price update:     {next_pu_str}\n"
        f"Last eBay sales check: {last_es_str}\n"
        f"Next sales check:      {next_es_str}"
    )
    embed.add_field(name="⏱️ Timing", value=f"```{timing_lines}```", inline=False)

    pc_lines = (
        f"Success rate:    {pc_health_str}\n"
        f"429 rate-limits: {err_count} (last cycle)"
    )
    embed.add_field(name="📊 PriceCharting Health", value=f"```{pc_lines}```", inline=False)

    inv_lines = (
        f"Items in inventory:    {inv_count}\n"
        f"Active eBay listings:  {ebay_count}  (tracked in bot)\n"
        f"Unverified prices:     {unverified}"
    )
    embed.add_field(name="📦 Inventory vs eBay", value=f"```{inv_lines}```", inline=False)

    sys_lines = (
        f"Price update interval: every {config.UPDATE_INTERVAL_HOURS:.0f}h\n"
        f"Sales check interval:  every 30min"
    )
    embed.add_field(name="⚙️ System", value=f"```{sys_lines}```", inline=False)

    await interaction.followup.send(embed=embed)


# ---------------------------------------------------------------------------
# /listvinted
# ---------------------------------------------------------------------------

@bot.tree.command(name="listvinted", description="List a card on Vinted UK.")
@app_commands.describe(
    item_id="Item ID to list",
    image1="Primary photo (required)",
    image2="Second photo (optional)",
    price_override="Override the suggested price in GBP (optional)",
)
async def slash_listvinted(
    interaction: discord.Interaction,
    item_id: int,
    image1: discord.Attachment,
    image2: Optional[discord.Attachment] = None,
    price_override: float = 0.0,
) -> None:
    await interaction.response.defer()

    try:
        item = await excel_db.get_item_async(item_id)
    except ValueError as exc:
        await interaction.followup.send(f"❌ Item not found: {exc}")
        return

    if item.get("status") != STATUS_INVENTORY:
        await interaction.followup.send(f"❌ Item #{item_id} is not in Inventory.")
        return

    if not Path(config.VINTED_STATE_PATH).exists():
        await interaction.followup.send(
            "❌ No Vinted session found. Run `python generate_cookies.py vinted` to log in."
        )
        return

    attachments = [a for a in [image1, image2] if a is not None]

    if price_override > 0:
        listing_price = price_override
    elif item.get("live_price") is not None:
        listing_price = max(round(float(item["live_price"]) * 0.95, 2), config.EBAY_MIN_PRICE_GBP)
    elif item.get("purchase_price") is not None:
        listing_price = float(item["purchase_price"])
    else:
        await interaction.followup.send(
            f"❌ Item #{item_id} has no price recorded. Use `price_override`."
        )
        return

    status_msg = await interaction.followup.send(
        f"⏳ Generating listing for **{item['card_name']}** at £{listing_price:.2f}…"
    )

    try:
        ai_content = await ai_helper.generate_listing_content(
            item_name=item["card_name"],
            condition=item.get("condition") or "ungraded",
            uk_avg_price_gbp=item.get("live_price"),
        )
        ai_title       = ai_content["title"]
        ai_description = ai_content["description"]
    except Exception as exc:
        print(f"[listvinted] AI helper failed: {exc}")
        ai_title       = item["card_name"]
        ai_description = ""

    temp_dir = Path(config.TEMP_IMAGES_DIR) / f"vinted_{item_id}"
    try:
        await status_msg.edit(content=f"⏳ Uploading **{item['card_name']}** to Vinted…")
        image_paths = await _download_attachments(attachments, temp_dir)

        vinted_cond = lister_vinted.INVENTORY_CONDITION_MAP.get(item.get("condition") or "", "good")

        result = await lister_vinted.list_item_on_vinted(
            item_name=ai_title,
            price_gbp=listing_price,
            image_paths=image_paths,
            condition=vinted_cond,
            description=ai_description[:2000],
        )
    finally:
        if temp_dir.exists():
            await asyncio.to_thread(shutil.rmtree, temp_dir, ignore_errors=True)

    if result.success:
        try:
            await excel_db.mark_vinted_listed_async(item_id, result.listing_url or "")
            audit.log_mutation("listvinted", item_id, "listed on Vinted", {
                "listing_url": result.listing_url,
                "price":       listing_price,
            })
        except Exception as exc:
            print(f"[listvinted] Failed to record Vinted listing for item {item_id}: {exc}")

        if _SYNC_ENABLED:
            try:
                bot_sync.sync_vinted_listed(item_id, result.listing_url or "")
            except Exception as exc:
                print(f"[bot_sync] vinted listed sync failed for item {item_id}: {exc}")

        embed_desc = ai_description[:500] + "…" if len(ai_description) > 500 else ai_description
        embed = discord.Embed(title="✅ Listed on Vinted", colour=discord.Colour.green())
        embed.add_field(name="Card",  value=item["card_name"],    inline=False)
        embed.add_field(name="Title", value=f"`{ai_title}`",      inline=False)
        if embed_desc:
            embed.add_field(name="Description", value=embed_desc, inline=False)
        embed.add_field(name="Price", value=f"£{listing_price:.2f}", inline=True)
        embed.add_field(name="Images", value=str(len(attachments)), inline=True)
        if result.listing_url:
            embed.add_field(name="Listing URL", value=result.listing_url, inline=False)
        await status_msg.edit(content=None, embed=embed)
    else:
        error_msg = result.error or "Unknown error"
        if "session expired" in error_msg.lower() or "generate_cookies" in error_msg.lower() or "import_cookies" in error_msg.lower():
            reply = "❌ Vinted session expired. Run `python generate_cookies.py vinted` to refresh."
        else:
            reply = f"❌ Vinted listing failed: {error_msg}"
        await status_msg.edit(content=reply)


# ---------------------------------------------------------------------------
# /watch  /watchlist  /unwatch  (buy watchlist)
# ---------------------------------------------------------------------------

async def _check_watchlist() -> None:
    """Check watchlist prices and fire alerts when targets are hit."""
    watchlist = await excel_db.get_watchlist_async()
    if not watchlist:
        return

    channel = (
        bot.get_channel(config.PRICE_UPDATE_CHANNEL_ID)
        if config.PRICE_UPDATE_CHANNEL_ID
        else None
    )

    for entry in watchlist:
        try:
            _, current_price = await scraper.scrape_card(entry["pc_url"], "Near mint or better")
            if current_price is None:
                continue

            await excel_db.update_watchlist_price_async(entry["watch_id"], current_price)

            # Reset alert when price recovers above target + 5% buffer
            if entry["alert_sent"] and current_price > entry["target_price_gbp"] * 1.05:
                await excel_db.reset_watchlist_alert_async(entry["watch_id"])
                continue

            # Fire alert if not yet alerted and price at or below target
            if not entry["alert_sent"] and current_price <= entry["target_price_gbp"]:
                embed = discord.Embed(title="🔔 Buy Alert Triggered!", colour=discord.Colour.green())
                embed.add_field(name="Card",          value=entry["card_name"],                            inline=False)
                embed.add_field(name="Current price", value=f"£{current_price:.2f}",                      inline=True)
                embed.add_field(name="Your target",   value=f"£{entry['target_price_gbp']:.2f}",          inline=True)
                embed.add_field(name="Below target by", value=f"£{entry['target_price_gbp'] - current_price:.2f}", inline=True)
                embed.add_field(name="PriceCharting", value=f"[View]({entry['pc_url']})",                 inline=False)
                embed.set_footer(text=f"Watch ID: {entry['watch_id']} — use /unwatch to remove")
                if channel:
                    await channel.send(embed=embed)
                await excel_db.mark_watchlist_alerted_async(entry["watch_id"])

        except Exception as exc:
            print(f"[watchlist] Error checking {entry['card_name']}: {exc}")

        await asyncio.sleep(2.0)


@bot.tree.command(name="watch", description="Add a card to your buy watchlist.")
@app_commands.describe(
    pc_url="PriceCharting URL for the card",
    target_price="Alert when price drops to or below this (£)",
)
async def slash_watch(interaction: discord.Interaction, pc_url: str, target_price: float) -> None:
    _PC_PREFIX = "https://www.pricecharting.com/game/"
    if not pc_url.lower().startswith(_PC_PREFIX):
        await interaction.response.send_message(
            "❌ That doesn't look like a PriceCharting product URL.", ephemeral=True
        )
        return

    if target_price <= 0:
        await interaction.response.send_message("❌ `target_price` must be greater than 0.", ephemeral=True)
        return

    await interaction.response.defer()

    card_name, current_price = await scraper.scrape_card(pc_url, "Near mint or better")

    watch_id = await excel_db.add_watchlist_entry_async(card_name, pc_url, target_price, current_price)

    current_str = f"£{current_price:.2f}" if current_price is not None else "unknown"
    await interaction.followup.send(
        f"👁️ Watching **{card_name}** — alert when ≤ £{target_price:.2f} "
        f"(currently {current_str})  `Watch ID: {watch_id}`"
    )


@bot.tree.command(name="watchlist", description="Show your buy watchlist.")
async def slash_watchlist(interaction: discord.Interaction) -> None:
    await interaction.response.defer()

    watchlist = await excel_db.get_watchlist_async()
    if not watchlist:
        await interaction.followup.send("👁️ Your watchlist is empty. Use `/watch` to add a card.")
        return

    lines = [f"{'ID':<4} {'Card':<32} {'Target':>8}  {'Current':>8}  {'Status'}"]
    lines.append("─" * 70)
    for entry in watchlist:
        cur = f"£{entry['current_price_gbp']:.2f}" if entry["current_price_gbp"] is not None else "—"
        status = "🔔 TRIGGERED" if entry["alert_sent"] and (
            entry["current_price_gbp"] is not None
            and entry["current_price_gbp"] <= entry["target_price_gbp"]
        ) else "watching"
        lines.append(
            f"{entry['watch_id']:<4} {entry['card_name'][:32]:<32} "
            f"£{entry['target_price_gbp']:>6.2f}  {cur:>8}  {status}"
        )

    embed = discord.Embed(
        title=f"👁️ Buy Watchlist ({len(watchlist)} item{'s' if len(watchlist) != 1 else ''})",
        description=f"```{chr(10).join(lines)}```",
        colour=discord.Colour.blurple(),
    )
    embed.set_footer(text="Triggered entries re-alert if price rises above target then drops again.")
    await interaction.followup.send(embed=embed)


@bot.tree.command(name="unwatch", description="Remove a card from your watchlist.")
@app_commands.describe(watch_id="Watch ID to remove (from /watchlist)")
async def slash_unwatch(interaction: discord.Interaction, watch_id: int) -> None:
    await interaction.response.defer()
    try:
        await excel_db.remove_watchlist_entry_async(watch_id)
        await interaction.followup.send(f"✅ Removed watch ID {watch_id} from your watchlist.")
    except ValueError as exc:
        await interaction.followup.send(f"❌ {exc}")


# ---------------------------------------------------------------------------
# /verifylisting — check which eBay listings are still active
# ---------------------------------------------------------------------------

@bot.tree.command(name="verifylisting", description="Check which eBay listings are still active and clean up ended ones.")
@app_commands.describe(dry_run="Preview what would be cleared without making changes (default True)")
async def slash_verifylisting(
    interaction: discord.Interaction,
    dry_run: bool = True,
) -> None:
    await interaction.response.defer()

    channel = bot.get_channel(config.PRICE_UPDATE_CHANNEL_ID)

    try:
        await interaction.followup.send(
            f"⏳ Verifying eBay listings… "
            f"{'(dry run — no changes)' if dry_run else '(live — will clear ended listings)'}\n"
            f"Results will post to <#{config.PRICE_UPDATE_CHANNEL_ID}> when done."
        )
    except Exception:
        pass

    all_items = await excel_db.get_stock_async(status_filter=None)
    listed_items = [
        item for item in all_items
        if str(item.get("ebay_listed", "") or "") == "Yes"
        and str(item.get("ebay_listing_id", "") or "").strip().isdigit()
        and str(item.get("status", "")) == "Inventory"
    ]

    print(f"[ebay_verify] Checking {len(listed_items)} listed items…")

    still_active: list[dict] = []
    ended:        list[dict] = []
    clear_errors: list[dict] = []

    for item in listed_items:
        listing_id = str(item["ebay_listing_id"]).strip()
        is_active  = await lister_ebay.check_listing_active(listing_id)
        if is_active:
            still_active.append(item)
        else:
            ended.append(item)
            print(f"[ebay_verify] ENDED: #{item['item_id']} {item['card_name']} (listing {listing_id})")
        await asyncio.sleep(0.3)

    cleared = 0
    if not dry_run:
        for item in ended:
            try:
                await excel_db.edit_item_async(item["item_id"], "ebay_listed",    "No")
                await excel_db.edit_item_async(item["item_id"], "ebay_listing_id", "")
                cleared += 1
            except Exception as exc:
                print(f"[ebay_verify] Error clearing item {item['item_id']}: {exc}")
                clear_errors.append(item)

    embed = discord.Embed(
        title=f"{'🔍 DRY RUN — ' if dry_run else ''}eBay Listing Verification",
        colour=discord.Colour.green() if not ended else discord.Colour.orange(),
    )
    embed.add_field(name="Checked",       value=str(len(listed_items)), inline=True)
    embed.add_field(name="Still active",  value=str(len(still_active)), inline=True)
    embed.add_field(name="Ended/missing", value=str(len(ended)),        inline=True)

    if ended:
        lines = []
        for item in ended[:15]:
            lines.append(f"• #{item['item_id']} {item['card_name'][:35]}\n  Listing: {item['ebay_listing_id']}")
        ended_text = "\n".join(lines[:10])
        if len(ended) > 10:
            ended_text += f"\n…and {len(ended) - 10} more"
        embed.add_field(
            name=f"{'Would clear' if dry_run else 'Cleared'} ({len(ended)})",
            value=ended_text,
            inline=False,
        )

    if dry_run and ended:
        embed.set_footer(text="Run /verifylisting dry_run:False to apply changes")
    elif not dry_run:
        note = f"Cleared {cleared} ended listings from inventory"
        if clear_errors:
            note += f" ({len(clear_errors)} failed — check logs)"
        embed.set_footer(text=note)
    else:
        embed.set_footer(text="All listings verified active ✅")

    if channel:
        await channel.send(embed=embed)

    audit.log_mutation("verifylisting", None, "listing_verification", {
        "checked": len(listed_items),
        "active":  len(still_active),
        "ended":   len(ended),
        "cleared": cleared,
        "dry_run": dry_run,
    })


# ---------------------------------------------------------------------------
# /datahealth — price data quality report
# ---------------------------------------------------------------------------

@bot.tree.command(name="datahealth", description="Show inventory items with missing or stale price data.")
async def slash_datahealth(interaction: discord.Interaction) -> None:
    await interaction.response.defer()
    try:
        health = await _check_price_health()
    except Exception as exc:
        await interaction.followup.send(f"❌ Health check failed: {exc}")
        return

    total = len(health["zero_price"]) + len(health["no_pc_url"]) + len(health["stale_price"])

    if total == 0:
        await interaction.followup.send("✅ All inventory items have healthy price data.")
        return

    embed = discord.Embed(
        title=f"🔍 Data Health Check ({total} issues)",
        colour=discord.Colour.orange(),
    )

    if health["zero_price"]:
        lines = [f"#{i['item_id']} {i['card_name'][:35]}" for i in health["zero_price"][:10]]
        embed.add_field(
            name=f"💰 Zero/missing price ({len(health['zero_price'])})",
            value="\n".join(lines),
            inline=False,
        )

    if health["no_pc_url"]:
        lines = [f"#{i['item_id']} {i['card_name'][:35]}" for i in health["no_pc_url"][:10]]
        embed.add_field(
            name=f"🔗 No PriceCharting URL ({len(health['no_pc_url'])})",
            value="\n".join(lines),
            inline=False,
        )

    if health["stale_price"]:
        lines = [f"#{i['item_id']} {i['card_name'][:35]}" for i in health["stale_price"][:10]]
        embed.add_field(
            name=f"⏰ Stale price 14+ days ({len(health['stale_price'])})",
            value="\n".join(lines),
            inline=False,
        )

    embed.set_footer(text="Use /edit item_id:X field:pc_url value:<url> to fix missing URLs")
    await interaction.followup.send(embed=embed)


# ---------------------------------------------------------------------------
# /restock — buying suggestions based on sales velocity and stock levels
# ---------------------------------------------------------------------------

@bot.tree.command(name="restock", description="Get buying suggestions based on sales velocity and current stock.")
async def slash_restock(interaction: discord.Interaction) -> None:
    await interaction.response.defer()

    try:
        suggestions = await excel_db.get_restock_suggestions()
    except Exception as exc:
        await interaction.followup.send(f"❌ Restock analysis failed: {exc}")
        return

    if not suggestions:
        await interaction.followup.send(
            "No strong restock signals right now — not enough sales data or stock is well-balanced."
        )
        return

    embed = discord.Embed(
        title="🔄 Restock Suggestions",
        description="Sets that sell fast but you're low on stock",
        colour=discord.Colour.blue(),
    )

    for s in suggestions[:8]:
        embed.add_field(
            name=s["set"],
            value=(
                f"Avg sell time: **{s['avg_days']}d**\n"
                f"Avg profit: **£{s['avg_profit']:.2f}**\n"
                f"Current stock: **{s['current_stock']}**\n"
                f"({s['total_sold']} historical sales)"
            ),
            inline=True,
        )

    try:
        watchlist = await excel_db.get_watchlist_async()
        matched_watches = []
        for s in suggestions[:5]:
            for w in watchlist:
                if s["set"].lower() in (w.get("card_name") or "").lower():
                    matched_watches.append((s["set"], w))

        if matched_watches:
            lines = [
                f"👁️ **{set_name}**: {w['card_name']} — "
                f"target £{float(w.get('target_price_gbp', 0)):.2f}, "
                f"current £{float(w.get('current_price_gbp') or 0):.2f}"
                for set_name, w in matched_watches
            ]
            embed.add_field(
                name="🎯 Matching Watchlist Entries",
                value="\n".join(lines),
                inline=False,
            )
    except Exception:
        pass

    embed.set_footer(text="Combine with /watch to track buy prices for specific cards in these sets")
    await interaction.followup.send(embed=embed)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    bot.run(config.BOT_TOKEN)