from fastapi import APIRouter, Depends, HTTPException
from web.auth import get_current_user
from web.database import get_db
from web.notifications import send_discord_notification
import config
import requests
import logging

logger = logging.getLogger(__name__)
router = APIRouter()


@router.get("")
async def get_settings(user: dict = Depends(get_current_user)):
    """Return current settings — mask sensitive values."""
    return {
        "display_name":      user.get("display_name", ""),
        "email":             user.get("email", ""),
        "plan":              user.get("plan", "free"),
        "has_ebay":          bool(user.get("ebay_app_id")),
        "has_gemini":        bool(user.get("gemini_api_key")),
        "has_discord":       bool(user.get("discord_webhook_url")),
        "ebay_fee_rate":     user.get("ebay_fee_rate", 0.1235),
        "postage_cost":      user.get("postage_cost", 1.50),
        "promoted_listing_pct": float(user.get("promoted_listing_pct") or 0),
        "auto_sync_ebay":    user.get("auto_sync_ebay_prices", True),
        "korean_multiplier": user.get("korean_price_multiplier", 0.7),
        "effort_minutes":    user.get("effort_minutes_per_card", 15),
        "hourly_rate":       user.get("hourly_rate_gbp", 12.0),
        "ebay_fulfillment_policy_id": user.get("ebay_fulfillment_policy_id", ""),
        "ebay_payment_policy_id":     user.get("ebay_payment_policy_id", ""),
        "ebay_return_policy_id":      user.get("ebay_return_policy_id", ""),
        "onboarding_dismissed": user.get("onboarding_dismissed", False),
    }


@router.patch("")
async def update_settings(body: dict, user: dict = Depends(get_current_user)):
    """Update user settings and API keys."""
    db = get_db()
    allowed = {
        "display_name", "ebay_fee_rate", "postage_cost", "promoted_listing_pct",
        "auto_sync_ebay_prices", "korean_price_multiplier",
        "effort_minutes_per_card", "hourly_rate_gbp",
        # API keys — only update if non-empty string provided
        "ebay_app_id", "ebay_dev_id", "ebay_cert_id", "ebay_refresh_token",
        "ebay_fulfillment_policy_id", "ebay_payment_policy_id", "ebay_return_policy_id",
        "gemini_api_key", "discord_webhook_url",
        "onboarding_dismissed",
    }
    updates = {k: v for k, v in body.items() if k in allowed and v is not None}
    if not updates:
        return {"success": False, "error": "Nothing to update"}

    db.table("user_profiles").update(updates).eq("id", user["id"]).execute()
    return {"success": True, "updated": list(updates.keys())}


@router.delete("/api-key/{key_name}")
async def clear_api_key(key_name: str, user: dict = Depends(get_current_user)):
    """Clear a specific API key."""
    clearable = {
        "ebay_app_id", "ebay_dev_id", "ebay_cert_id", "ebay_refresh_token",
        "gemini_api_key", "discord_webhook_url",
    }
    if key_name not in clearable:
        return {"success": False, "error": "Unknown key"}
    db = get_db()
    db.table("user_profiles").update({key_name: None}).eq("id", user["id"]).execute()
    return {"success": True}


@router.post("/migrate-excel")
async def migrate_from_excel(user: dict = Depends(get_current_user)):
    """
    Import inventory.xlsx into the user's Supabase account.
    Only works if inventory.xlsx exists on the server.
    Idempotent — skips items that already exist.
    """
    from pathlib import Path
    import openpyxl
    import config

    xlsx_path = Path(config.EXCEL_FILE)
    if not xlsx_path.exists():
        return {"success": False, "error": "inventory.xlsx not found on server"}

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
        ws = wb.active
        headers = [c.value for c in ws[1]]

        def col(name):
            return headers.index(name) if name in headers else -1

        db = get_db()
        migrated = skipped = errors = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            item_id = row[col("Item_ID")] if col("Item_ID") >= 0 else None
            if not item_id:
                continue

            # Check if already exists
            existing = db.table("inventory_items").select("id") \
                .eq("user_id", user["id"]).eq("item_id", int(item_id)).execute()
            if existing.data:
                skipped += 1
                continue

            def v(name):
                idx = col(name)
                return row[idx] if idx >= 0 else None

            try:
                db.table("inventory_items").insert({
                    "user_id":        user["id"],
                    "item_id":        int(item_id),
                    "card_name":      str(v("Card_Name") or ""),
                    "pc_url":         str(v("PC_URL") or ""),
                    "condition":      str(v("Condition") or "Near mint or better"),
                    "region":         str(v("Region") or ""),
                    "purchase_price": float(v("Purchase_Price") or 0),
                    "live_price":     float(v("Live_Price") or 0) or None,
                    "quick_price":    float(v("Quick_Price") or 0) or None,
                    "potential_profit": float(v("Potential_Profit") or 0) or None,
                    "sell_price":     float(v("Sell_Price") or 0) or None,
                    "profit":         float(v("Profit") or 0) or None,
                    "status":         str(v("Status") or "Inventory"),
                    "ebay_listing_id": str(v("eBay_Listing_ID") or ""),
                    "ebay_listed":    str(v("eBay_Listed") or "No"),
                    "date_added":     str(v("Date_Added") or "")[:10] or None,
                    "date_sold":      str(v("Date_Sold") or "")[:10] or None,
                    "price_verified": bool(v("Price_Verified")),
                    "source":         str(v("Source") or ""),
                    "image_urls":     str(v("Image_URLs") or ""),
                }).execute()
                migrated += 1
            except Exception as e:
                logger.info(f"[migrate] Error on item {item_id}: {e}")
                errors += 1

        return {
            "success": True,
            "migrated": migrated,
            "skipped":  skipped,
            "errors":   errors,
            "total":    migrated + skipped + errors,
        }

    except Exception as e:
        return {"success": False, "error": str(e)}


@router.post("/test-discord")
async def test_discord_webhook(user: dict = Depends(get_current_user)):
    """Send a test Discord notification to verify webhook is working."""
    await send_discord_notification(
        user["id"],
        "✅ PokeManager Connected",
        "Discord notifications are working! You'll receive alerts for sales, price changes, and account updates.",
        5763719,
    )
    return {"success": True, "message": "Test notification sent"}


# ── eBay OAuth Token Management ───────────────────────────────────────────

@router.get("/ebay/auth-url")
async def get_ebay_auth_url(user: dict = Depends(get_current_user)):
    """Generate and return the eBay OAuth authorization URL."""
    from urllib.parse import urlencode

    app_id = user.get("ebay_app_id", "").strip()
    if not app_id:
        raise HTTPException(
            status_code=400,
            detail="eBay App ID not configured in Settings. Please add your eBay App ID first."
        )

    scopes = " ".join([
        "https://api.ebay.com/oauth/api_scope/sell.inventory",
        "https://api.ebay.com/oauth/api_scope/sell.account",
        "https://api.ebay.com/oauth/api_scope/sell.fulfillment.readonly",
    ])

    auth_url = (
        "https://auth.ebay.com/oauth2/authorize?"
        + urlencode({
            "client_id": app_id,
            "redirect_uri": config.EBAY_REDIRECT_URI,
            "response_type": "code",
            "scope": scopes,
        })
    )

    return {
        "auth_url": auth_url,
        "instructions": "Log in to eBay, grant permission, then copy the full redirect URL from your browser and paste it in the Settings page."
    }


@router.post("/ebay/callback")
async def exchange_ebay_token(body: dict, user: dict = Depends(get_current_user)):
    """Exchange eBay auth code for refresh token and save to user profile."""
    import base64
    from urllib.parse import parse_qs, urlparse

    redirect_url = body.get("redirect_url", "").strip()
    if not redirect_url:
        raise HTTPException(status_code=400, detail="redirect_url is required")

    # Extract auth code from redirect URL
    try:
        parsed = urlparse(redirect_url)
        qs = parse_qs(parsed.query)
        code = qs.get("code", [None])[0]
        if not code:
            raise ValueError("No code in URL")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid redirect URL: {e}")

    app_id = user.get("ebay_app_id", "").strip()
    cert_id = user.get("ebay_cert_id", "").strip()

    if not app_id:
        raise HTTPException(
            status_code=400,
            detail="eBay App ID not configured in Settings. Please add your eBay App ID first."
        )
    if not cert_id:
        raise HTTPException(
            status_code=400,
            detail="eBay Cert ID not configured in Settings. Please add your eBay Cert ID first."
        )

    # Exchange code for tokens
    try:
        credentials = base64.b64encode(f"{app_id}:{cert_id}".encode()).decode()
        resp = requests.post(
            "https://api.ebay.com/identity/v1/oauth2/token",
            headers={
                "Authorization": f"Basic {credentials}",
                "Content-Type": "application/x-www-form-urlencoded",
            },
            data={
                "grant_type": "authorization_code",
                "code": code,
                "redirect_uri": config.EBAY_REDIRECT_URI,
            },
            timeout=30,
        )

        if resp.status_code != 200:
            raise ValueError(f"eBay token exchange failed: {resp.text}")

        tokens = resp.json()
        refresh_token = tokens.get("refresh_token")
        if not refresh_token:
            raise ValueError("No refresh token in eBay response")

        # Save refresh token to user profile
        db = get_db()
        db.table("user_profiles").update({
            "ebay_refresh_token": refresh_token
        }).eq("id", user["id"]).execute()

        logger.info(f"[ebay-oauth] User {user['id']} connected eBay account")

        return {
            "success": True,
            "message": "eBay account connected successfully"
        }

    except Exception as e:
        logger.error(f"[ebay-oauth] Token exchange failed for user {user['id']}: {e}")
        raise HTTPException(status_code=400, detail=f"Token exchange failed: {e}")
