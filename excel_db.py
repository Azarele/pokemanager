"""
Excel inventory database (openpyxl).

Schema
------
Item_ID        : int   — auto-incrementing unique key
Card_Name      : str   — scraped from PriceCharting
PC_URL         : str   — PriceCharting product URL
Condition      : str   — e.g. ungraded, graded, near_mint
Purchase_Price : float — what you paid (GBP)
Live_Price     : float | None — latest PriceCharting estimate in GBP
Sell_Price     : float | None
Profit         : float | None  (Sell_Price - Purchase_Price)
Status         : "Inventory" | "Sold"
Quantity       : int   — copies at this entry (always 1; multiple copies get separate rows)
Date_Added     : str   — ISO date when /add was called (YYYY-MM-DD)
Date_Sold      : str   — ISO date when /sell was called (YYYY-MM-DD)

Concurrency contract
--------------------
All public entry points are async (the *_async variants at the bottom).
A single asyncio.Lock (_lock) serialises every load→mutate→save cycle so
that two concurrent operations can never interleave their disk reads and writes.
The underlying sync functions are only called via asyncio.to_thread inside
the async wrappers; never call them directly from async code.
"""

import asyncio
import time
import urllib.parse
import zipfile
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Optional
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import audit
import backups
import config

# ---------------------------------------------------------------------------
# Concurrency lock — one lock, shared by all async wrappers
# ---------------------------------------------------------------------------

_lock = asyncio.Lock()

# ---------------------------------------------------------------------------
# Column layout
# ---------------------------------------------------------------------------

COLUMNS = [
    "Item_ID",
    "Card_Name",
    "PC_URL",
    "Condition",
    "Region",
    "Purchase_Price",
    "Live_Price",
    "Quick_Price",
    "Potential_Profit",
    "Sell_Price",
    "Profit",
    "Status",
    "Quantity",
    "eBay_Listing_ID",
    "eBay_Listed",
    "Vinted_Listed",
    "Vinted_Listing_URL",
    "Image_URLs",
    "Date_Added",
    "Date_Sold",
    "Price_Verified",
    "Source",
]

_C = {name: idx for idx, name in enumerate(COLUMNS)}

_HEADER_BG  = "2F75B6"
_HEADER_FG  = "FFFFFF"
_COL_WIDTHS = [8, 36, 55, 14, 8, 15, 12, 12, 15, 12, 10, 10, 8, 18, 10, 10, 30, 60, 12, 12, 14, 14]

# Price_History sheet layout
_PH_SHEET   = "Price_History"
_PH_COLUMNS = ["Timestamp", "Item_ID", "Card_Name", "Region", "Live_Price_GBP", "Source"]
_PH_WIDTHS  = [22, 8, 36, 8, 15, 16]

# Watchlist sheet layout
_WL_SHEET   = "Watchlist"
_WL_COLUMNS = ["Watch_ID", "Card_Name", "PC_URL", "Target_Price_GBP", "Current_Price_GBP", "Added_Date", "Alert_Sent"]
_WL_WIDTHS  = [10, 36, 55, 16, 16, 12, 10]
_WL_C       = {name: idx for idx, name in enumerate(_WL_COLUMNS)}


# ---------------------------------------------------------------------------
# Init
# ---------------------------------------------------------------------------

def init_excel() -> None:
    """
    Ensure inventory.xlsx exists with the current schema.

    Migration path (no data loss): if the file exists with the old 9-column
    schema (missing Potential_Profit), the column is inserted in-place and
    values are computed for all Inventory rows.  A full rebuild only happens
    when the file is absent or the first header cell is wrong.
    """
    path = Path(config.EXCEL_FILE)

    if path.exists():
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            if ws.max_row >= 1 and ws.cell(1, 1).value == COLUMNS[0]:
                existing = {ws.cell(1, col).value for col in range(1, ws.max_column + 1)}
                migrated = False
                # Apply migrations in column-order so each insert lands at the right index.
                if "Region" not in existing:
                    _migrate_add_region(wb)
                    migrated = True
                    print(f"[excel_db] Migrated {path}: added Region column")
                if "Potential_Profit" not in existing:
                    _migrate_add_potential_profit(wb)
                    migrated = True
                    print(f"[excel_db] Migrated {path}: added Potential_Profit column")
                if "Quick_Price" not in existing:
                    _migrate_add_quick_price(wb)
                    migrated = True
                    print(f"[excel_db] Migrated {path}: added Quick_Price column")
                if "eBay_Listing_ID" not in existing or "eBay_Listed" not in existing:
                    _migrate_add_ebay_columns(wb)
                    migrated = True
                    print(f"[excel_db] Migrated {path}: added eBay_Listing_ID / eBay_Listed columns")
                if "Quantity" not in existing:
                    _migrate_add_quantity(wb)
                    migrated = True
                    print(f"[excel_db] Migrated {path}: added Quantity column")
                if "Date_Added" not in existing or "Date_Sold" not in existing:
                    _migrate_add_date_columns(wb)
                    migrated = True
                    print(f"[excel_db] Migrated {path}: added Date_Added / Date_Sold columns")
                if "Price_Verified" not in existing:
                    _migrate_add_price_verified(wb)
                    migrated = True
                    print(f"[excel_db] Migrated {path}: added Price_Verified column")
                if "Vinted_Listed" not in existing or "Vinted_Listing_URL" not in existing:
                    _migrate_add_vinted_columns(wb)
                    migrated = True
                    print(f"[excel_db] Migrated {path}: added Vinted_Listed / Vinted_Listing_URL columns")
                if _ensure_price_history_sheet(wb):
                    migrated = True
                    print(f"[excel_db] Migrated {path}: added Price_History sheet")
                if _ensure_watchlist_sheet(wb):
                    migrated = True
                    print(f"[excel_db] Migrated {path}: added Watchlist sheet")
                if "Source" not in existing:
                    _migrate_add_source(wb)
                    migrated = True
                    print(f"[excel_db] Migrated {path}: added Source column")
                if "Image_URLs" not in existing:
                    _migrate_add_image_urls(wb)
                    migrated = True
                    print(f"[excel_db] Migrated {path}: added Image_URLs column")
                ebay_fixed       = _migrate_sync_ebay_listed(wb.active)
                dates_backfilled = _backfill_missing_dates(wb)
                backfill_missing_sold_dates()
                if migrated:
                    _save_with_totals(wb)
                elif ebay_fixed or dates_backfilled:
                    wb.save(path)
                fix_url_encoded_names()
                return  # schema is correct (or just fixed)
        except Exception:
            pass
        print(f"[excel_db] Schema mismatch or corrupt file — recreating {path}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    for col_num, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font      = Font(bold=True, color=_HEADER_FG)
        cell.fill      = PatternFill(
            start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")

    for col_num, width in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    ws.freeze_panes = "A2"
    _ensure_price_history_sheet(wb)
    _ensure_watchlist_sheet(wb)
    wb.save(path)
    print(f"[excel_db] Created {path}")


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _load() -> openpyxl.Workbook:
    last_exc: Exception | None = None
    for attempt in range(5):
        try:
            return openpyxl.load_workbook(config.EXCEL_FILE)
        except (zipfile.BadZipFile, EOFError, Exception) as e:
            last_exc = e
            if attempt < 4:
                time.sleep(0.2 * (attempt + 1))
    raise last_exc  # type: ignore[misc]


def _ensure_price_history_sheet(wb: openpyxl.Workbook) -> bool:
    """
    Create the Price_History sheet if it doesn't already exist.
    Returns True if the sheet was newly created, False if it already existed.
    """
    if _PH_SHEET in wb.sheetnames:
        return False
    ws = wb.create_sheet(_PH_SHEET)
    for col_num, (header, width) in enumerate(zip(_PH_COLUMNS, _PH_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font      = Font(bold=True, color=_HEADER_FG)
        cell.fill      = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_num)].width = width
    return True


def _ensure_watchlist_sheet(wb: openpyxl.Workbook) -> bool:
    """
    Create the Watchlist sheet if it doesn't already exist.
    Returns True if the sheet was newly created, False if it already existed.
    """
    if _WL_SHEET in wb.sheetnames:
        return False
    ws = wb.create_sheet(_WL_SHEET)
    for col_num, (header, width) in enumerate(zip(_WL_COLUMNS, _WL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font      = Font(bold=True, color=_HEADER_FG)
        cell.fill      = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_num)].width = width
    return True


def _get_next_id(ws) -> int:
    """
    Return max(existing IDs) + 1.

    IDs are NEVER reused even after /remove deletes a row, because IDs may be
    referenced in bot logs, user screenshots, or external systems. Gaps in the
    sequence (e.g. 1, 2, 4, 5 after item 3 is removed) are normal and expected.
    """
    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[_C["Item_ID"]] is not None:
            try:
                max_id = max(max_id, int(row[_C["Item_ID"]]))
            except (ValueError, TypeError):
                pass
    return max_id + 1


def _migrate_add_region(wb: openpyxl.Workbook) -> None:
    """Insert the Region column after Condition (in-place, no data loss). Defaults to ""."""
    ws = wb.active
    condition_col_1 = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(1, col).value == "Condition":
            condition_col_1 = col
            break
    if condition_col_1 is None:
        return
    insert_col = condition_col_1 + 1
    ws.insert_cols(insert_col)
    hc = ws.cell(row=1, column=insert_col, value="Region")
    hc.font      = Font(bold=True, color=_HEADER_FG)
    hc.fill      = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
    hc.alignment = Alignment(horizontal="center")
    ws.column_dimensions[get_column_letter(insert_col)].width = 8
    # Default all existing data rows to "" (TOTAL rows with Item_ID=None get "" too, harmless)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=insert_col)
        if cell.value is None:
            cell.value = ""


def _migrate_add_potential_profit(wb: openpyxl.Workbook) -> None:
    """Insert the Potential_Profit column after Live_Price (in-place, no data loss)."""
    ws = wb.active
    live_col_1 = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(1, col).value == "Live_Price":
            live_col_1 = col
            break
    if live_col_1 is None:
        return
    insert_col = live_col_1 + 1
    ws.insert_cols(insert_col)
    hc = ws.cell(row=1, column=insert_col, value="Potential_Profit")
    hc.font      = Font(bold=True, color=_HEADER_FG)
    hc.fill      = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
    hc.alignment = Alignment(horizontal="center")
    ws.column_dimensions[get_column_letter(insert_col)].width = 15


def _migrate_add_quick_price(wb: openpyxl.Workbook) -> None:
    """Insert Quick_Price column after Live_Price (in-place, no data loss). Defaults to None."""
    ws = wb.active
    live_col_1 = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(1, col).value == "Live_Price":
            live_col_1 = col
            break
    if live_col_1 is None:
        return
    insert_col = live_col_1 + 1
    ws.insert_cols(insert_col)
    hc = ws.cell(row=1, column=insert_col, value="Quick_Price")
    hc.font      = Font(bold=True, color=_HEADER_FG)
    hc.fill      = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
    hc.alignment = Alignment(horizontal="center")
    ws.column_dimensions[get_column_letter(insert_col)].width = 12
    # Existing rows default to None — populated on the next /update cycle


def _migrate_add_ebay_columns(wb: openpyxl.Workbook) -> None:
    """
    Insert eBay_Listing_ID and eBay_Listed after Status (in-place, no data loss).
    Handles partial migrations: inserts only the column(s) that are missing.
    Defaults all existing data rows to "".
    """
    ws = wb.active
    header_map = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}

    if "eBay_Listing_ID" in header_map and "eBay_Listed" in header_map:
        return  # already fully migrated

    # Determine the column after which to start inserting
    if "eBay_Listing_ID" in header_map:
        insert_after = header_map["eBay_Listing_ID"]
        cols_to_add  = [("eBay_Listed", 10)]
    else:
        insert_after = header_map.get("Status")
        if insert_after is None:
            return
        cols_to_add = [("eBay_Listing_ID", 18)]
        if "eBay_Listed" not in header_map:
            cols_to_add.append(("eBay_Listed", 10))

    ws.insert_cols(insert_after + 1, len(cols_to_add))
    for i, (name, width) in enumerate(cols_to_add):
        col_idx = insert_after + 1 + i
        hc = ws.cell(row=1, column=col_idx, value=name)
        hc.font      = Font(bold=True, color=_HEADER_FG)
        hc.fill      = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
        hc.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                cell.value = ""


def _migrate_add_quantity(wb: openpyxl.Workbook) -> None:
    """Insert Quantity column after Status, before eBay_Listing_ID. Defaults to 1."""
    ws = wb.active
    header_map = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
    status_col = header_map.get("Status")
    if status_col is None:
        return
    insert_col = status_col + 1
    ws.insert_cols(insert_col)
    hc = ws.cell(row=1, column=insert_col, value="Quantity")
    hc.font      = Font(bold=True, color=_HEADER_FG)
    hc.fill      = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
    hc.alignment = Alignment(horizontal="center")
    ws.column_dimensions[get_column_letter(insert_col)].width = 8
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=insert_col)
        if cell.value is None:
            cell.value = 1


def _migrate_add_date_columns(wb: openpyxl.Workbook) -> None:
    """Append Date_Added and Date_Sold after eBay_Listed. Defaults existing rows to ''."""
    ws = wb.active
    header_map = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}

    cols_to_add = []
    if "Date_Added" not in header_map:
        cols_to_add.append(("Date_Added", 12))
    if "Date_Sold" not in header_map:
        cols_to_add.append(("Date_Sold", 12))

    if not cols_to_add:
        return

    # Append after the current last column
    start_col = ws.max_column + 1
    for i, (name, width) in enumerate(cols_to_add):
        col_idx = start_col + i
        hc = ws.cell(row=1, column=col_idx, value=name)
        hc.font      = Font(bold=True, color=_HEADER_FG)
        hc.fill      = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
        hc.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                cell.value = ""


def _migrate_add_price_verified(wb: openpyxl.Workbook) -> None:
    """Append Price_Verified column after Date_Sold. Defaults all existing rows to True."""
    ws = wb.active
    insert_col = ws.max_column + 1
    hc = ws.cell(row=1, column=insert_col, value="Price_Verified")
    hc.font      = Font(bold=True, color=_HEADER_FG)
    hc.fill      = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
    hc.alignment = Alignment(horizontal="center")
    ws.column_dimensions[get_column_letter(insert_col)].width = 14
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=insert_col)
        if cell.value is None:
            cell.value = True


def _migrate_add_source(wb: openpyxl.Workbook) -> None:
    """Append Source column after Price_Verified. Defaults all existing rows to ""."""
    ws = wb.active
    insert_col = ws.max_column + 1
    hc = ws.cell(row=1, column=insert_col, value="Source")
    hc.font      = Font(bold=True, color=_HEADER_FG)
    hc.fill      = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
    hc.alignment = Alignment(horizontal="center")
    ws.column_dimensions[get_column_letter(insert_col)].width = 14
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=insert_col)
        if cell.value is None:
            cell.value = ""


def _migrate_add_image_urls(wb: openpyxl.Workbook) -> None:
    """Insert Image_URLs column after Vinted_Listing_URL. Defaults all existing rows to ""."""
    ws = wb.active
    header_map = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
    vinted_url_col = header_map.get("Vinted_Listing_URL")
    if vinted_url_col is None:
        return
    insert_col = vinted_url_col + 1
    ws.insert_cols(insert_col)
    hc = ws.cell(row=1, column=insert_col, value="Image_URLs")
    hc.font      = Font(bold=True, color=_HEADER_FG)
    hc.fill      = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
    hc.alignment = Alignment(horizontal="center")
    ws.column_dimensions[get_column_letter(insert_col)].width = 60
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=insert_col)
        if cell.value is None:
            cell.value = ""


def _migrate_add_vinted_columns(wb: openpyxl.Workbook) -> None:
    """
    Insert Vinted_Listed and Vinted_Listing_URL after eBay_Listed (in-place, no data loss).
    Handles partial migrations: inserts only the column(s) that are missing.
    Defaults all existing data rows to "".
    """
    ws = wb.active
    header_map = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}

    if "Vinted_Listed" in header_map and "Vinted_Listing_URL" in header_map:
        return

    if "Vinted_Listed" in header_map:
        insert_after = header_map["Vinted_Listed"]
        cols_to_add  = [("Vinted_Listing_URL", 30)]
    else:
        insert_after = header_map.get("eBay_Listed")
        if insert_after is None:
            return
        cols_to_add = [("Vinted_Listed", 10)]
        if "Vinted_Listing_URL" not in header_map:
            cols_to_add.append(("Vinted_Listing_URL", 30))

    ws.insert_cols(insert_after + 1, len(cols_to_add))
    for i, (name, width) in enumerate(cols_to_add):
        col_idx = insert_after + 1 + i
        hc = ws.cell(row=1, column=col_idx, value=name)
        hc.font      = Font(bold=True, color=_HEADER_FG)
        hc.fill      = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
        hc.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                cell.value = ""


def _backfill_missing_dates(wb: openpyxl.Workbook) -> int:
    """
    One-time backfill for rows that predate date tracking.
    Sets Date_Added = 2026-05-01 for all rows missing it.
    Sets Date_Sold  = 2026-05-31 for Sold rows missing a sell date.
    Safe to run on every startup — only touches blank cells.
    Returns the number of cells changed.
    """
    ws = wb.active
    changed = 0
    for row in ws.iter_rows(min_row=2):
        if row[_C["Item_ID"]].value is None:
            continue
        status          = str(row[_C["Status"]].value or "")
        date_added_cell = row[_C["Date_Added"]]
        date_sold_cell  = row[_C["Date_Sold"]]

        if not date_added_cell.value:
            date_added_cell.value = "2026-05-01"
            changed += 1

        if not date_sold_cell.value and status == "Sold":
            date_sold_cell.value = "2026-05-31"
            changed += 1

    if changed:
        print(f"[excel_db] Backfilled dates for {changed} cell(s) (assumed May 2026)")
    return changed


def backfill_missing_sold_dates() -> int:
    """
    One-time fix for sold items missing Date_Added or Date_Sold.
    Sets Date_Added → 2026-05-01, Date_Sold → 2026-05-31 for blank cells.
    Safe to run repeatedly — only touches blank cells on Sold rows.
    """
    wb = _load()
    ws = wb.active
    updated = 0
    for row in ws.iter_rows(min_row=2):
        if row[_C["Item_ID"]].value is None:
            continue
        if str(row[_C["Status"]].value or "") != "Sold":
            continue
        if not row[_C["Date_Added"]].value:
            row[_C["Date_Added"]].value = "2026-05-01"
            updated += 1
        if not row[_C["Date_Sold"]].value:
            row[_C["Date_Sold"]].value = "2026-05-31"
            updated += 1
    if updated:
        _save_with_totals(wb)
        print(f"[excel_db] backfill_missing_sold_dates: fixed {updated} cell(s)")
    return updated


def is_valid_ebay_listing_id(value: str) -> bool:
    """eBay listing IDs are 9-13 digit numbers."""
    if not value:
        return False
    stripped = value.strip()
    return stripped.isdigit() and 9 <= len(stripped) <= 13


def _migrate_sync_ebay_listed(ws) -> int:
    """
    For every row where eBay_Listing_ID is a valid listing ID but eBay_Listed != "Yes",
    set eBay_Listed = "Yes". Handles IDs typed manually into Excel.
    Returns the number of rows fixed.
    """
    fixed = 0
    for row in ws.iter_rows(min_row=2):
        if row[_C["Item_ID"]].value is None:
            continue
        listing_id = str(row[_C["eBay_Listing_ID"]].value or "").strip()
        if is_valid_ebay_listing_id(listing_id) and row[_C["eBay_Listed"]].value != "Yes":
            row[_C["eBay_Listed"]].value = "Yes"
            fixed += 1
    if fixed:
        print(f"[excel_db] Synced eBay_Listed flag for {fixed} rows with manual listing IDs.")
    return fixed


def fix_url_encoded_names() -> int:
    """One-time cleanup: decode URL-encoded characters in existing Card_Name values
    and fix str.title() apostrophe quirk ('S -> 's)."""
    import re as _re

    def _clean(name: str) -> str:
        decoded = urllib.parse.unquote(name)
        return _re.sub(r"(?<=\w)'([A-Z])", lambda m: "'" + m.group(1).lower(), decoded)

    wb = _load()
    ws = wb.active
    fixed = 0
    for row in ws.iter_rows(min_row=2):
        name_cell = row[_C["Card_Name"]]
        name = name_cell.value
        if not name:
            continue
        cleaned = _clean(str(name))
        if cleaned != str(name):
            name_cell.value = cleaned
            fixed += 1
    if fixed:
        _save_with_totals(wb)
        print(f"[excel_db] Fixed {fixed} URL-encoded card name(s)")
    return fixed


def _recompute_potential_profit(ws) -> None:
    """Recompute Potential_Profit for every Inventory row. Called inside _save_with_totals."""
    for row in ws.iter_rows(min_row=2):
        if row[_C["Item_ID"]].value is None:
            continue
        if row[_C["Status"]].value == "Inventory":
            buy  = row[_C["Purchase_Price"]].value or 0.0
            live = row[_C["Live_Price"]].value
            row[_C["Potential_Profit"]].value = round(live - buy, 2) if live is not None else None
        else:
            row[_C["Potential_Profit"]].value = None


def _find_row(ws, item_id: int):
    """
    Return the tuple of cells (not values) for item_id, or None if not found.
    TOTAL summary rows (Item_ID is None) are naturally skipped.
    """
    for excel_row in ws.iter_rows(min_row=2):
        if excel_row[_C["Item_ID"]].value == item_id:
            return excel_row
    return None


def _strip_summary_rows(ws) -> None:
    """
    Remove TOTAL summary rows and trim any trailing empty rows.

    Must be called *before* ws.append() so that new data never lands past the
    summary block. _save_with_totals calls this too, so all mutating paths are
    covered. The trailing-empty-row trim self-heals historically corrupted files.
    """
    # Delete rows whose Card_Name starts with "TOTAL"
    to_delete = [
        idx
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2)
        if isinstance(row[_C["Card_Name"]], str)
        and row[_C["Card_Name"]].startswith("TOTAL")
    ]
    for idx in reversed(to_delete):
        ws.delete_rows(idx)

    # Trim trailing empty rows left by historical corruption
    while ws.max_row > 1:
        last_row = [ws.cell(ws.max_row, c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or v == "" for v in last_row):
            ws.delete_rows(ws.max_row)
        else:
            break


def _save_with_totals(wb: openpyxl.Workbook) -> None:
    """
    Recompute inventory-aggregate rows and save the workbook.

    Four labelled summary rows are written two rows below the last data row.
    Rows with Item_ID == None (i.e. summary rows) are skipped by all CRUD
    iterators so they never interfere with get_stock / get_summary.
    """
    ws = wb.active

    # ── 1. Remove pre-existing summary rows and trailing empty rows ───────────
    _strip_summary_rows(ws)

    # ── 2. Recompute Potential_Profit for all Inventory rows ──────────────────
    _recompute_potential_profit(ws)

    # ── 3. Compute totals ─────────────────────────────────────────────────────
    total_cost             = 0.0
    total_potential        = 0.0
    total_potential_profit = 0.0
    total_sold_revenue     = 0.0
    total_lifetime_cost    = 0.0
    total_lifetime_profit  = 0.0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[_C["Item_ID"]] is None:
            continue
        total_lifetime_cost += row[_C["Purchase_Price"]] or 0.0
        if row[_C["Status"]] == "Inventory":
            total_cost             += row[_C["Purchase_Price"]]   or 0.0
            total_potential        += row[_C["Live_Price"]]       or 0.0
            total_potential_profit += row[_C["Potential_Profit"]] or 0.0
        else:
            total_sold_revenue    += row[_C["Sell_Price"]] or 0.0
            total_lifetime_profit += row[_C["Profit"]]     or 0.0

    # ── 4. Append summary rows two rows below last data row ───────────────────
    summary_start = ws.max_row + 2
    bold = Font(bold=True)

    for offset, (label, col_key, value) in enumerate([
        ("TOTAL COST",              "Purchase_Price",   round(total_cost, 2)),
        ("TOTAL POTENTIAL",         "Live_Price",       round(total_potential, 2)),
        ("TOTAL POTENTIAL PROFIT",  "Potential_Profit", round(total_potential_profit, 2)),
        ("TOTAL SOLD REVENUE",      "Sell_Price",       round(total_sold_revenue, 2)),
        ("TOTAL LIFETIME COST",     "Purchase_Price",   round(total_lifetime_cost, 2)),
        ("TOTAL LIFETIME PROFIT",   "Profit",           round(total_lifetime_profit, 2)),
        ("TOTAL LIFETIME POTENTIAL","Potential_Profit", round(total_lifetime_profit + total_potential_profit, 2)),
    ]):
        r = summary_start + offset
        lc = ws.cell(row=r, column=_C["Card_Name"] + 1, value=label)
        lc.font = bold
        vc = ws.cell(row=r, column=_C[col_key] + 1, value=value)
        vc.font = bold

    wb.save(config.EXCEL_FILE)


# ---------------------------------------------------------------------------
# Public CRUD
# ---------------------------------------------------------------------------

def add_item(
    card_name: str,
    pc_url: str,
    condition: str,
    region: str,
    purchase_price: float,
    live_price: Optional[float],
    quantity: int = 1,
    source: str = "",
) -> int:
    """
    Append a new row and return the generated Item_ID.
    Status is always set to 'Inventory' on creation.
    The quantity parameter is accepted but ignored — each copy gets its own row
    via the bot. Quantity=1 is always stored per row.
    """
    wb = _load()
    ws = wb.active

    _strip_summary_rows(ws)   # must precede append so the row lands after real data
    item_id  = _get_next_id(ws)
    row_data = [None] * len(COLUMNS)

    row_data[_C["Item_ID"]]          = item_id
    row_data[_C["Card_Name"]]        = card_name
    row_data[_C["PC_URL"]]           = pc_url
    row_data[_C["Condition"]]        = condition
    row_data[_C["Region"]]           = region or ""
    row_data[_C["Purchase_Price"]]   = round(purchase_price, 2)
    row_data[_C["Live_Price"]]       = round(live_price, 2) if live_price is not None else None
    row_data[_C["Quick_Price"]]      = None
    row_data[_C["Sell_Price"]]       = None
    row_data[_C["Profit"]]           = None
    row_data[_C["Status"]]           = "Inventory"
    row_data[_C["Quantity"]]         = 1
    row_data[_C["eBay_Listing_ID"]]      = ""
    row_data[_C["eBay_Listed"]]          = ""
    row_data[_C["Vinted_Listed"]]        = ""
    row_data[_C["Vinted_Listing_URL"]]   = ""
    row_data[_C["Image_URLs"]]           = ""
    row_data[_C["Date_Added"]]           = datetime.now().strftime("%Y-%m-%d")
    row_data[_C["Date_Sold"]]            = ""
    row_data[_C["Price_Verified"]]       = True
    if "Source" in _C:
        row_data[_C["Source"]]           = source or ""

    ws.append(row_data)
    _save_with_totals(wb)
    return item_id


def sell_item(item_id: int, sell_price: float) -> dict:
    """
    Mark an item as Sold, record the sell price, and calculate profit.

    Returns
    -------
    dict with keys: card_name, purchase_price, sell_price, profit

    Raises
    ------
    ValueError  if the item is not found or is already marked Sold.
    """
    wb = _load()
    ws = wb.active

    excel_row = _find_row(ws, item_id)
    if excel_row is None:
        raise ValueError(f"Item ID {item_id} not found in inventory.")

    if excel_row[_C["Status"]].value == "Sold":
        raise ValueError(f"Item {item_id} is already marked as Sold.")

    card_name      = excel_row[_C["Card_Name"]].value
    purchase_price = excel_row[_C["Purchase_Price"]].value or 0.0
    profit         = round(sell_price - purchase_price, 2)

    excel_row[_C["Sell_Price"]].value = round(sell_price, 2)
    excel_row[_C["Profit"]].value     = profit
    excel_row[_C["Status"]].value     = "Sold"
    excel_row[_C["Date_Sold"]].value  = datetime.now().strftime("%Y-%m-%d")

    _save_with_totals(wb)
    return {
        "card_name":      card_name,
        "purchase_price": purchase_price,
        "sell_price":     round(sell_price, 2),
        "profit":         profit,
    }


def unsell_item(item_id: int) -> dict:
    """
    Revert a Sold item back to Inventory (e.g. after an eBay order cancellation).
    Clears sell price, profit, and date sold.

    Raises
    ------
    ValueError  if the item is not found or is not currently marked Sold.
    """
    wb = _load()
    ws = wb.active

    excel_row = _find_row(ws, item_id)
    if excel_row is None:
        raise ValueError(f"Item ID {item_id} not found in inventory.")

    if excel_row[_C["Status"]].value != "Sold":
        raise ValueError(f"Item {item_id} is not marked as Sold.")

    card_name = excel_row[_C["Card_Name"]].value

    excel_row[_C["Status"]].value     = "Inventory"
    excel_row[_C["Sell_Price"]].value = None
    excel_row[_C["Profit"]].value     = None
    excel_row[_C["Date_Sold"]].value  = None

    _save_with_totals(wb)
    return {"card_name": card_name}


# Editable fields: command_name → (column_key, type_cast)
_EDITABLE_FIELDS: dict[str, tuple[str, type]] = {
    "card_name":      ("Card_Name",      str),
    "purchase_price": ("Purchase_Price", float),
    "pc_url":         ("PC_URL",         str),
    "condition":      ("Condition",      str),
    "region":         ("Region",         str),
    "sell_price":     ("Sell_Price",     float),
    "ebay_listed":    ("eBay_Listed",    str),
}


def edit_item(item_id: int, field: str, value: str) -> tuple:
    """Update a single field. Returns (old_value, new_value). Raises ValueError if not found."""
    field_info = _EDITABLE_FIELDS.get(field)
    if field_info is None:
        raise ValueError(f"Unknown field {field!r}. Valid: {', '.join(_EDITABLE_FIELDS)}")
    col_key, type_cast = field_info

    wb = _load()
    ws = wb.active
    excel_row = _find_row(ws, item_id)
    if excel_row is None:
        raise ValueError(f"Item ID {item_id} not found in inventory.")

    old_value = excel_row[_C[col_key]].value
    try:
        new_value = type_cast(value)
    except (ValueError, TypeError) as exc:
        raise ValueError(f"Cannot convert {value!r} to {type_cast.__name__}: {exc}") from exc

    if field == "ebay_listed" and new_value not in ("Yes", "No"):
        raise ValueError(f"ebay_listed must be 'Yes' or 'No', got {new_value!r}")

    excel_row[_C[col_key]].value = new_value

    # Keep derived columns consistent after price edits
    if field == "purchase_price":
        live = excel_row[_C["Live_Price"]].value
        if live is not None:
            excel_row[_C["Potential_Profit"]].value = round(float(live) - new_value, 2)
        if excel_row[_C["Status"]].value == "Sold":
            sell = excel_row[_C["Sell_Price"]].value or 0.0
            excel_row[_C["Profit"]].value = round(float(sell) - new_value, 2)
    elif field == "sell_price" and excel_row[_C["Status"]].value == "Sold":
        purchase = excel_row[_C["Purchase_Price"]].value or 0.0
        excel_row[_C["Profit"]].value = round(new_value - float(purchase), 2)

    _save_with_totals(wb)
    return old_value, new_value


def get_item(item_id: int) -> dict:
    """
    Return all fields for a single item as a dict.

    Raises
    ------
    ValueError  if no row with the given Item_ID exists.
    """
    wb = _load()
    ws = wb.active

    cells = _find_row(ws, item_id)
    if cells is None:
        raise ValueError(f"Item ID {item_id} not found in inventory.")

    _img_raw = cells[_C["Image_URLs"]].value if "Image_URLs" in _C else ""
    return {
        "item_id":            cells[_C["Item_ID"]].value,
        "card_name":          cells[_C["Card_Name"]].value,
        "pc_url":             cells[_C["PC_URL"]].value,
        "condition":          cells[_C["Condition"]].value,
        "region":             cells[_C["Region"]].value or "",
        "purchase_price":     cells[_C["Purchase_Price"]].value,
        "live_price":         cells[_C["Live_Price"]].value,
        "quick_price":        cells[_C["Quick_Price"]].value or None,
        "potential_profit":   cells[_C["Potential_Profit"]].value,
        "sell_price":         cells[_C["Sell_Price"]].value,
        "profit":             cells[_C["Profit"]].value,
        "status":             cells[_C["Status"]].value,
        "quantity":           cells[_C["Quantity"]].value or 1,
        "ebay_listing_id":    cells[_C["eBay_Listing_ID"]].value or "",
        "ebay_listed":        cells[_C["eBay_Listed"]].value or "",
        "vinted_listed":      cells[_C["Vinted_Listed"]].value or "",
        "vinted_listing_url": cells[_C["Vinted_Listing_URL"]].value or "",
        "image_urls":         [u for u in str(_img_raw or "").split("|") if u] if _img_raw else [],
        "date_added":         cells[_C["Date_Added"]].value or "",
        "date_sold":          cells[_C["Date_Sold"]].value or "",
        "price_verified":     cells[_C["Price_Verified"]].value is not False,
        "source":             cells[_C["Source"]].value or "" if "Source" in _C else "",
    }


def get_stock(status_filter: Optional[str] = "Inventory") -> list[dict]:
    """
    Return all rows matching status_filter (default: Inventory).
    Pass status_filter=None to return every row.
    """
    wb = _load()
    ws = wb.active

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[_C["Item_ID"]] is None:
            continue
        if status_filter and row[_C["Status"]] != status_filter:
            continue
        _img_raw = (row[_C["Image_URLs"]] if "Image_URLs" in _C and _C["Image_URLs"] < len(row) else None) or ""
        items.append({
            "item_id":            row[_C["Item_ID"]],
            "card_name":          row[_C["Card_Name"]],
            "pc_url":             row[_C["PC_URL"]],
            "condition":          row[_C["Condition"]],
            "region":             row[_C["Region"]] or "",
            "purchase_price":     row[_C["Purchase_Price"]],
            "live_price":         row[_C["Live_Price"]],
            "quick_price":        row[_C["Quick_Price"]] or None,
            "potential_profit":   row[_C["Potential_Profit"]],
            "sell_price":         row[_C["Sell_Price"]],
            "profit":             row[_C["Profit"]],
            "status":             row[_C["Status"]],
            "quantity":           row[_C["Quantity"]] or 1,
            "ebay_listing_id":    row[_C["eBay_Listing_ID"]] or "",
            "ebay_listed":        row[_C["eBay_Listed"]] or "",
            "vinted_listed":      row[_C["Vinted_Listed"]] or "",
            "vinted_listing_url": row[_C["Vinted_Listing_URL"]] or "",
            "image_urls":         [u for u in str(_img_raw).split("|") if u] if _img_raw else [],
            "date_added":         row[_C["Date_Added"]] or "",
            "date_sold":          row[_C["Date_Sold"]] or "",
            "price_verified":     row[_C["Price_Verified"]] is not False,
            "source":             (row[_C["Source"]] if "Source" in _C and _C["Source"] < len(row) else None) or "",
        })
    return items


def update_live_price(item_id: int, live_price_gbp: float, old_price: Optional[float] = None) -> None:
    """
    Refresh the Live_Price column for an existing inventory item.
    Sets Price_Verified = False when the price actually changed (old_price provided and differs).

    Raises
    ------
    ValueError  if item_id is not found.
    """
    wb = _load()
    ws = wb.active

    excel_row = _find_row(ws, item_id)
    if excel_row is None:
        raise ValueError(f"Item ID {item_id} not found in inventory.")

    new_price = round(live_price_gbp, 2)
    excel_row[_C["Live_Price"]].value = new_price

    if old_price is not None and new_price != round(old_price, 2):
        excel_row[_C["Price_Verified"]].value = False

    _save_with_totals(wb)


def update_quick_price(item_id: int, price: Optional[float]) -> None:
    """
    Update the Quick_Price column for an existing inventory item.

    Raises
    ------
    ValueError  if item_id is not found.
    """
    wb = _load()
    ws = wb.active
    excel_row = _find_row(ws, item_id)
    if excel_row is None:
        raise ValueError(f"Item ID {item_id} not found in inventory.")
    excel_row[_C["Quick_Price"]].value = round(price, 2) if price is not None else None
    _save_with_totals(wb)


def update_sell_price(item_id: int, price: float) -> None:
    wb = _load()
    ws = wb.active
    excel_row = _find_row(ws, item_id)
    if excel_row is None:
        raise ValueError(f"Item ID {item_id} not found in inventory.")
    excel_row[_C["Sell_Price"]].value = round(price, 2)
    _save_with_totals(wb)


async def update_sell_price_async(item_id: int, price: float) -> None:
    async with _lock:
        await asyncio.to_thread(update_sell_price, item_id, price)


def update_item_listing(item_id: int, listing_id: str, sell_price: float) -> None:
    """Mark item as eBay listed with listing ID and proportional sell price."""
    wb = _load()
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[_C["Item_ID"]].value == item_id:
            row[_C["eBay_Listed"]].value = "Yes"
            row[_C["eBay_Listing_ID"]].value = listing_id
            row[_C["Sell_Price"]].value = round(sell_price, 2)
            _save_with_totals(wb)
            return
    raise ValueError(f"Item {item_id} not found")

async def update_item_listing_async(item_id: int, listing_id: str, sell_price: float) -> None:
    async with _lock:
        await asyncio.to_thread(update_item_listing, item_id, listing_id, sell_price)


def remove_item(item_id: int) -> dict:
    """
    Permanently delete an Inventory row from the workbook.

    Returns a dict with the removed item's details for the bot to echo back.

    Raises
    ------
    ValueError  if item_id is not found, or if the item is already Sold
                (sold rows are protected; edit the file directly if truly needed).
    """
    wb = _load()
    ws = wb.active

    _strip_summary_rows(ws)

    excel_row = _find_row(ws, item_id)
    if excel_row is None:
        raise ValueError(f"Item ID {item_id} not found in inventory.")
    if excel_row[_C["Status"]].value == "Sold":
        raise ValueError(f"Item {item_id} is marked Sold and cannot be removed.")

    removed = {
        "item_id":        excel_row[_C["Item_ID"]].value,
        "card_name":      excel_row[_C["Card_Name"]].value,
        "condition":      excel_row[_C["Condition"]].value,
        "region":         excel_row[_C["Region"]].value or "",
        "purchase_price": excel_row[_C["Purchase_Price"]].value,
        "live_price":     excel_row[_C["Live_Price"]].value,
    }

    ws.delete_rows(excel_row[0].row)
    _save_with_totals(wb)
    return removed


def mark_ebay_listed(item_id: int, ebay_listing_id: str) -> None:
    """Record that an item has been published to eBay."""
    wb = _load()
    ws = wb.active
    excel_row = _find_row(ws, item_id)
    if excel_row is None:
        raise ValueError(f"Item ID {item_id} not found in inventory.")
    excel_row[_C["eBay_Listing_ID"]].value = ebay_listing_id
    excel_row[_C["eBay_Listed"]].value     = "Yes"
    _save_with_totals(wb)


def mark_ebay_delisted(item_id: int) -> None:
    """Clear the eBay listing record when a listing is ended."""
    wb = _load()
    ws = wb.active
    excel_row = _find_row(ws, item_id)
    if excel_row is None:
        raise ValueError(f"Item ID {item_id} not found in inventory.")
    excel_row[_C["eBay_Listing_ID"]].value = ""
    excel_row[_C["eBay_Listed"]].value     = ""
    _save_with_totals(wb)


def mark_vinted_listed(item_id: int, listing_url: str) -> None:
    """Record that an item has been published to Vinted."""
    wb = _load()
    ws = wb.active
    excel_row = _find_row(ws, item_id)
    if excel_row is None:
        raise ValueError(f"Item ID {item_id} not found in inventory.")
    excel_row[_C["Vinted_Listed"]].value       = "Yes"
    excel_row[_C["Vinted_Listing_URL"]].value  = listing_url or ""
    _save_with_totals(wb)


async def mark_vinted_listed_async(item_id: int, listing_url: str) -> None:
    async with _lock:
        await asyncio.to_thread(mark_vinted_listed, item_id, listing_url)


def get_summary() -> dict:
    """Return aggregate profit/loss stats across all items, including month-to-date fields."""
    wb = _load()
    ws = wb.active

    month_prefix = datetime.now().strftime("%Y-%m")

    total_invested                  = 0.0
    total_revenue                   = 0.0
    total_profit                    = 0.0
    total_lifetime_cost             = 0.0
    total_cost_in_stock             = 0.0
    total_potential_in_stock        = 0.0
    total_potential_profit_in_stock = 0.0
    sold_count                      = 0
    stock_count                     = 0

    mtd_sold_count   = 0
    mtd_revenue      = 0.0
    mtd_profit       = 0.0
    mtd_cost_of_sold = 0.0
    mtd_added_count  = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[_C["Item_ID"]] is None:
            continue
        buy = row[_C["Purchase_Price"]] or 0.0
        total_lifetime_cost += buy
        if row[_C["Status"]] == "Sold":
            sold_count     += 1
            total_invested += buy
            total_revenue  += (row[_C["Sell_Price"]] or 0.0)
            total_profit   += (row[_C["Profit"]]     or 0.0)
            date_sold = str(row[_C["Date_Sold"]] or "")
            if date_sold.startswith(month_prefix):
                mtd_sold_count   += 1
                mtd_revenue      += (row[_C["Sell_Price"]] or 0.0)
                mtd_profit       += (row[_C["Profit"]]     or 0.0)
                mtd_cost_of_sold += buy
        else:
            stock_count                     += 1
            total_cost_in_stock             += buy
            total_potential_in_stock        += (row[_C["Live_Price"]]       or 0.0)
            total_potential_profit_in_stock += (row[_C["Potential_Profit"]] or 0.0)

        date_added = str(row[_C["Date_Added"]] or "")
        if date_added.startswith(month_prefix):
            mtd_added_count += 1

    roi_pct             = (total_profit / total_lifetime_cost * 100) if total_lifetime_cost else 0
    avg_sell_price      = total_revenue / sold_count if sold_count else 0
    avg_profit_per_item = total_profit / sold_count if sold_count else 0
    avg_margin_pct      = (avg_profit_per_item / avg_sell_price * 100) if avg_sell_price else 0
    stock_roi_pct       = (total_potential_profit_in_stock / total_cost_in_stock * 100) if total_cost_in_stock else 0
    mtd_monthly_roi     = (mtd_profit / mtd_cost_of_sold * 100) if mtd_cost_of_sold else 0

    return {
        "in_stock":                        stock_count,
        "sold":                            sold_count,
        "total_invested":                  round(total_invested, 2),
        "total_revenue":                   round(total_revenue, 2),
        "total_sold_revenue":              round(total_revenue, 2),  # alias used by footer
        "total_profit":                    round(total_profit, 2),
        "total_lifetime_cost":             round(total_lifetime_cost, 2),
        "total_lifetime_profit":           round(total_profit, 2),
        "total_lifetime_potential":        round(total_profit + total_potential_profit_in_stock, 2),
        "total_cost_in_stock":             round(total_cost_in_stock, 2),
        "total_potential_in_stock":        round(total_potential_in_stock, 2),
        "total_potential_profit_in_stock": round(total_potential_profit_in_stock, 2),
        "roi_pct":                         round(roi_pct, 1),
        "avg_margin_pct":                  round(avg_margin_pct, 1),
        "stock_roi_pct":                   round(stock_roi_pct, 1),
        "avg_sell_price":                  round(avg_sell_price, 2),
        "avg_profit":                      round(avg_profit_per_item, 2),
        "mtd_sold_count":                  mtd_sold_count,
        "mtd_revenue":                     round(mtd_revenue, 2),
        "mtd_profit":                      round(mtd_profit, 2),
        "mtd_cost_of_sold":                round(mtd_cost_of_sold, 2),
        "mtd_monthly_roi":                 round(mtd_monthly_roi, 1),
        "mtd_added_count":                 mtd_added_count,
    }


def get_best_sellers(limit: int = 5) -> dict:
    """Returns top sold cards by absolute profit, margin %, and sales volume."""
    sold_items = [
        item for item in get_stock(status_filter="Sold")
        if float(item.get("sell_price") or 0) > 0
        and float(item.get("purchase_price") or 0) > 0
    ]

    by_profit = sorted(
        sold_items,
        key=lambda i: float(i.get("sell_price", 0)) - float(i.get("purchase_price", 0)),
        reverse=True,
    )[:limit]

    def _margin_pct(i):
        sell = float(i.get("sell_price") or 0)
        buy  = float(i.get("purchase_price") or 0)
        return ((sell - buy) / sell * 100) if sell else 0

    by_margin = sorted(sold_items, key=_margin_pct, reverse=True)[:limit]

    name_counts   = Counter(i["card_name"] for i in sold_items)
    top_by_volume = name_counts.most_common(limit)

    return {
        "by_profit": by_profit,
        "by_margin": by_margin,
        "by_volume": top_by_volume,
    }


# ---------------------------------------------------------------------------
# Price history
# ---------------------------------------------------------------------------

def append_price_observation(
    item_id: int,
    card_name: str,
    region: str,
    price_gbp: float,
    source: str = "pricecharting",
) -> None:
    """
    Append one row to the Price_History sheet with the current timestamp.

    History is strictly append-only — no backfill for existing items.
    Records start from when this function is first called after deployment.
    Saves directly with wb.save() (not _save_with_totals) since this sheet
    has no totals and is independent of the Inventory sheet.
    """
    wb = _load()
    _ensure_price_history_sheet(wb)   # defensive: create sheet if somehow absent
    ws = wb[_PH_SHEET]
    ws.append([
        datetime.now().isoformat(timespec="seconds"),
        item_id,
        card_name,
        region or "",
        round(price_gbp, 2),
        source,
    ])
    wb.save(config.EXCEL_FILE)


# ---------------------------------------------------------------------------
# Monthly stats / reporting
# ---------------------------------------------------------------------------

def get_monthly_stats(year: int, month: int) -> dict:
    """Return sales and addition stats for a given calendar month."""
    month_prefix = f"{year}-{month:02d}"
    wb = _load()
    ws = wb.active

    sold_items: list[dict] = []
    added_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[_C["Item_ID"]] is None:
            continue
        date_added = str(row[_C["Date_Added"]] or "")
        if date_added.startswith(month_prefix):
            added_count += 1
        if row[_C["Status"]] == "Sold":
            date_sold = str(row[_C["Date_Sold"]] or "")
            if date_sold.startswith(month_prefix):
                profit = row[_C["Profit"]] or 0.0
                sold_items.append({
                    "item_id":        row[_C["Item_ID"]],
                    "card_name":      row[_C["Card_Name"]],
                    "condition":      row[_C["Condition"]],
                    "region":         row[_C["Region"]] or "",
                    "purchase_price": row[_C["Purchase_Price"]] or 0.0,
                    "sell_price":     row[_C["Sell_Price"]] or 0.0,
                    "profit":         profit,
                    "date_sold":      date_sold,
                })

    total_revenue = sum(i["sell_price"] for i in sold_items)
    total_cost    = sum(i["purchase_price"] for i in sold_items)
    total_profit  = sum(i["profit"] for i in sold_items)
    avg_profit    = total_profit / len(sold_items) if sold_items else 0.0

    best = max(sold_items, key=lambda i: i["profit"], default=None)
    worst = min(sold_items, key=lambda i: i["profit"], default=None)

    # Current stock snapshot
    stock_count = 0
    stock_cost  = 0.0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[_C["Item_ID"]] is None:
            continue
        if row[_C["Status"]] == "Inventory":
            stock_count += 1
            stock_cost  += row[_C["Purchase_Price"]] or 0.0

    return {
        "month_prefix":  month_prefix,
        "sold_count":    len(sold_items),
        "total_revenue": round(total_revenue, 2),
        "total_cost":    round(total_cost, 2),
        "total_profit":  round(total_profit, 2),
        "avg_profit":    round(avg_profit, 2),
        "added_count":   added_count,
        "best_name":     best["card_name"] if best else None,
        "best_profit":   best["profit"]    if best else None,
        "worst_name":    worst["card_name"] if worst else None,
        "worst_profit":  worst["profit"]    if worst else None,
        "stock_count":   stock_count,
        "stock_cost":    round(stock_cost, 2),
    }


_MS_HEADERS = ["Item_ID", "Card_Name", "Condition", "Region", "Purchase_Price", "Sell_Price", "Profit", "Date_Sold"]
_MS_WIDTHS  = [8, 36, 14, 8, 15, 12, 10, 12]


def ensure_monthly_sheet(year: int, month: int) -> bool:
    """
    Create a YYYY-MM summary sheet in the workbook if it doesn't already exist.
    Returns True if a new sheet was created, False if it already existed.
    """
    sheet_name   = f"{year}-{month:02d}"
    month_prefix = sheet_name
    wb = _load()
    if sheet_name in wb.sheetnames:
        return False

    ws = wb.create_sheet(sheet_name)

    # --- Section 1 header row ---
    for col, (h, w) in enumerate(zip(_MS_HEADERS, _MS_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color=_HEADER_FG)
        cell.fill      = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = w

    # --- Collect sold items for this month from the Inventory sheet ---
    inv_ws = wb.active
    sold_items: list[dict] = []
    for row in inv_ws.iter_rows(min_row=2, values_only=True):
        if row[_C["Item_ID"]] is None:
            continue
        if row[_C["Status"]] == "Sold":
            date_sold = str(row[_C["Date_Sold"]] or "")
            if date_sold.startswith(month_prefix):
                sold_items.append({
                    "item_id":        row[_C["Item_ID"]],
                    "card_name":      row[_C["Card_Name"]],
                    "condition":      row[_C["Condition"]] or "",
                    "region":         row[_C["Region"]] or "",
                    "purchase_price": row[_C["Purchase_Price"]] or 0.0,
                    "sell_price":     row[_C["Sell_Price"]] or 0.0,
                    "profit":         row[_C["Profit"]] or 0.0,
                    "date_sold":      date_sold,
                })

    for r_idx, item in enumerate(sold_items, 2):
        ws.cell(r_idx, 1, item["item_id"])
        ws.cell(r_idx, 2, item["card_name"])
        ws.cell(r_idx, 3, item["condition"])
        ws.cell(r_idx, 4, item["region"])
        ws.cell(r_idx, 5, item["purchase_price"])
        ws.cell(r_idx, 6, item["sell_price"])
        ws.cell(r_idx, 7, item["profit"])
        ws.cell(r_idx, 8, item["date_sold"])

    if not sold_items:
        ws.cell(2, 1, "No sales this month")

    # --- Section 2: Summary block two rows below sold items ---
    summary_row = (len(sold_items) + 2) + 2  # header row + data rows + blank row

    total_revenue = sum(i["sell_price"] for i in sold_items)
    total_cost    = sum(i["purchase_price"] for i in sold_items)
    total_profit  = sum(i["profit"] for i in sold_items)
    avg_profit    = total_profit / len(sold_items) if sold_items else 0.0

    # Current stock snapshot
    stock_count = 0
    stock_cost  = 0.0
    stock_potential = 0.0
    for row in inv_ws.iter_rows(min_row=2, values_only=True):
        if row[_C["Item_ID"]] is None:
            continue
        if row[_C["Status"]] == "Inventory":
            stock_count += 1
            stock_cost  += row[_C["Purchase_Price"]] or 0.0
            stock_potential += row[_C["Live_Price"]] or 0.0

    bold = Font(bold=True)
    summary_lines = [
        (f"Total items sold:",      str(len(sold_items))),
        (f"Total revenue:",         f"£{total_revenue:.2f}"),
        (f"Total cost of sold:",    f"£{total_cost:.2f}"),
        (f"Total profit:",          f"£{total_profit:.2f}"),
        (f"Avg profit per item:",   f"£{avg_profit:.2f}"),
        ("",                        ""),
        ("Total currently in stock:", str(stock_count)),
        ("Total cost in stock:",    f"£{stock_cost:.2f}"),
        ("Total potential value:",  f"£{stock_potential:.2f}"),
    ]
    for i, (label, value) in enumerate(summary_lines):
        lc = ws.cell(summary_row + i, 1, label)
        lc.font = bold
        ws.cell(summary_row + i, 2, value)

    wb.save(config.EXCEL_FILE)
    return True


# ---------------------------------------------------------------------------
# Async public API — always use these from bot.py
# ---------------------------------------------------------------------------
# Each wrapper acquires _lock for the full duration of the sync call so that
# no two operations can interleave their load→mutate→save cycles.
#
# Pattern for mutating wrappers:
#   1. snapshot_inventory()     — before lock (cheap, shouldn't block waiters)
#   2. async with _lock: ...    — atomic load→mutate→save
#   3. audit.log_mutation()     — after lock (pure file write, no lock needed)
#   4. append_price_observation_async()  — after lock (acquires lock itself;
#      must be OUTSIDE the main lock block to avoid asyncio.Lock deadlock)

async def init_excel_async() -> None:
    async with _lock:
        await asyncio.to_thread(init_excel)


async def add_item_async(
    card_name: str,
    pc_url: str,
    condition: str,
    region: str,
    purchase_price: float,
    live_price: Optional[float],
    quantity: int = 1,
    source: str = "",
) -> int:
    await backups.snapshot_inventory("pre-add")
    async with _lock:
        item_id = await asyncio.to_thread(
            add_item, card_name, pc_url, condition, region, purchase_price, live_price, quantity, source
        )
    audit.log_mutation("add", item_id, "added", {
        "card_name":      card_name,
        "condition":      condition,
        "region":         region,
        "purchase_price": purchase_price,
        "live_price":     live_price,
    })
    if live_price is not None:
        await append_price_observation_async(item_id, card_name, region, live_price)
    return item_id


async def sell_item_async(item_id: int, sell_price: float) -> dict:
    await backups.snapshot_inventory("pre-sell")
    async with _lock:
        result = await asyncio.to_thread(sell_item, item_id, sell_price)
    audit.log_mutation("sell", item_id, "sold", {
        "card_name":      result["card_name"],
        "purchase_price": result["purchase_price"],
        "sell_price":     result["sell_price"],
        "profit":         result["profit"],
    })
    return result


async def unsell_item_async(item_id: int) -> dict:
    await backups.snapshot_inventory("pre-unsell")
    async with _lock:
        result = await asyncio.to_thread(unsell_item, item_id)
    audit.log_mutation("unsell", item_id, "reverted_to_inventory", {
        "card_name": result["card_name"],
    })
    return result


async def get_item_async(item_id: int) -> dict:
    async with _lock:
        return await asyncio.to_thread(get_item, item_id)


async def get_stock_async(status_filter: Optional[str] = "Inventory") -> list[dict]:
    async with _lock:
        return await asyncio.to_thread(get_stock, status_filter)


async def update_live_price_async(
    item_id: int,
    live_price_gbp: float,
    *,
    card_name: str = "",
    region: str = "",
    old_price: Optional[float] = None,
) -> None:
    async with _lock:
        await asyncio.to_thread(update_live_price, item_id, live_price_gbp, old_price)
    audit.log_mutation("price_refresh", item_id, "price refreshed", {
        "old_price": old_price,
        "new_price": live_price_gbp,
        "region":    region,
    })
    if card_name:
        await append_price_observation_async(item_id, card_name, region, live_price_gbp)


async def update_quick_price_async(item_id: int, price: Optional[float]) -> None:
    async with _lock:
        await asyncio.to_thread(update_quick_price, item_id, price)


async def remove_item_async(item_id: int) -> dict:
    await backups.snapshot_inventory("pre-remove")
    async with _lock:
        removed = await asyncio.to_thread(remove_item, item_id)
    audit.log_mutation("remove", item_id, "removed", {
        "card_name":      removed["card_name"],
        "condition":      removed["condition"],
        "region":         removed["region"],
        "purchase_price": removed["purchase_price"],
        "live_price":     removed["live_price"],
    })
    return removed


async def append_price_observation_async(
    item_id: int,
    card_name: str,
    region: str,
    price_gbp: float,
    source: str = "pricecharting",
) -> None:
    async with _lock:
        await asyncio.to_thread(
            append_price_observation, item_id, card_name, region, price_gbp, source
        )


async def get_summary_async() -> dict:
    async with _lock:
        return await asyncio.to_thread(get_summary)


async def get_best_sellers_async(limit: int = 5) -> dict:
    async with _lock:
        return await asyncio.to_thread(get_best_sellers, limit)


async def get_monthly_stats_async(year: int, month: int) -> dict:
    async with _lock:
        return await asyncio.to_thread(get_monthly_stats, year, month)


async def ensure_monthly_sheet_async(year: int, month: int) -> bool:
    async with _lock:
        return await asyncio.to_thread(ensure_monthly_sheet, year, month)


async def mark_ebay_listed_async(item_id: int, ebay_listing_id: str) -> None:
    async with _lock:
        await asyncio.to_thread(mark_ebay_listed, item_id, ebay_listing_id)


async def mark_ebay_delisted_async(item_id: int) -> None:
    async with _lock:
        await asyncio.to_thread(mark_ebay_delisted, item_id)


async def edit_item_async(item_id: int, field: str, value: str) -> tuple:
    await backups.snapshot_inventory("pre-edit")
    async with _lock:
        old_val, new_val = await asyncio.to_thread(edit_item, item_id, field, value)
    audit.log_mutation("edit", item_id, "edited", {
        "field":     field,
        "old_value": str(old_val) if old_val is not None else "",
        "new_value": str(new_val),
    })
    return old_val, new_val


# ---------------------------------------------------------------------------
# Price verification
# ---------------------------------------------------------------------------

def get_unverified_price_changes() -> list[dict]:
    """Return all Inventory items with Price_Verified == False."""
    wb = _load()
    ws = wb.active
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[_C["Item_ID"]] is None:
            continue
        if row[_C["Status"]] != "Inventory":
            continue
        if row[_C["Price_Verified"]] is False:
            items.append({
                "item_id":        row[_C["Item_ID"]],
                "card_name":      row[_C["Card_Name"]],
                "live_price":     row[_C["Live_Price"]],
                "purchase_price": row[_C["Purchase_Price"]],
            })
    return items


def set_price_verified_bulk(item_ids: list[int], verified: bool = True) -> int:
    """Set Price_Verified for the given item IDs. Returns count of rows updated."""
    wb = _load()
    ws = wb.active
    id_set = set(item_ids)
    updated = 0
    for row in ws.iter_rows(min_row=2):
        if row[_C["Item_ID"]].value in id_set:
            row[_C["Price_Verified"]].value = verified
            updated += 1
    _save_with_totals(wb)
    return updated


def get_item_price_history(item_id: int, limit: int = 2) -> list[dict]:
    """Return the last `limit` price observations for item_id from Price_History."""
    wb = _load()
    if _PH_SHEET not in wb.sheetnames:
        return []
    ws = wb[_PH_SHEET]
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == item_id:
            entries.append({
                "timestamp":      row[0],
                "live_price_gbp": row[4],
            })
    return entries[-limit:]


async def get_unverified_price_changes_async() -> list[dict]:
    async with _lock:
        return await asyncio.to_thread(get_unverified_price_changes)


async def set_price_verified_bulk_async(item_ids: list[int], verified: bool = True) -> int:
    async with _lock:
        return await asyncio.to_thread(set_price_verified_bulk, item_ids, verified)


async def get_item_price_history_async(item_id: int, limit: int = 2) -> list[dict]:
    async with _lock:
        return await asyncio.to_thread(get_item_price_history, item_id, limit)


# ---------------------------------------------------------------------------
# Watchlist CRUD
# ---------------------------------------------------------------------------

def _get_next_watch_id(ws) -> int:
    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[_WL_C["Watch_ID"]] is not None:
            try:
                max_id = max(max_id, int(row[_WL_C["Watch_ID"]]))
            except (ValueError, TypeError):
                pass
    return max_id + 1


def get_watchlist() -> list[dict]:
    wb = _load()
    _ensure_watchlist_sheet(wb)
    ws = wb[_WL_SHEET]
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[_WL_C["Watch_ID"]] is None:
            continue
        entries.append({
            "watch_id":         row[_WL_C["Watch_ID"]],
            "card_name":        row[_WL_C["Card_Name"]],
            "pc_url":           row[_WL_C["PC_URL"]],
            "target_price_gbp": row[_WL_C["Target_Price_GBP"]],
            "current_price_gbp": row[_WL_C["Current_Price_GBP"]],
            "added_date":       row[_WL_C["Added_Date"]] or "",
            "alert_sent":       bool(row[_WL_C["Alert_Sent"]]),
        })
    return entries


def add_watchlist_entry(card_name: str, pc_url: str, target_price: float, current_price: Optional[float] = None) -> int:
    wb = _load()
    _ensure_watchlist_sheet(wb)
    ws = wb[_WL_SHEET]
    watch_id = _get_next_watch_id(ws)
    ws.append([
        watch_id,
        card_name,
        pc_url,
        round(target_price, 2),
        round(current_price, 2) if current_price is not None else None,
        datetime.now().strftime("%Y-%m-%d"),
        False,
    ])
    wb.save(config.EXCEL_FILE)
    return watch_id


def _find_watchlist_row(ws, watch_id: int):
    for row in ws.iter_rows(min_row=2):
        if row[_WL_C["Watch_ID"]].value == watch_id:
            return row
    return None


def update_watchlist_price(watch_id: int, current_price: float) -> None:
    wb = _load()
    _ensure_watchlist_sheet(wb)
    ws = wb[_WL_SHEET]
    row = _find_watchlist_row(ws, watch_id)
    if row:
        row[_WL_C["Current_Price_GBP"]].value = round(current_price, 2)
    wb.save(config.EXCEL_FILE)


def mark_watchlist_alerted(watch_id: int) -> None:
    wb = _load()
    _ensure_watchlist_sheet(wb)
    ws = wb[_WL_SHEET]
    row = _find_watchlist_row(ws, watch_id)
    if row:
        row[_WL_C["Alert_Sent"]].value = True
    wb.save(config.EXCEL_FILE)


def reset_watchlist_alert(watch_id: int) -> None:
    wb = _load()
    _ensure_watchlist_sheet(wb)
    ws = wb[_WL_SHEET]
    row = _find_watchlist_row(ws, watch_id)
    if row:
        row[_WL_C["Alert_Sent"]].value = False
    wb.save(config.EXCEL_FILE)


def remove_watchlist_entry(watch_id: int) -> None:
    wb = _load()
    _ensure_watchlist_sheet(wb)
    ws = wb[_WL_SHEET]
    row = _find_watchlist_row(ws, watch_id)
    if row is None:
        raise ValueError(f"Watch ID {watch_id} not found.")
    ws.delete_rows(row[0].row)
    wb.save(config.EXCEL_FILE)


async def get_watchlist_async() -> list[dict]:
    async with _lock:
        return await asyncio.to_thread(get_watchlist)


async def add_watchlist_entry_async(card_name: str, pc_url: str, target_price: float, current_price: Optional[float] = None) -> int:
    async with _lock:
        return await asyncio.to_thread(add_watchlist_entry, card_name, pc_url, target_price, current_price)


async def update_watchlist_price_async(watch_id: int, current_price: float) -> None:
    async with _lock:
        await asyncio.to_thread(update_watchlist_price, watch_id, current_price)


async def mark_watchlist_alerted_async(watch_id: int) -> None:
    async with _lock:
        await asyncio.to_thread(mark_watchlist_alerted, watch_id)


async def reset_watchlist_alert_async(watch_id: int) -> None:
    async with _lock:
        await asyncio.to_thread(reset_watchlist_alert, watch_id)


async def remove_watchlist_entry_async(watch_id: int) -> None:
    async with _lock:
        await asyncio.to_thread(remove_watchlist_entry, watch_id)


async def get_restock_suggestions() -> list[dict]:
    """
    Identify sets that sell fast but have low current stock.
    Requires at least 3 historical sales for a set to be included.
    Returns up to 20 suggestions sorted by priority score.
    """
    import re as _re
    from collections import defaultdict as _dd

    sold_items = await get_stock_async(status_filter="Sold")
    inv_items  = await get_stock_async(status_filter="Inventory")

    def _set_name(card_name: str) -> str:
        m = _re.search(r'\(Pokemon (.+?)\)', card_name or "")
        return m.group(1) if m else "Unknown"

    set_velocity: dict = _dd(list)
    for item in sold_items:
        date_added = item.get("date_added")
        date_sold  = item.get("date_sold")
        if not date_added or not date_sold:
            continue
        try:
            from datetime import datetime as _dt
            added  = _dt.strptime(str(date_added)[:10], "%Y-%m-%d")
            sold   = _dt.strptime(str(date_sold)[:10],  "%Y-%m-%d")
            days   = max((sold - added).days, 1)
            profit = float(item.get("sell_price") or 0) - float(item.get("purchase_price") or 0)
            set_velocity[_set_name(item.get("card_name", ""))].append({"days": days, "profit": profit})
        except Exception:
            continue

    set_stock: dict = _dd(int)
    for item in inv_items:
        set_stock[_set_name(item.get("card_name", ""))] += 1

    suggestions = []
    for set_name, sales in set_velocity.items():
        if len(sales) < 3:
            continue
        avg_days   = sum(s["days"]   for s in sales) / len(sales)
        avg_profit = sum(s["profit"] for s in sales) / len(sales)
        current_stock = set_stock.get(set_name, 0)
        if avg_days < 14 and current_stock < 5:
            suggestions.append({
                "set":            set_name,
                "avg_days":       round(avg_days, 1),
                "avg_profit":     round(avg_profit, 2),
                "current_stock":  current_stock,
                "total_sold":     len(sales),
                "priority_score": round((avg_profit / max(avg_days, 1)) * (5 - current_stock), 2),
            })

    suggestions.sort(key=lambda x: x["priority_score"], reverse=True)
    return suggestions[:20]


def get_velocity_stats(group_by: str = "set") -> list[dict]:
    """
    Calculate average days-to-sell grouped by set, condition, or price range.
    Only includes sold items with both Date_Added and Date_Sold populated.
    """
    import re
    from datetime import datetime
    from collections import defaultdict

    sold_items = get_stock(status_filter="Sold")

    groups:        dict[str, list[float]] = defaultdict(list)
    group_profits: dict[str, list[float]] = defaultdict(list)

    for item in sold_items:
        date_added = item.get("date_added")
        date_sold  = item.get("date_sold")
        if not date_added or not date_sold:
            continue
        try:
            added = datetime.strptime(str(date_added)[:10], "%Y-%m-%d")
            sold  = datetime.strptime(str(date_sold)[:10],  "%Y-%m-%d")
            days  = max((sold - added).days, 1)
        except ValueError:
            continue

        sell_price     = float(item.get("sell_price")     or 0)
        purchase_price = float(item.get("purchase_price") or 0)
        profit         = sell_price - purchase_price

        if group_by == "set":
            name = item.get("card_name", "")
            m    = re.search(r'\(Pokemon (.+?)\)', name)
            key  = m.group(1) if m else "Unknown"
        elif group_by == "condition":
            key = item.get("condition") or "Unknown"
        elif group_by == "price":
            if sell_price < 3:
                key = "Under £3"
            elif sell_price < 10:
                key = "£3–£10"
            elif sell_price < 25:
                key = "£10–£25"
            elif sell_price < 50:
                key = "£25–£50"
            else:
                key = "£50+"
        else:
            key = "All"

        groups[key].append(days)
        group_profits[key].append(profit)

    results = []
    for key, days_list in groups.items():
        results.append({
            "group":      key,
            "avg_days":   round(sum(days_list) / len(days_list), 1),
            "avg_profit": round(sum(group_profits[key]) / len(group_profits[key]), 2),
            "count":      len(days_list),
        })

    results.sort(key=lambda x: x["avg_days"])
    return results


async def get_velocity_stats_async(group_by: str = "set") -> list[dict]:
    async with _lock:
        return await asyncio.to_thread(get_velocity_stats, group_by)


def store_image_urls(item_id: int, urls: list[str]) -> None:
    """Store image URLs for an item (pipe-separated)."""
    wb = _load()
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[_C["Item_ID"]].value == item_id:
            row[_C["Image_URLs"]].value = "|".join(u for u in urls if u)
            _save_with_totals(wb)
            return
    raise ValueError(f"Item {item_id} not found")


async def store_image_urls_async(item_id: int, urls: list[str]) -> None:
    async with _lock:
        await asyncio.to_thread(store_image_urls, item_id, urls)


if __name__ == "__main__":
    import os, tempfile
    config.EXCEL_FILE = os.path.join(tempfile.mkdtemp(), "test_inventory.xlsx")
    init_excel()
    iid = add_item("Test Card", "http://example.com", "ungraded", "KR", 10.0, 12.0)
    wb  = _load()
    ws  = wb.active
    row = _find_row(ws, iid)
    assert row is not None, "_find_row returned None for existing item"
    assert row[_C["Item_ID"]].value    == iid
    assert row[_C["Region"]].value     == "KR", f"Expected Region='KR', got {row[_C['Region']].value}"
    assert row[_C["Potential_Profit"]].value == 2.0, \
        f"Expected Potential_Profit=2.0, got {row[_C['Potential_Profit']].value}"
    assert _find_row(ws, 9999) is None, "_find_row should return None for missing item"

    s = get_summary()
    assert s["total_cost_in_stock"]                == 10.0
    assert s["total_potential_in_stock"]           == 12.0
    assert s["total_potential_profit_in_stock"]    == 2.0, \
        f"Expected total_potential_profit_in_stock=2.0, got {s['total_potential_profit_in_stock']}"
    assert "total_sold_revenue" in s, "get_summary missing total_sold_revenue key"

    import shutil

    # Migration test A: file missing Potential_Profit only
    old_pp_path = os.path.join(tempfile.mkdtemp(), "old_noPP.xlsx")
    shutil.copy(config.EXCEL_FILE, old_pp_path)
    _wb = openpyxl.load_workbook(old_pp_path)
    _ws = _wb.active
    _ws.delete_cols(_C["Potential_Profit"] + 1)
    _wb.save(old_pp_path)
    config.EXCEL_FILE = old_pp_path
    init_excel()
    _mig = _find_row(_load().active, iid)
    assert _mig is not None and _mig[_C["Potential_Profit"]].value == 2.0, \
        "Migration A: Potential_Profit not restored"
    assert _mig[_C["Region"]].value == "KR", "Migration A: Region column shifted"

    # Migration test B: file missing both Region and Potential_Profit (oldest schema)
    old_both_path = os.path.join(tempfile.mkdtemp(), "old_noBoth.xlsx")
    shutil.copy(config.EXCEL_FILE, old_both_path)
    _wb = openpyxl.load_workbook(old_both_path)
    _ws = _wb.active
    # Delete both columns (higher index first to avoid shifting)
    _ws.delete_cols(_C["Potential_Profit"] + 1)
    _ws.delete_cols(_C["Region"] + 1)
    _wb.save(old_both_path)
    config.EXCEL_FILE = old_both_path
    init_excel()
    _mig = _find_row(_load().active, iid)
    assert _mig is not None, "Migration B: row lost"
    # Region defaults to "" when added to a file that never had the column
    assert _mig[_C["Region"]].value     in ("", None), \
        f"Migration B: Region should be empty, got {_mig[_C['Region']].value!r}"
    assert _mig[_C["Potential_Profit"]].value == 2.0, \
        f"Migration B: Potential_Profit wrong: {_mig[_C['Potential_Profit']].value}"
    assert _mig[_C["Card_Name"]].value  == "Test Card", "Migration B: Card_Name shifted"
    assert _mig[_C["Purchase_Price"]].value == 10.0, "Migration B: Purchase_Price shifted"

    print("excel_db smoke test: all assertions passed (including double-migration).")

    # --- No-gaps test: verify add_item never creates phantom empty rows -------
    gaps_path = os.path.join(tempfile.mkdtemp(), "gaps_test.xlsx")
    config.EXCEL_FILE = gaps_path
    init_excel()
    for i in range(5):
        add_item(f"Test {i}", f"http://x.com/{i}", "ungraded", "", 10.0, 15.0)

    _gwb = openpyxl.load_workbook(gaps_path)
    _gws = _gwb.active
    total_row    = None
    last_data_row = None
    for r in range(2, _gws.max_row + 1):
        item_id_val = _gws.cell(r, _C["Item_ID"] + 1).value
        card_val    = _gws.cell(r, _C["Card_Name"] + 1).value
        if isinstance(card_val, str) and card_val.startswith("TOTAL"):
            if total_row is None:
                total_row = r
        elif item_id_val is not None:
            last_data_row = r

    assert last_data_row is not None, "No data rows found"
    assert total_row     is not None, "No TOTAL row found"
    gap = total_row - last_data_row
    assert gap == 2, f"Expected gap of 2 (one blank row), got {gap} (data={last_data_row}, total={total_row})"
    print(f"No-gaps test: OK (last_data_row={last_data_row}, total_row={total_row}, gap=1 empty row)")

    # --- Self-heal test: corrupted file (items with gaps) cleans up on next save ---
    heal_path = os.path.join(tempfile.mkdtemp(), "heal_test.xlsx")
    config.EXCEL_FILE = heal_path
    init_excel()
    add_item("Card A", "http://x.com/a", "ungraded", "", 10.0, 15.0)
    # Manually inject phantom empty rows to simulate the old corruption
    _hwb = openpyxl.load_workbook(heal_path)
    _hws = _hwb.active
    for _ in range(20):
        _hws.append([None] * len(COLUMNS))
    _hwb.save(heal_path)
    assert _hws.max_row > 10, "Corrupt file setup failed"
    # Next add_item should strip the phantom rows via _strip_summary_rows
    add_item("Card B", "http://x.com/b", "ungraded", "", 20.0, 25.0)
    _hwb2 = openpyxl.load_workbook(heal_path)
    _hws2 = _hwb2.active
    # max_row: 1 header + 2 data + 1 blank + 7 totals = 11
    assert _hws2.max_row <= 12, f"Self-heal failed: max_row={_hws2.max_row} (expected ≤12)"
    print(f"Self-heal test: OK (max_row={_hws2.max_row} after healing 20 phantom rows)")
